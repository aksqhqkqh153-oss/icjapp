from __future__ import annotations
import io
import json
import logging
import math
import os
import random
import re
import time
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from fastapi import Depends, FastAPI, File, Header, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .db import (
    DB_ENGINE,
    DB_LABEL,
    get_conn,
    get_user_by_token,
    grade_label,
    hash_password,
    init_db,
    insert_notification,
    json_loads,
    make_token,
    generate_account_unique_id,
    mark_deleted_imported_account,
    row_to_dict,
    user_public_dict,
    utcnow,
    _ensure_columns,
)
from .settings import settings, get_settings
from .storage import StorageError, save_upload
from .settlement_sync import settlement_sync_service, _credential_summary, save_auth_state_json, get_auth_session_guide
from .soomgo_review_api import router as soomgo_review_router
from .warehouse_service import get_state as get_warehouse_state, save_state as save_warehouse_state, update_cell as update_warehouse_cell, update_layout as update_warehouse_layout
from .storage_status_service import get_state as get_storage_status_state, replace_rows as replace_storage_status_rows, save_state as save_storage_status_state

EMAIL_DEMO_MODE = settings.email_demo_mode
logging.basicConfig(level=getattr(logging, settings.log_level, logging.INFO), format='%(asctime)s %(levelname)s %(name)s %(message)s')
logger = logging.getLogger('icj24app')

GEOCODE_CACHE: dict[str, dict[str, Any]] = {}
GEOCODE_CACHE_TTL_SECONDS = 60 * 60 * 24 * 30
KOREA_ADDRESS_FALLBACK_CENTERS: dict[str, dict[str, float]] = {
    '서울 강서구': {'lat': 37.5509, 'lng': 126.8495},
    '서울 양천구': {'lat': 37.5169, 'lng': 126.8666},
    '서울 구로구': {'lat': 37.4954, 'lng': 126.8874},
    '서울 금천구': {'lat': 37.4569, 'lng': 126.8956},
    '서울 영등포구': {'lat': 37.5264, 'lng': 126.8962},
    '서울 동작구': {'lat': 37.5124, 'lng': 126.9393},
    '서울 관악구': {'lat': 37.4782, 'lng': 126.9515},
    '서울 서초구': {'lat': 37.4837, 'lng': 127.0324},
    '서울 강남구': {'lat': 37.5172, 'lng': 127.0473},
    '서울 송파구': {'lat': 37.5145, 'lng': 127.1059},
    '서울 강동구': {'lat': 37.5301, 'lng': 127.1238},
    '서울 마포구': {'lat': 37.5663, 'lng': 126.9019},
    '서울 서대문구': {'lat': 37.5792, 'lng': 126.9368},
    '서울 은평구': {'lat': 37.6176, 'lng': 126.9227},
    '서울 종로구': {'lat': 37.5735, 'lng': 126.9790},
    '서울 중구': {'lat': 37.5636, 'lng': 126.9976},
    '서울 용산구': {'lat': 37.5324, 'lng': 126.9900},
    '서울 성동구': {'lat': 37.5634, 'lng': 127.0369},
    '서울 광진구': {'lat': 37.5384, 'lng': 127.0822},
    '서울 동대문구': {'lat': 37.5744, 'lng': 127.0396},
    '서울 중랑구': {'lat': 37.6066, 'lng': 127.0926},
    '서울 성북구': {'lat': 37.5894, 'lng': 127.0167},
    '서울 강북구': {'lat': 37.6398, 'lng': 127.0257},
    '서울 도봉구': {'lat': 37.6688, 'lng': 127.0471},
    '서울 노원구': {'lat': 37.6542, 'lng': 127.0568},
    '경기 고양시': {'lat': 37.6584, 'lng': 126.8320},
    '경기 파주시': {'lat': 37.7600, 'lng': 126.7802},
    '경기 의정부시': {'lat': 37.7381, 'lng': 127.0338},
    '경기 양주시': {'lat': 37.7853, 'lng': 127.0458},
    '경기 남양주시': {'lat': 37.6360, 'lng': 127.2165},
    '경기 구리시': {'lat': 37.5943, 'lng': 127.1296},
    '경기 하남시': {'lat': 37.5392, 'lng': 127.2149},
    '경기 성남시': {'lat': 37.4200, 'lng': 127.1267},
    '경기 용인시': {'lat': 37.2411, 'lng': 127.1776},
    '경기 수원시': {'lat': 37.2636, 'lng': 127.0286},
    '경기 부천시': {'lat': 37.5034, 'lng': 126.7660},
    '경기 안양시': {'lat': 37.3943, 'lng': 126.9568},
    '경기 광명시': {'lat': 37.4786, 'lng': 126.8646},
    '경기 김포시': {'lat': 37.6152, 'lng': 126.7156},
    '경기 시흥시': {'lat': 37.3803, 'lng': 126.8029},
    '경기 안산시': {'lat': 37.3219, 'lng': 126.8309},
    '인천 부평구': {'lat': 37.5070, 'lng': 126.7219},
    '인천 계양구': {'lat': 37.5371, 'lng': 126.7378},
    '인천 서구': {'lat': 37.5453, 'lng': 126.6758},
    '인천 남동구': {'lat': 37.4473, 'lng': 126.7314},
    '인천 미추홀구': {'lat': 37.4635, 'lng': 126.6506},
}

app = FastAPI(title="이청잘 앱 API", version="1.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.allowed_origins,
    allow_origin_regex=settings.allowed_origin_regex or None,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=1024, compresslevel=6)
app.include_router(soomgo_review_router)

ALLOWED_GENDERS = {'남성', '여성'}


def _origin_allowed(origin: str) -> bool:
    value = str(origin or '').strip()
    if not value:
        return False
    if value in (settings.allowed_origins or []):
        return True
    pattern = str(settings.allowed_origin_regex or '').strip()
    if pattern:
        try:
            return re.match(pattern, value) is not None
        except re.error:
            return False
    return False

def _normalize_administrative_address(value: str) -> str:
    return (
        str(value or '')
        .replace('서울특별시', '서울')
        .replace('부산광역시', '부산')
        .replace('대구광역시', '대구')
        .replace('인천광역시', '인천')
        .replace('광주광역시', '광주')
        .replace('대전광역시', '대전')
        .replace('울산광역시', '울산')
        .replace('세종특별자치시', '세종')
        .replace('경기도', '경기')
        .replace('강원특별자치도', '강원')
        .replace('강원도', '강원')
        .replace('충청북도', '충북')
        .replace('충청남도', '충남')
        .replace('전라북도', '전북')
        .replace('전북특별자치도', '전북')
        .replace('전라남도', '전남')
        .replace('경상북도', '경북')
        .replace('경상남도', '경남')
        .replace('제주특별자치도', '제주')
        .replace('제주도', '제주')
        .replace('특별시', '')
        .replace('광역시', '')
        .replace('특별자치시', '')
        .replace('특별자치도', '')
        .replace('자치시', '')
        .replace('자치도', '')
    ).strip()


def _derive_fallback_geocode(address: str) -> dict[str, Any] | None:
    normalized = re.sub(r'\s+', ' ', _normalize_administrative_address(address)).strip()
    if not normalized:
        return None
    for key, point in KOREA_ADDRESS_FALLBACK_CENTERS.items():
        if normalized.startswith(key):
            return {'lat': point['lat'], 'lng': point['lng'], 'label': normalized, 'cached': False, 'approximate': True}
    tokens = [token for token in normalized.split(' ') if token]
    for size in (3, 2):
        if len(tokens) >= size:
            key = ' '.join(tokens[:size])
            point = KOREA_ADDRESS_FALLBACK_CENTERS.get(key)
            if point:
                return {'lat': point['lat'], 'lng': point['lng'], 'label': normalized, 'cached': False, 'approximate': True}
    return None


def _validate_gender_value(value: str, allow_empty: bool = True) -> str:
    gender = str(value or '').strip()
    if not gender and allow_empty:
        return ''
    normalized_map = {
        '남': '남성',
        '남자': '남성',
        'male': '남성',
        'm': '남성',
        '여': '여성',
        '여자': '여성',
        'female': '여성',
        'f': '여성',
    }
    gender = normalized_map.get(gender.lower(), gender)
    if gender not in ALLOWED_GENDERS:
        raise HTTPException(status_code=400, detail='성별은 남성 또는 여성만 선택할 수 있습니다.')
    return gender
class SignupIn(BaseModel):
    login_id: str
    email: str = ""
    google_email: str = ""
    password: str
    nickname: str
    gender: str = ""
    birth_year: int = 1990
    region: str = "서울"
    phone: str = ""
    recovery_email: str = ""
    vehicle_number: str = ""
    branch_no: Optional[int] = None
class LoginIn(BaseModel):
    login_id: str = ""
    email: str = ""
    password: str
class AccountFindIn(BaseModel):
    nickname: str
    phone: str
    recovery_email: str
class PasswordResetRequestIn(BaseModel):
    recovery_email: str
class PasswordResetConfirmIn(BaseModel):
    recovery_email: str
    code: str
    email: str
    new_password: str
class ProfileIn(BaseModel):
    login_id: str = ""
    email: str = ""
    nickname: str
    position_title: str = ''
    region: str = "서울"
    bio: str = ""
    one_liner: str = ""
    interests: list[str] = []
    photo_url: str = ""
    phone: str = ""
    recovery_email: str = ""
    gender: str = ""
    birth_year: int = 1990
    vehicle_number: str = ""
    branch_no: Optional[int] = None
    marital_status: str = ""
    resident_address: str = ""
    business_name: str = ""
    business_number: str = ""
    business_type: str = ""
    business_item: str = ""
    business_address: str = ""
    bank_account: str = ""
    bank_name: str = ""
    mbti: str = ""
    google_email: str = ""
    resident_id: str = ""
    new_password: str = ""
class LocationIn(BaseModel):
    latitude: float
    longitude: float
    region: str = "서울"
class LocationShareConsentIn(BaseModel):
    enabled: bool

class StorageStatusRowIn(BaseModel):
    id: str = ''
    status: str = ''
    customer_name: str = ''
    manager_name: str = ''
    start_date: str = ''
    end_date: str = ''
    scale: str = ''


class StorageStatusStateIn(BaseModel):
    rows: list[StorageStatusRowIn] = Field(default_factory=list)

class DisposalJurisdictionRowIn(BaseModel):
    id: Optional[int] = None
    category: str = '기본'
    place_prefix: str
    district_name: str
    report_link: str = ''

class DisposalJurisdictionBulkSaveIn(BaseModel):
    rows: list[DisposalJurisdictionRowIn] = Field(default_factory=list)

class DisposalJurisdictionResolveOut(BaseModel):
    matched: bool
    place_prefix: str = ''
    district_name: str = ''
    report_link: str = ''

class FeedPostIn(BaseModel):
    content: str
    image_url: str = ""
class CommentIn(BaseModel):
    content: str
class FriendRespondIn(BaseModel):
    action: str  # accepted / rejected
class GroupRoomIn(BaseModel):
    title: str
    description: str = ""
    region: str = ""
    member_ids: list[int] = Field(default_factory=list)
class MessageIn(BaseModel):
    message: str = ""
    attachment_name: str = ""
    attachment_url: str = ""
    attachment_type: str = ""
    reply_to_id: Optional[int] = None
    mention_user_id: Optional[int] = None
class ChatRoomSettingIn(BaseModel):
    custom_name: str = ""
    pinned: Optional[bool] = None
    favorite: Optional[bool] = None
    muted: Optional[bool] = None
    hidden: Optional[bool] = None
class ChatInviteIn(BaseModel):
    user_id: int
class ReactionIn(BaseModel):
    emoji: str
class VoiceSignalIn(BaseModel):
    payload: dict
class MeetupIn(BaseModel):
    title: str
    place: str
    meetup_date: str
    start_time: str
    end_time: str
    content: str = ""
    cautions: str = ""
    notes: str = ""
class MeetupReviewIn(BaseModel):
    schedule_id: int
    content: str
class BoardPostIn(BaseModel):
    title: str
    content: str
    image_url: str = ""

class SettlementAuthStateIn(BaseModel):
    storage_state: str
    platform: str = '숨고'

class SettlementCredentialIn(BaseModel):
    platform: str = '숨고'
    email: str = ''
    password: str = ''

class SettlementReflectIn(BaseModel):
    settlement_date: str
    category: str = 'daily'
    title: str = ''
    block: dict[str, Any]

class CalendarEventIn(BaseModel):
    title: str
    content: str = ""
    event_date: str
    start_time: str
    end_time: str
    location: str = ""
    color: str = "#2563eb"
    visit_time: str = ""
    move_start_date: str = ""
    move_end_date: str = ""
    move_end_start_time: str = ""
    move_end_end_time: str = ""
    start_address: str = ""
    end_address: str = ""
    platform: str = ""
    customer_name: str = ""
    department_info: str = ""
    schedule_type: str = "A"
    status_a_count: int = 0
    status_b_count: int = 0
    status_c_count: int = 0
    amount1: str = ""
    amount2: str = ""
    amount_item: str = ""
    deposit_method: str = ""
    deposit_amount: str = ""
    representative1: str = ""
    representative2: str = ""
    representative3: str = ""
    staff1: str = ""
    staff2: str = ""
    staff3: str = ""
    image_data: str = ""
    deposit_datetime: str = ""
    reservation_name: str = ""
    reservation_phone: str = ""

class CalendarEventCommentIn(BaseModel):
    content: str = ""
    image_data: str = ""

class CalendarDepartmentReplaceIn(BaseModel):
    from_values: list[str] = []
    to_value: str = '미정'
    to_color: str = '#000000'

class WorkScheduleEntryIn(BaseModel):
    schedule_date: str
    schedule_time: str = ""
    customer_name: str = ""
    representative_names: str = ""
    staff_names: str = ""
    memo: str = ""
class WorkScheduleDayNoteIn(BaseModel):
    schedule_date: str
    excluded_business: str = ""
    excluded_staff: str = ""
    excluded_business_details: list[dict] = []
    excluded_staff_details: list[dict] = []
    available_vehicle_count: int = 0
    status_a_count: int = 0
    status_b_count: int = 0
    status_c_count: int = 0
    day_memo: str = ""
    is_handless_day: bool = False
class WorkdayToggleIn(BaseModel):
    action: str = ''

class HandlessBulkIn(BaseModel):
    month: str = ""
    visible_dates: list[str] = []
    selected_dates: list[str] = []
class AdminModeConfigIn(BaseModel):
    total_vehicle_count: str = ""
    branch_count_override: str = ""
    admin_mode_access_grade: int = 1
    role_assign_actor_max_grade: int = 3
    role_assign_target_min_grade: int = 3
    account_suspend_actor_max_grade: int = 3
    account_suspend_target_min_grade: int = 3
    signup_approve_actor_max_grade: int = 3
    signup_approve_target_min_grade: int = 7
    menu_permissions_json: str = ""
    menu_locks_json: str = ""
class AdminAccountUpdateIn(BaseModel):
    grade: int = 6
    approved: Optional[bool] = None
    position_title: str = ''
    vehicle_available: bool = True
    id: Optional[int] = None
class AdminAccountsBulkUpdateIn(BaseModel):
    accounts: list[AdminAccountUpdateIn] = []
class AdminDeleteAccountsIn(BaseModel):
    ids: list[int] = []

class AdminAccountTypeSwitchIn(BaseModel):
    user_id: int
    target_type: str = 'employee'

class AdminUserDetailIn(BaseModel):
    id: int
    login_id: str = ''
    account_status: str = 'active'
    group_number: str = "0"
    name: str = ''
    nickname: str = ''
    new_password: str = ''
    account_unique_id: str = ''
    position_title: str = ''
    gender: str = ''
    birth_year: int = 1995
    region: str = ''
    phone: str = ''
    recovery_email: str = ''
    vehicle_number: str = ''
    branch_no: Optional[int] = None
    marital_status: str = ''
    resident_address: str = ''
    business_name: str = ''
    business_number: str = ''
    business_type: str = ''
    business_item: str = ''
    business_address: str = ''
    bank_account: str = ''
    bank_name: str = ''
    mbti: str = ''
    email: str = ''
    google_email: str = ''
    resident_id: str = ''
    vehicle_available: bool = True
    show_in_branch_status: bool = False
    show_in_employee_status: bool = False
    show_in_field_employee_status: bool = False
    show_in_hq_status: bool = False
    archived_in_branch_status: bool = False
class AdminUserDetailsBulkIn(BaseModel):
    users: list[AdminUserDetailIn] = []
class AdminCreateAccountIn(BaseModel):
    login_id: str
    email: str = ''
    google_email: str = ''
    account_status: str = 'active'
    password: str
    name: str = ''
    group_number: str = "0"
    position_title: str = ''
    nickname: str
    gender: str = ''
    birth_year: int = 1995
    region: str = '서울'
    phone: str = ''
    recovery_email: str = ''
    vehicle_number: str = ''
    branch_no: Optional[int] = None
    grade: int = 6
    approved: bool = True
    vehicle_available: bool = True

class VehicleExclusionIn(BaseModel):
    start_date: str
    end_date: str
    reason: str = ''

class MaterialPurchaseItemIn(BaseModel):
    product_id: int
    quantity: int = 0
    memo: str = ''

class MaterialPurchaseCreateIn(BaseModel):
    items: list[MaterialPurchaseItemIn] = []
    request_note: str = ''

class MaterialSettlementProcessIn(BaseModel):
    request_ids: list[int] = []

class MaterialRequestDeleteIn(BaseModel):
    request_ids: list[int] = []

class MaterialRequestUpdateRowIn(BaseModel):
    product_id: int
    quantity: int = 0

class MaterialRequestUpdateIn(BaseModel):
    request_ids: list[int] = []
    rows: list[MaterialRequestUpdateRowIn] = []

class MaterialInventoryRowIn(BaseModel):
    product_id: int
    incoming_qty: int = 0
    outgoing_qty: int = 0
    note: str = ''

class MaterialInventorySaveIn(BaseModel):
    rows: list[MaterialInventoryRowIn] = []

class MaterialIncomingSaveIn(BaseModel):
    entry_date: str = ''
    force_apply: bool = False
    rows: list[MaterialInventoryRowIn] = []

class InquiryIn(BaseModel):
    category: str
    title: str
    content: str
class QuoteFormSubmitIn(BaseModel):
    form_type: str = 'same_day'
    requester_name: str = ''
    contact_phone: str = ''
    desired_date: str = ''
    summary_title: str = ''
    privacy_agreed: bool = False
    payload: dict[str, Any] = {}
class ReportIn(BaseModel):
    reason: str
    detail: str = ""
class BlockIn(BaseModel):
    reason: str = ""
class PreferenceIn(BaseModel):
    data: dict


class PasswordVerifyIn(BaseModel):
    password: str


def _ensure_policy_storage_ready(conn: Any) -> bool:
    try:
        conn.execute(
            "CREATE TABLE IF NOT EXISTS policy_contents (policy_key TEXT PRIMARY KEY, content TEXT NOT NULL DEFAULT '', updated_at TEXT NOT NULL DEFAULT '')"
        )
        try:
            _ensure_columns(conn, 'policy_contents', {'updated_at': "TEXT NOT NULL DEFAULT ''"})
        except Exception:
            pass
        return True
    except Exception:
        logger.exception('failed to ensure policy_contents table')
        return False


def _load_policy_contents(conn: Any) -> dict[str, str]:
    contents = dict(POLICY_CONTENT_DEFAULTS)
    storage_ready = _ensure_policy_storage_ready(conn)
    if storage_ready:
        try:
            rows = conn.execute("SELECT policy_key, content FROM policy_contents").fetchall()
            for row in rows:
                normalized = str(row['policy_key'] or '').strip()
                if normalized:
                    contents[normalized] = str(row['content'] or '')
            return contents
        except Exception:
            logger.exception('failed to read policy_contents rows')
    try:
        raw = _get_admin_setting(conn, 'policy_contents_json', '')
        if raw:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                for key, value in parsed.items():
                    normalized = str(key or '').strip()
                    if normalized:
                        contents[normalized] = str(value or '')
    except Exception:
        logger.exception('failed to read fallback policy_contents_json')
    return contents


def _save_policy_content(conn: Any, normalized: str, content: str) -> dict[str, str]:
    now = utcnow()
    storage_ready = _ensure_policy_storage_ready(conn)
    if storage_ready:
        try:
            existing = conn.execute("SELECT policy_key FROM policy_contents WHERE policy_key = ?", (normalized,)).fetchone()
            if existing:
                conn.execute("UPDATE policy_contents SET content = ?, updated_at = ? WHERE policy_key = ?", (content, now, normalized))
            else:
                conn.execute("INSERT INTO policy_contents(policy_key, content, updated_at) VALUES (?, ?, ?)", (normalized, content, now))
            return _load_policy_contents(conn)
        except Exception:
            logger.exception('failed to persist policy_contents row, fallback to admin_settings')
    contents = _load_policy_contents(conn)
    contents[normalized] = content
    conn.execute(
        "INSERT INTO admin_settings(key, value, updated_at) VALUES (?, ?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at",
        ('policy_contents_json', json.dumps(contents, ensure_ascii=False), now),
    )
    return contents

POLICY_CONTENT_DEFAULTS = {
    'vacation:business': '개요\n\n사업자 연차 사용 규정\n\n구분\n분기마다 4일의 연차\n1분기 4일 / 2분기 4일 / 3분기 4일 / 4기 4일\n\n분기구분\n1월 / 2월 / 3월 / 4월 / 5월 / 6월 / 7월 / 8월 / 9월 / 10월 / 11월 / 12월\n\n연간 가능한 총 사용일수\n총 16일\n\n기본신청기준\n- 가능: 2주(14일) 전 미리 신청시 가능\n- 가능: 주말, 공휴일, 손 없는 날 전부 사용 가능\n- 불가: 14일 이내로 신청시 불가\n- 불가: 이미 풀 스케쥴일 경우 불가\n- 예외: 급작스런 경조사 및 특수한 날은 사유에 따라 연차 승인 가능\n\n특별신청기준\n결혼식 / 신혼여행시 기타로 분류\n\n개요\n\n사업자 월차 사용 규정\n\n구분\n월마다 1일의 월차\n1월~12월 각 월 1일\n\n연간 가능한 총 사용일수\n총 12일\n\n기본신청기준\n- 가능: 1주(7일) 전 미리 신청시 가능\n- 불가: 7일 이내로 신청시 불가\n- 불가: 주말, 공휴일, 손 없는 날, 이미 풀 스케쥴일 경우 불가\n- 불가: 월차와 연차를 같은 달에 동시 사용시 불가',
    'vacation:field': '현장직원 휴가 규정을 입력해 주세요.',
    'vacation:office': '본사직원 휴가 규정을 입력해 주세요.',
    'welfare:business': '사업자 복지 규정을 입력해 주세요.',
    'welfare:field': '현장직원 복지 규정을 입력해 주세요.',
    'welfare:office': '본사직원 복지 규정을 입력해 주세요.',
    'schedule:common': '공용 스케줄 규정을 입력해 주세요.',
}

class WarehouseCellUpdateIn(BaseModel):
    sheet_name: str
    row: int
    col: int
    value: Any = ''
class WarehouseLayoutUpdateIn(BaseModel):
    sheet_name: str
    col_widths: dict[str, Any] = {}
    row_heights: dict[str, Any] = {}
class StorageStatusSaveIn(BaseModel):
    rows: list[dict[str, Any]] = Field(default_factory=list)
def _bearer_token(authorization: Optional[str]) -> Optional[str]:
    if not authorization:
        return None
    if authorization.startswith("Bearer "):
        return authorization[7:]
    return authorization
GRADE_LABELS = {1: '관리자', 2: '부관리자', 3: '중간관리자', 4: '사업자', 5: '직원', 6: '일반', 7: '기타'}
def _get_admin_setting(conn, key: str, default: str = '') -> str:
    try:
        row = conn.execute("SELECT value FROM admin_settings WHERE key = ?", (key,)).fetchone()
    except Exception:
        return default
    if not row or row['value'] in (None, ''):
        return default
    return str(row['value'])
def _get_permission_config(conn) -> dict:
    return {
        'admin_mode_access_grade': int(_get_admin_setting(conn, 'admin_mode_access_grade', '2') or 2),
        'role_assign_actor_max_grade': int(_get_admin_setting(conn, 'role_assign_actor_max_grade', '3') or 3),
        'role_assign_target_min_grade': int(_get_admin_setting(conn, 'role_assign_target_min_grade', '3') or 3),
        'account_suspend_actor_max_grade': int(_get_admin_setting(conn, 'account_suspend_actor_max_grade', '3') or 3),
        'account_suspend_target_min_grade': int(_get_admin_setting(conn, 'account_suspend_target_min_grade', '3') or 3),
        'signup_approve_actor_max_grade': int(_get_admin_setting(conn, 'signup_approve_actor_max_grade', '3') or 3),
        'signup_approve_target_min_grade': int(_get_admin_setting(conn, 'signup_approve_target_min_grade', '7') or 7),
        'menu_permissions_json': _get_admin_setting(conn, 'menu_permissions_json', ''),
        'menu_locks_json': _get_admin_setting(conn, 'menu_locks_json', ''),
    }
def _get_admin_total_vehicle_count(conn) -> int:
    row = conn.execute("SELECT COUNT(*) FROM users WHERE COALESCE(vehicle_available, 1) = 1 AND COALESCE(approved, 1) = 1 AND CAST(COALESCE(grade, '6') AS INTEGER) <= 4").fetchone()
    return int(row[0] or 0)
def _get_branch_count_override(conn) -> int:
    raw = _get_admin_setting(conn, 'branch_count_override', '')
    if raw.isdigit():
        return int(raw)
    row = conn.execute("SELECT COUNT(*) FROM users WHERE COALESCE(approved, 1) = 1 AND COALESCE(archived_in_branch_status, 0) = 0 AND (CAST(COALESCE(grade, '6') AS INTEGER) = 4 OR COALESCE(show_in_branch_status, 0) = 1 OR branch_no IS NOT NULL)").fetchone()
    return int(row[0] or 0)
def _grade_of(user: dict) -> int:
    return int(user.get('grade') or 6)

def _is_staff_grade(grade_value) -> bool:
    try:
        return int(grade_value or 0) == 5
    except Exception:
        return False

def _is_employee_restricted_user(user: dict) -> bool:
    account_type = str(user.get('account_type') or '').strip().lower()
    if account_type == 'employee':
        return True
    return _is_staff_grade(user.get('grade'))

def _can_access_admin_mode(user: dict, conn) -> bool:
    return _grade_of(user) <= 2 or _grade_of(user) <= _get_permission_config(conn)['admin_mode_access_grade']
def _can_manage_grade(actor: dict, target_grade: int, conn) -> bool:
    actor_grade = _grade_of(actor)
    cfg = _get_permission_config(conn)
    return actor_grade <= cfg['role_assign_actor_max_grade'] and target_grade >= cfg['role_assign_target_min_grade'] and actor_grade < target_grade
def _can_actor_apply(actor: dict, actor_key: str, target_key: str, target_grade: int, conn) -> bool:
    actor_grade = _grade_of(actor)
    cfg = _get_permission_config(conn)
    return actor_grade <= int(cfg.get(actor_key, 1)) and target_grade >= int(cfg.get(target_key, 7)) and actor_grade < target_grade


def _can_edit_calendar_event(user: dict) -> bool:
    grade = _grade_of(user)
    position_title = str(user.get('position_title') or '').strip()
    allowed_titles = {'대표', '부대표', '호점대표', '팀장', '부팀장', '본부장', '상담실장', '상담팀장', '상담사원'}
    return grade <= 2 or position_title in allowed_titles

def _normalize_account_type(row: Any) -> str:
    data = row_to_dict(row)
    grade = _grade_of(data)
    position_title = str(data.get('position_title') or '').strip()
    if grade <= 3:
        return 'admin'
    if position_title in {'대표', '부대표', '호점대표'}:
        return 'business'
    if position_title in {'팀장', '부팀장', '직원'}:
        return 'employee_field'
    if position_title in {'본부장', '상담실장', '상담팀장', '상담사원'}:
        return 'employee_hq'
    if grade == 4:
        return 'business'
    if grade == 5:
        return 'employee_field'
    return 'general'


def _normalize_email_value(value: Any) -> str:
    return str(value or '').strip().lower()

def _normalize_login_id_value(value: Any) -> str:
    return ''.join(ch for ch in str(value or '').strip().lower() if not ch.isspace())

def _validate_login_id_value(value: Any) -> str:
    login_id = _normalize_login_id_value(value)
    if not login_id:
        raise HTTPException(status_code=400, detail='아이디를 입력해 주세요.')
    if len(login_id) > 30:
        raise HTTPException(status_code=400, detail='아이디는 30자 이하로 입력해 주세요.')
    if not re.fullmatch(r'[^\W_]+', login_id, re.UNICODE):
        raise HTTPException(status_code=400, detail='아이디는 특수문자, -, _ 없이 입력해 주세요.')
    return login_id

def _normalize_account_status_value(value: Any, approved: Any = None, grade: Any = None) -> str:
    status = str(value or '').strip().lower()
    aliases = {'승인대기': 'pending', '사용중': 'active', '일시정지': 'suspended', '퇴사/종료': 'retired', '계정삭제': 'deleted'}
    status = aliases.get(status, status or '')
    if status in {'pending','active','suspended','retired','deleted'}:
        return status
    try:
        if int(approved if approved is not None else 1) == 0 or int(grade or 6) == 7:
            return 'pending'
    except Exception:
        pass
    return 'active'

def _find_user_by_login_id_ci(conn, login_id: str, exclude_user_id: int | None = None):
    normalized = _normalize_login_id_value(login_id)
    if not normalized:
        return None
    sql = "SELECT * FROM users WHERE LOWER(TRIM(COALESCE(login_id, ''))) = ?"
    params = [normalized]
    if exclude_user_id is not None:
        sql += " AND id != ?"
        params.append(int(exclude_user_id))
    sql += " ORDER BY id LIMIT 1"
    row = conn.execute(sql, tuple(params)).fetchone()
    if row:
        return row
    legacy_sql = "SELECT * FROM users WHERE LOWER(TRIM(COALESCE(email, ''))) = ?"
    legacy_params = [normalized]
    if exclude_user_id is not None:
        legacy_sql += " AND id != ?"
        legacy_params.append(int(exclude_user_id))
    legacy_sql += " ORDER BY id LIMIT 1"
    return conn.execute(legacy_sql, tuple(legacy_params)).fetchone()


def _find_user_by_email_ci(conn, email: str, exclude_user_id: int | None = None):
    normalized = _normalize_email_value(email)
    if not normalized:
        return None
    if exclude_user_id is None:
        return conn.execute("SELECT * FROM users WHERE LOWER(TRIM(email)) = ? ORDER BY id LIMIT 1", (normalized,)).fetchone()
    return conn.execute("SELECT * FROM users WHERE LOWER(TRIM(email)) = ? AND id != ? ORDER BY id LIMIT 1", (normalized, int(exclude_user_id))).fetchone()


def _canonical_role_for_user(grade_value: Any, branch_no: Any, position_title: Any, current_role: Any = '') -> str:
    try:
        grade = int(grade_value or 6)
    except Exception:
        grade = 6
    account_type = _normalize_account_type({'grade': grade, 'branch_no': branch_no, 'position_title': position_title, 'role': current_role})
    if grade <= 3:
        return 'admin'
    if account_type == 'business':
        return 'business'
    if account_type in {'employee_field', 'employee_hq'}:
        return 'employee'
    return 'user'


def _normalize_account_admin_flags(data: dict) -> dict:
    normalized = dict(data)
    try:
        grade = int(normalized.get('grade') or 6)
    except Exception:
        grade = 6
    branch_no = normalized.get('branch_no')
    position_title = str(normalized.get('position_title') or '').strip()
    account_type = _normalize_account_type({'grade': grade, 'branch_no': branch_no, 'position_title': position_title, 'role': normalized.get('role')})
    if account_type == 'business' and branch_no in (None, ''):
        normalized['branch_no'] = -1
        branch_no = -1
    normalized['show_in_branch_status'] = 1 if account_type == 'business' else 0
    normalized['account_type'] = account_type
    normalized['show_in_employee_status'] = 1 if account_type in {'employee_field', 'employee_hq'} else 0
    normalized['show_in_field_employee_status'] = 1 if account_type == 'employee_field' else 0
    normalized['show_in_hq_status'] = 1 if account_type == 'employee_hq' else 0
    normalized['role'] = _canonical_role_for_user(grade, branch_no, position_title, normalized.get('role') or 'user')
    normalized['vehicle_available'] = 1 if account_type == 'business' else 0
    if not normalized.get('position_title') and account_type == 'business':
        normalized['position_title'] = '호점대표'
    normalized['branch_code'] = 'TEMP_BRANCH' if normalized.get('branch_no') == -1 else (f"BRANCH_{int(normalized.get('branch_no') or 0)}" if normalized.get('branch_no') not in (None, '', -1) and int(normalized.get('branch_no') or 0) > 0 else '')
    normalized['permission_codes_json'] = normalized.get('permission_codes_json') or '[]'
    normalized['account_status'] = _normalize_account_status_value(normalized.get('account_status'), normalized.get('approved'), grade)
    return normalized


def _serialize_admin_user_row(row: Any) -> dict[str, Any]:
    item = row_to_dict(row)
    item['group_number'] = str((item.get('group_number_text') if item.get('group_number_text') not in (None, '') else item.get('group_number')) or '0')
    item['grade_label'] = grade_label(item.get('grade'))
    item['approved'] = bool(item.get('approved', 1))
    item['vehicle_available'] = False if _is_staff_grade(item.get('grade')) else bool(item.get('vehicle_available', 1))
    item['account_type'] = _normalize_account_type(item)
    item['branch_code'] = str(item.get('branch_code') or ('TEMP_BRANCH' if item.get('branch_no') == -1 else (f"BRANCH_{int(item.get('branch_no') or 0)}" if item.get('branch_no') not in (None, '') and int(item.get('branch_no') or 0) > 0 else '')))
    item['permission_codes_json'] = json_loads(item.get('permission_codes_json'), []) if isinstance(item.get('permission_codes_json'), str) else (item.get('permission_codes_json') or [])
    branch_flag = item.get('show_in_branch_status')
    employee_flag = item.get('show_in_employee_status')
    field_employee_flag = item.get('show_in_field_employee_status')
    hq_flag = item.get('show_in_hq_status')
    inferred_employee = bool(employee_flag) if employee_flag is not None else item['account_type'] == 'employee'
    inferred_hq = bool(hq_flag) if hq_flag is not None else _is_head_office_staff(item)
    inferred_field = bool(field_employee_flag) if field_employee_flag is not None else (inferred_employee and not inferred_hq)
    item['show_in_branch_status'] = bool(branch_flag) if branch_flag is not None else item['account_type'] == 'business'
    item['show_in_employee_status'] = inferred_employee
    item['show_in_field_employee_status'] = inferred_field
    item['show_in_hq_status'] = inferred_hq
    item['archived_in_branch_status'] = bool(item.get('archived_in_branch_status', 0))
    return item



def _is_head_office_staff(item: dict) -> bool:
    email = str(item.get('email') or '').strip()
    name = str(item.get('name') or '').strip()
    nickname = str(item.get('nickname') or '').strip()
    head_office_emails = {'이청잘A', '이청잘B', '이청잘C'}
    head_office_names = {'최성규', '이준희', '손지민'}
    return email in head_office_emails or name in head_office_names or nickname in head_office_names

def _split_names_for_match(*values: str) -> list[str]:
    tokens: list[str] = []
    for value in values:
        for token in re.split(r'[\n,/|]+', str(value or '')):
            cleaned = token.strip()
            if not cleaned:
                continue
            tokens.append(cleaned)
            bracket_parts = re.findall(r'\[([^\]]+)\]', cleaned)
            for part in bracket_parts:
                normalized = str(part or '').strip()
                if normalized:
                    tokens.append(normalized)
            compact = re.sub(r'\s+', '', cleaned)
            if compact and compact != cleaned:
                tokens.append(compact)
    unique: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        normalized = str(token or '').strip()
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique.append(normalized)
    return unique


def _user_assignment_tokens(user: dict) -> set[str]:
    tokens = set()
    display_name = str(user.get('name') or user.get('nickname') or user.get('email') or '').strip()
    for raw in [user.get('name'), user.get('nickname'), user.get('email'), user.get('vehicle_number')]:
        value = str(raw or '').strip()
        if value:
            tokens.add(value)
    branch_no = user.get('branch_no')
    branch_label = ''
    if branch_no not in (None, ''):
        branch_label = f"{branch_no}호점"
        tokens.add(branch_label)
        tokens.add(str(branch_no))
    if branch_label and display_name:
        tokens.add(f'[{branch_label}][{display_name}]')
        tokens.add(f'{branch_label}{display_name}')
    return {token for token in tokens if str(token or '').strip()}


def _row_assigned_to_user(user: dict, row: dict) -> bool:
    tokens = _user_assignment_tokens(user)
    if not tokens:
        return False
    row_values = []
    for key in ['representative1', 'representative2', 'representative3', 'staff1', 'staff2', 'staff3', 'representative_names', 'staff_names']:
        row_values.extend(_split_names_for_match(row.get(key, '')))
    row_set = {str(item).strip() for item in row_values if str(item).strip()}
    if row_set & tokens:
        return True
    merged = ' '.join(row_set)
    return any(token and token in merged for token in tokens)


def _parse_time_value(value: str | None):
    text = str(value or '').strip()
    if not text or text == '미정':
        return None
    match = re.match(r'^(\d{1,2}):(\d{2})$', text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour == 24 and minute == 0:
        hour = 23
        minute = 59
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return datetime.strptime(f"{hour:02d}:{minute:02d}", '%H:%M').time()


def _schedule_time_window(target_date: date, start_raw: str | None, end_raw: str | None):
    start_t = _parse_time_value(start_raw)
    end_t = _parse_time_value(end_raw)
    if start_t is None and end_t is None:
        return None, None
    base_start_dt = datetime.combine(target_date, start_t or datetime.strptime('00:00', '%H:%M').time())
    if end_t is None:
        base_end_dt = base_start_dt + timedelta(hours=2)
    else:
        base_end_dt = datetime.combine(target_date, end_t)
        if base_end_dt < base_start_dt:
            base_end_dt = base_start_dt + timedelta(hours=2)
    return base_start_dt - timedelta(hours=1), base_end_dt + timedelta(minutes=30)


def _assignment_display_text(current_user: dict, value: str) -> str:
    tokens = _user_assignment_tokens(current_user)
    names = _split_names_for_match(value)
    out = []
    for name in names:
        out.append(f"{name}(본인)" if name in tokens else name)
    return ' / '.join(out) if out else '-'

def _work_assignment_target_ids(conn, representative_names: str, staff_names: str, actor_id: int | None = None) -> list[int]:
    target_tokens = {token for token in _split_names_for_match(representative_names, staff_names) if token}
    if not target_tokens:
        return []
    rows = conn.execute("SELECT id, email, nickname, vehicle_number, branch_no FROM users ORDER BY id").fetchall()
    matched_ids: list[int] = []
    for row in rows:
        data = row_to_dict(row)
        user_id = int(data.get('id') or 0)
        tokens = _user_assignment_tokens(data)
        if target_tokens & tokens or any(token and token in ' '.join(tokens) for token in target_tokens):
            if actor_id is not None and user_id == int(actor_id):
                matched_ids.append(user_id)
            elif user_id not in matched_ids:
                matched_ids.append(user_id)
    return matched_ids


def _notify_work_schedule_assignments(conn, actor: dict, schedule_date: str, schedule_time: str, customer_name: str, representative_names: str, staff_names: str, previous_ids: set[int] | None = None):
    previous_ids = previous_ids or set()
    target_ids = set(_work_assignment_target_ids(conn, representative_names, staff_names, actor.get('id')))
    if not target_ids:
        return
    date_text = ''
    time_text = str(schedule_time or '미정').strip() or '미정'
    try:
        dt = datetime.strptime(str(schedule_date), '%Y-%m-%d')
        date_text = f"{dt.month:02d}월 {dt.day:02d}일"
    except Exception:
        date_text = str(schedule_date or '')
    reps = ' / '.join(_split_names_for_match(representative_names)) or '-'
    staffs = ' / '.join(_split_names_for_match(staff_names)) or '-'
    customer = str(customer_name or '(고객명)').strip() or '(고객명)'
    body = f"{date_text} {time_text}에 {customer}고객님 일정으로 {reps} 사업자와 {staffs} 직원이 배정되었습니다. 일정 확인을 해주세요"
    for user_id in target_ids:
        if user_id in previous_ids:
            continue
        insert_notification(conn, user_id, 'work_schedule_assignment', '스케줄 배정', body)




def _format_notice_date(date_value: str) -> str:
    try:
        dt = datetime.strptime(str(date_value), '%Y-%m-%d')
        return f"{dt.month}월 {dt.day}일"
    except Exception:
        return str(date_value or '')


def _join_assignment_names(*values: str) -> str:
    names: list[str] = []
    for value in values:
        for token in _split_names_for_match(value):
            if token and token not in names:
                names.append(token)
    return ' / '.join(names) if names else '-'


def _notify_schedule_change(conn, target_ids: set[int], type_: str, title: str, body: str, actor_id: int | None = None):
    for user_id in sorted({int(item) for item in target_ids if int(item or 0) > 0}):
        if actor_id is not None and user_id == int(actor_id):
            continue
        insert_notification(conn, user_id, type_, title, body)


def _calendar_assignment_names(row: dict) -> tuple[str, str]:
    reps = _join_assignment_names(row.get('representative1'), row.get('representative2'), row.get('representative3'))
    staffs = _join_assignment_names(row.get('staff1'), row.get('staff2'), row.get('staff3'))
    return reps, staffs


def _schedule_assignment_notice_payload(date_value: str, time_value: str, customer_name: str, before_text: str, after_text: str) -> tuple[str, str]:
    date_text = _format_notice_date(date_value)
    customer = str(customer_name or '(고객명)').strip() or '(고객명)'
    added_name = '-'
    removed_name = '-'
    after_match = re.search(r'([가-힣A-Za-z0-9_]+)\s*(대표|부대표|호점대표|팀장|부팀장|본부장|상담실장|상담팀장|상담사원|직원)', str(after_text or ''))
    before_match = re.search(r'([가-힣A-Za-z0-9_]+)\s*(대표|부대표|호점대표|팀장|부팀장|본부장|상담실장|상담팀장|상담사원|직원)', str(before_text or ''))
    if after_match:
        added_name = f"{after_match.group(1)} {after_match.group(2)}"
    if before_match:
        removed_name = f"{before_match.group(1)} {before_match.group(2)}"
    title = '스케줄 변경'
    body = f"{date_text} {customer} 고객 | {removed_name} 삭제\n{date_text} {customer} 고객 | {added_name} 배정"
    return title, body


def _schedule_assignment_membership_notice_payload(date_value: str, customer_name: str, action: str, assignee_name: str = '') -> tuple[str, str]:
    date_text = _format_notice_date(date_value)
    customer = str(customer_name or '(고객명)').strip() or '(고객명)'
    assignee = str(assignee_name or '담당자').strip() or '담당자'
    if action == 'removed':
        return '스케줄 변경', f"{date_text} {customer} 고객 | {assignee} 삭제"
    return '스케줄 변경', f"{date_text} {customer} 고객 | {assignee} 배정"


def _schedule_time_notice_payload(date_value: str, before_time: str, after_time: str, customer_name: str, representative_names: str, staff_names: str) -> tuple[str, str]:
    date_text = _format_notice_date(date_value)
    customer = str(customer_name or '(고객명)').strip() or '(고객명)'
    before_text = str(before_time or '미정').strip() or '미정'
    after_text = str(after_time or '미정').strip() or '미정'
    reps = representative_names or '-'
    staffs = staff_names or '-'
    title = '이사시간 변경'
    body = f"{date_text} {after_text} {customer} 고객님의 이사시간 변경([{before_text}] → [{after_text}])되었습니다.\n* 투입되는 인원[{reps}][{staffs}]은 {after_text}에 맞춰 출발지로 방문드려주세요."
    return title, body


def _schedule_address_notice_payload(date_value: str, time_value: str, before_address: str, after_address: str, customer_name: str, representative_names: str, staff_names: str) -> tuple[str, str]:
    date_text = _format_notice_date(date_value)
    time_text = str(time_value or '미정').strip() or '미정'
    customer = str(customer_name or '(고객명)').strip() or '(고객명)'
    before_text = str(before_address or '-').strip() or '-'
    after_text = str(after_address or '-').strip() or '-'
    reps = representative_names or '-'
    staffs = staff_names or '-'
    title = '출발지 주소변경'
    body = f"{date_text} {time_text} {customer} 고객님의 출발지 주소변경([{before_text}] → [{after_text}])되었습니다.\n* 투입되는 인원[{reps}][{staffs}]은 {after_text}에 맞춰 출발지로 방문드려주세요."
    return title, body


def _notify_work_schedule_entry_changes(conn, actor: dict, previous_row: dict, next_row: dict):
    previous_ids = set(_work_assignment_target_ids(conn, previous_row.get('representative_names') or '', previous_row.get('staff_names') or '', None))
    next_ids = set(_work_assignment_target_ids(conn, next_row.get('representative_names') or '', next_row.get('staff_names') or '', None))
    target_ids = previous_ids | next_ids
    if not target_ids:
        return
    previous_assignment = f"대표 {previous_row.get('representative_names') or '-'} / 직원 {previous_row.get('staff_names') or '-'}"
    next_assignment = f"대표 {next_row.get('representative_names') or '-'} / 직원 {next_row.get('staff_names') or '-'}"
    if previous_assignment != next_assignment:
        event_date = next_row.get('schedule_date') or previous_row.get('schedule_date') or ''
        customer_name = next_row.get('customer_name') or previous_row.get('customer_name') or ''
        added_ids = next_ids - previous_ids
        removed_ids = previous_ids - next_ids
        stayed_ids = previous_ids & next_ids
        if stayed_ids:
            title, body = _schedule_assignment_notice_payload(event_date, next_row.get('schedule_time') or previous_row.get('schedule_time') or '', customer_name, previous_assignment, next_assignment)
            _notify_schedule_change(conn, stayed_ids, 'work_schedule_assignment_change', title, body, actor.get('id'))
        if added_ids:
            added_name = _join_assignment_names(next_row.get('representative_names') or '', next_row.get('staff_names') or '')
            title, body = _schedule_assignment_membership_notice_payload(event_date, customer_name, 'added', added_name)
            _notify_schedule_change(conn, added_ids, 'work_schedule_assignment_added', title, body, actor.get('id'))
        if removed_ids:
            removed_name = _join_assignment_names(previous_row.get('representative_names') or '', previous_row.get('staff_names') or '')
            title, body = _schedule_assignment_membership_notice_payload(event_date, customer_name, 'removed', removed_name)
            _notify_schedule_change(conn, removed_ids, 'work_schedule_assignment_removed', title, body, actor.get('id'))
    if (previous_row.get('schedule_time') or '') != (next_row.get('schedule_time') or ''):
        title, body = _schedule_time_notice_payload(next_row.get('schedule_date') or previous_row.get('schedule_date') or '', previous_row.get('schedule_time') or '', next_row.get('schedule_time') or '', next_row.get('customer_name') or previous_row.get('customer_name') or '', next_row.get('representative_names') or '', next_row.get('staff_names') or '')
        _notify_schedule_change(conn, target_ids, 'work_schedule_time_change', title, body, actor.get('id'))
    if (previous_row.get('start_address') or '') != (next_row.get('start_address') or '') and ((previous_row.get('start_address') or '') or (next_row.get('start_address') or '')):
        title, body = _schedule_address_notice_payload(next_row.get('schedule_date') or previous_row.get('schedule_date') or '', next_row.get('schedule_time') or previous_row.get('schedule_time') or '', previous_row.get('start_address') or '', next_row.get('start_address') or '', next_row.get('customer_name') or previous_row.get('customer_name') or '', next_row.get('representative_names') or '', next_row.get('staff_names') or '')
        _notify_schedule_change(conn, target_ids, 'work_schedule_address_change', title, body, actor.get('id'))


def _notify_calendar_event_changes(conn, actor: dict, previous_row: dict, next_row: dict):
    prev_reps, prev_staffs = _calendar_assignment_names(previous_row)
    next_reps, next_staffs = _calendar_assignment_names(next_row)
    previous_ids = set(_work_assignment_target_ids(conn, prev_reps, prev_staffs, None))
    next_ids = set(_work_assignment_target_ids(conn, next_reps, next_staffs, None))
    target_ids = previous_ids | next_ids
    if not target_ids:
        return
    previous_assignment = f"대표 {prev_reps} / 직원 {prev_staffs}"
    next_assignment = f"대표 {next_reps} / 직원 {next_staffs}"
    event_date = next_row.get('event_date') or previous_row.get('event_date') or ''
    event_time = next_row.get('start_time') or next_row.get('visit_time') or previous_row.get('start_time') or previous_row.get('visit_time') or ''
    customer_name = next_row.get('customer_name') or previous_row.get('customer_name') or next_row.get('title') or previous_row.get('title') or ''
    if previous_assignment != next_assignment:
        added_ids = next_ids - previous_ids
        removed_ids = previous_ids - next_ids
        stayed_ids = previous_ids & next_ids
        if stayed_ids:
            title, body = _schedule_assignment_notice_payload(event_date, event_time, customer_name, previous_assignment, next_assignment)
            _notify_schedule_change(conn, stayed_ids, 'calendar_assignment_change', title, body, actor.get('id'))
        if added_ids:
            added_name = _join_assignment_names(next_reps, next_staffs)
            title, body = _schedule_assignment_membership_notice_payload(event_date, customer_name, 'added', added_name)
            _notify_schedule_change(conn, added_ids, 'calendar_assignment_added', title, body, actor.get('id'))
        if removed_ids:
            removed_name = _join_assignment_names(prev_reps, prev_staffs)
            title, body = _schedule_assignment_membership_notice_payload(event_date, customer_name, 'removed', removed_name)
            _notify_schedule_change(conn, removed_ids, 'calendar_assignment_removed', title, body, actor.get('id'))
    prev_time = previous_row.get('start_time') or previous_row.get('visit_time') or ''
    next_time = next_row.get('start_time') or next_row.get('visit_time') or ''
    if prev_time != next_time:
        title, body = _schedule_time_notice_payload(event_date, prev_time, next_time, customer_name, next_reps, next_staffs)
        _notify_schedule_change(conn, target_ids, 'calendar_time_change', title, body, actor.get('id'))
    prev_address = previous_row.get('start_address') or previous_row.get('location') or ''
    next_address = next_row.get('start_address') or next_row.get('location') or ''
    if prev_address != next_address and (prev_address or next_address):
        title, body = _schedule_address_notice_payload(event_date, event_time, prev_address, next_address, customer_name, next_reps, next_staffs)
        _notify_schedule_change(conn, target_ids, 'calendar_address_change', title, body, actor.get('id'))
def _calendar_row_summary(user: dict, row: dict) -> dict:
    rep_list = [str(row.get(key) or '').strip() for key in ['representative1', 'representative2', 'representative3'] if str(row.get(key) or '').strip()]
    staff_list = [str(row.get(key) or '').strip() for key in ['staff1', 'staff2', 'staff3'] if str(row.get(key) or '').strip()]
    target_date = datetime.strptime(row.get('event_date'), '%Y-%m-%d').date()
    return {
        'source': 'calendar',
        'id': row.get('id'),
        'schedule_date': row.get('event_date') or '',
        'time_text': row.get('start_time') or row.get('visit_time') or '미정',
        'customer_name': row.get('customer_name') or row.get('title') or '',
        'representative_text': _assignment_display_text(user, ' / '.join(rep_list)),
        'staff_text': _assignment_display_text(user, ' / '.join(staff_list)),
        'start_address': row.get('start_address') or row.get('location') or '',
        'end_address': row.get('end_address') or '',
        'window_start': _schedule_time_window(target_date, row.get('start_time') or row.get('visit_time'), row.get('end_time')),
    }


def _manual_row_summary(user: dict, row: dict) -> dict:
    date_value = datetime.strptime(row.get('schedule_date'), '%Y-%m-%d').date()
    return {
        'source': 'manual',
        'id': row.get('id'),
        'schedule_date': row.get('schedule_date') or '',
        'time_text': row.get('schedule_time') or '미정',
        'customer_name': row.get('customer_name') or '',
        'representative_text': _assignment_display_text(user, row.get('representative_names') or ''),
        'staff_text': _assignment_display_text(user, row.get('staff_names') or ''),
        'start_address': row.get('start_address') or row.get('location') or '',
        'end_address': row.get('end_address') or '',
        'window_start': _schedule_time_window(date_value, row.get('schedule_time'), None),
    }


def _assigned_schedule_items(conn, user: dict, start_date: date, end_date: date) -> list[dict]:
    start_key = start_date.isoformat()
    end_key = end_date.isoformat()
    items: list[dict] = []
    calendar_rows = conn.execute(
        """
        SELECT * FROM calendar_events
        WHERE event_date >= ? AND event_date <= ?
        ORDER BY event_date, CASE WHEN COALESCE(start_time, '') IN ('', '미정') THEN '99:99' ELSE start_time END, id
        """,
        (start_key, end_key),
    ).fetchall()
    for row in calendar_rows:
        data = row_to_dict(row)
        if not _row_assigned_to_user(user, data):
            continue
        items.append(_calendar_row_summary(user, data))
    manual_rows = conn.execute(
        """
        SELECT w.*, c.start_address AS linked_start_address, c.location AS linked_location, c.end_address AS linked_end_address
        FROM work_schedule_entries w
        LEFT JOIN calendar_events c
          ON c.event_date = w.schedule_date
         AND COALESCE(c.customer_name, '') = COALESCE(w.customer_name, '')
        WHERE w.schedule_date >= ? AND w.schedule_date <= ?
        ORDER BY w.schedule_date, CASE WHEN COALESCE(w.schedule_time, '') = '' THEN '99:99' ELSE w.schedule_time END, w.id
        """,
        (start_key, end_key),
    ).fetchall()
    for row in manual_rows:
        data = row_to_dict(row)
        if not _row_assigned_to_user(user, data):
            continue
        data['start_address'] = data.get('linked_start_address') or data.get('linked_location') or ''
        data['end_address'] = data.get('linked_end_address') or ''
        items.append(_manual_row_summary(user, data))
    items.sort(key=lambda item: (item['schedule_date'], '99:99' if item['time_text'] in ('', '미정') else item['time_text'], str(item['id'])))
    return items


def _location_share_status(conn, user: dict) -> dict:
    eligible = bool(str(user.get('vehicle_number') or '').strip()) and user.get('branch_no') not in (None, '')
    today = datetime.now().date()
    now_dt = datetime.now()
    assigned_today = [item for item in _assigned_schedule_items(conn, user, today, today) if item['schedule_date'] == today.isoformat()]
    active_item = None
    for item in assigned_today:
        start_dt, end_dt = item.get('window_start') or (None, None)
        if start_dt and end_dt and start_dt <= now_dt <= end_dt:
            active_item = item
            break
    return {
        'eligible': eligible,
        'consent_granted': bool(user.get('location_share_consent')),
        'sharing_enabled': bool(user.get('location_share_enabled')),
        'active_now': active_item is not None,
        'today_assignments': [
            {
                'schedule_date': item['schedule_date'],
                'time_text': item['time_text'],
                'customer_name': item['customer_name'],
                'start_address': item['start_address'],
            } for item in assigned_today
        ],
        'active_assignment': {
            'schedule_date': active_item['schedule_date'],
            'time_text': active_item['time_text'],
            'customer_name': active_item['customer_name'],
            'start_address': active_item['start_address'],
        } if active_item else None,
    }



def _materials_scope_allowed(user: dict, scope: str) -> bool:
    grade = _grade_of(user)
    employee_restricted = _is_employee_restricted_user(user)
    if scope == 'sales':
        return True
    if scope == 'my_requests':
        return not employee_restricted
    if scope == 'inventory':
        return grade <= 2
    if scope in {'requesters', 'settlements', 'history'}:
        return grade <= 2
    if scope == 'inventory_manage':
        return grade <= 2
    return False

def _require_materials_scope(user: dict, scope: str):
    if not _materials_scope_allowed(user, scope):
        raise HTTPException(status_code=403, detail='현재 권한으로는 자재 기능에 접근할 수 없습니다.')

def _material_alias_map(name: str, short_name: str) -> list[str]:
    aliases = {str(name or '').strip(), str(short_name or '').strip()}
    if str(short_name or '').strip() == '노비':
        aliases.add('노란비닐')
    if str(short_name or '').strip() == '흰비':
        aliases.add('흰색비닐')
    if str(short_name or '').strip() == '침비':
        aliases.add('침대비닐')
    if str(short_name or '').strip() == '스티커':
        aliases.add('스티커 인쇄물')
    if str(short_name or '').strip() == '테이프':
        aliases.add('이사테이프')
    return [item for item in aliases if item]

def _material_products(conn):
    return [row_to_dict(row) for row in conn.execute(
        "SELECT * FROM material_products WHERE COALESCE(is_active, 1) = 1 ORDER BY display_order, id"
    ).fetchall()]

def _material_permissions(user: dict) -> dict:
    return {
        'can_view_sales': _materials_scope_allowed(user, 'sales'),
        'can_view_inventory': _materials_scope_allowed(user, 'inventory'),
        'can_view_requesters': _materials_scope_allowed(user, 'requesters'),
        'can_view_settlements': _materials_scope_allowed(user, 'settlements'),
        'can_view_history': _materials_scope_allowed(user, 'history'),
        'can_manage_inventory': _materials_scope_allowed(user, 'inventory_manage'),
        'can_manage_incoming': _materials_scope_allowed(user, 'inventory_manage'),
        'can_view_my_requests': _materials_scope_allowed(user, 'my_requests'),
    }


def _material_branch_label_from_user(user_row: dict | None) -> str:
    data = row_to_dict(user_row)
    branch_no_raw = data.get('branch_no')
    branch_code = str(data.get('branch_code') or '').strip().upper()
    if branch_code == 'TEMP_BRANCH' or str(branch_no_raw).strip() == '-1':
        return '임시'
    branch_no = str(branch_no_raw or '').strip()
    if branch_no in {'0', '본점'}:
        return '본점'
    if branch_no.isdigit():
        return f"{int(branch_no)}호점"
    return ''


def _material_requester_display_name_from_user(user_row: dict | None) -> str:
    data = row_to_dict(user_row)
    return str(
        data.get('name')
        or data.get('nickname')
        or data.get('login_id')
        or data.get('email')
        or '구매신청자'
    ).strip()


def _notification_actor_prefix_and_name(user_row: dict | None) -> tuple[str, str]:
    data = row_to_dict(user_row)
    position_title = str(data.get('position_title') or '').strip()
    branch_label = _material_branch_label_from_user(data)
    display_name = str(
        data.get('name')
        or data.get('nickname')
        or data.get('login_id')
        or data.get('email')
        or '회원'
    ).strip()
    if position_title in {'대표', '부대표', '호점대표'} and branch_label:
        return branch_label, display_name
    if position_title:
        return position_title, display_name
    fallback_grade_label = str(grade_label(data.get('grade')) or '').strip()
    if fallback_grade_label:
        return fallback_grade_label, display_name
    return '직원', display_name


def _friend_notification_body(user_row: dict | None, suffix: str) -> str:
    prefix, display_name = _notification_actor_prefix_and_name(user_row)
    return f'[{prefix}] [{display_name}]님이 {suffix}'


def _material_request_detail(conn, request_row: dict) -> dict:
    items = [
        row_to_dict(row)
        for row in conn.execute(
            '''
            SELECT i.id, i.product_id, i.quantity, i.unit_price, i.line_total, i.memo, p.code, p.name, p.short_name, p.unit_label
            FROM material_purchase_request_items i
            JOIN material_products p ON p.id = i.product_id
            WHERE i.request_id = ?
            ORDER BY p.display_order, i.id
            ''',
            (request_row['id'],),
        ).fetchall()
    ]
    detail = {
        **request_row,
        'items': items,
    }
    user_id = int(detail.get('user_id') or 0)
    if user_id > 0:
        user_row = conn.execute(
            "SELECT id, login_id, email, google_email, name, nickname, account_unique_id, account_status, account_type, grade, position_title, branch_no, branch_code FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if user_row:
            user_dict = row_to_dict(user_row)
            branch_label = _material_branch_label_from_user(user_dict)
            display_name = _material_requester_display_name_from_user(user_dict)
            detail['requester_login_id'] = str(user_dict.get('login_id') or '').strip()
            detail['requester_email'] = str(user_dict.get('email') or '').strip()
            detail['requester_google_email'] = str(user_dict.get('google_email') or '').strip()
            detail['requester_account_status'] = str(user_dict.get('account_status') or '').strip()
            detail['requester_account_type'] = str(user_dict.get('account_type') or '').strip()
            detail['requester_position_title'] = str(user_dict.get('position_title') or '').strip()
            detail['requester_account_unique_id'] = str(user_dict.get('account_unique_id') or detail.get('requester_unique_id') or '').strip().lower()
            detail['requester_unique_id'] = detail['requester_account_unique_id']
            detail['requester_branch_no'] = user_dict.get('branch_no')
            detail['requester_branch_code'] = str(user_dict.get('branch_code') or '').strip()
            detail['requester_branch_label'] = branch_label
            detail['requester_display_name'] = display_name
            detail['requester_user_name'] = str(user_dict.get('name') or '').strip()
            detail['requester_nickname'] = str(user_dict.get('nickname') or '').strip()
            if branch_label and display_name:
                detail['requester_name'] = f"{branch_label} {display_name}".strip()
            elif display_name:
                detail['requester_name'] = display_name
    return detail

def _pending_material_quantity_map(conn) -> dict[int, int]:
    rows = conn.execute(
        """
        SELECT i.product_id, COALESCE(SUM(i.quantity), 0) AS total_qty
        FROM material_purchase_request_items i
        JOIN material_purchase_requests r ON r.id = i.request_id
        WHERE r.status = 'pending' AND COALESCE(i.quantity, 0) > 0
        GROUP BY i.product_id
        """
    ).fetchall()
    return {int(row['product_id']): int(row['total_qty'] or 0) for row in rows}


def _material_request_quantity_map(conn, request_ids: list[int]) -> dict[int, int]:
    ids = sorted({int(item) for item in request_ids if int(item or 0) > 0})
    if not ids:
        return {}
    placeholders = ','.join('?' for _ in ids)
    rows = conn.execute(
        f"SELECT product_id, COALESCE(SUM(quantity), 0) AS total_qty FROM material_purchase_request_items WHERE request_id IN ({placeholders}) AND COALESCE(quantity, 0) > 0 GROUP BY product_id",
        tuple(ids),
    ).fetchall()
    return {int(row['product_id']): int(row['total_qty'] or 0) for row in rows}


def _adjust_material_stock_by_product_map(conn, quantity_map: dict[int, int], direction: int, changed_at: str):
    if not quantity_map:
        return
    products = {
        int(row['id']): row_to_dict(row)
        for row in conn.execute("SELECT id, current_stock FROM material_products WHERE COALESCE(is_active, 1) = 1").fetchall()
    }
    for product_id, qty in quantity_map.items():
        if product_id not in products:
            continue
        delta = int(qty or 0) * int(direction or 0)
        if delta == 0:
            continue
        current_stock = max(0, int(products[product_id].get('current_stock') or 0) + delta)
        conn.execute(
            "UPDATE material_products SET current_stock = ?, updated_at = ? WHERE id = ?",
            (current_stock, changed_at, product_id),
        )


def _adjust_material_stock_by_request_ids(conn, request_ids: list[int], direction: int, changed_at: str):
    _adjust_material_stock_by_product_map(conn, _material_request_quantity_map(conn, request_ids), direction, changed_at)


def _material_today_inventory_rows(conn, target_date: str) -> list[dict]:
    products = _material_products(conn)
    pending_qty_map = _pending_material_quantity_map(conn)
    settled_rows = conn.execute(
        """
        SELECT i.product_id, COALESCE(SUM(i.quantity), 0) AS total_qty
        FROM material_purchase_request_items i
        JOIN material_purchase_requests r ON r.id = i.request_id
        WHERE r.status = 'settled' AND COALESCE(substr(r.settled_at, 1, 10), '') = ? AND COALESCE(i.quantity, 0) > 0
        GROUP BY i.product_id
        """,
        (target_date,),
    ).fetchall()
    settled_outgoing_map = {int(row['product_id']): int(row['total_qty'] or 0) for row in settled_rows}
    daily_rows = conn.execute(
        "SELECT * FROM material_inventory_daily WHERE inventory_date = ?",
        (target_date,),
    ).fetchall()
    daily_map = {int(row['product_id']): row_to_dict(row) for row in daily_rows}
    output = []
    for product in products:
        product_id = int(product['id'])
        row = daily_map.get(product_id, {})
        base_current_stock = max(0, int(product.get('current_stock') or 0))
        pending_qty = int(pending_qty_map.get(product_id, 0) or 0)
        incoming_qty = int(row.get('incoming_qty') or 0)
        manual_outgoing_qty = int(row.get('outgoing_qty') or 0)
        settled_outgoing_qty = int(settled_outgoing_map.get(product_id, 0) or 0)
        available_stock = max(0, base_current_stock - pending_qty)
        output.append({
            'product_id': product_id,
            'code': product.get('code', ''),
            'name': product.get('name', ''),
            'short_name': product.get('short_name', ''),
            'unit_price': int(product.get('unit_price') or 0),
            'base_current_stock': base_current_stock,
            'current_stock': available_stock,
            'pending_qty': pending_qty,
            'incoming_qty': incoming_qty,
            'outgoing_qty': settled_outgoing_qty + manual_outgoing_qty,
            'settled_outgoing_qty': settled_outgoing_qty,
            'manual_outgoing_qty': manual_outgoing_qty,
            'note': row.get('note', '') or '',
            'is_closed': bool(int(row.get('is_closed') or 0)) if row else False,
            'closed_at': row.get('closed_at', '') or '',
            'expected_stock': available_stock,
        })
    return output

def _material_request_identity_candidates(user: dict) -> dict[str, set[str]]:
    def _clean(value) -> str:
        return str(value or '').strip()

    branch_label = _material_branch_label_from_user(user)
    name = _clean(user.get('name'))
    nickname = _clean(user.get('nickname'))
    login_id = _clean(user.get('login_id')).lower()
    email = _clean(user.get('email')).lower()
    account_unique_id = _clean(user.get('account_unique_id')).lower()

    requester_names = set()
    display_candidates = [name, nickname, login_id, email]
    for candidate in display_candidates:
        if candidate:
            requester_names.add(candidate)
            if branch_label:
                requester_names.add(f'{branch_label} {candidate}'.strip())

    unique_keys = {value for value in {account_unique_id, login_id, email} if value}
    return {
        'requester_names': requester_names,
        'unique_keys': unique_keys,
    }


def _material_request_belongs_to_user(request_row: dict, user: dict) -> bool:
    if int(request_row.get('user_id') or 0) == int(user.get('id') or 0):
        return True
    identity = _material_request_identity_candidates(user)
    requester_unique_id = str(request_row.get('requester_unique_id') or '').strip().lower()
    if requester_unique_id and requester_unique_id in identity['unique_keys']:
        return True
    requester_name = str(request_row.get('requester_name') or '').strip()
    if requester_name and requester_name in identity['requester_names']:
        return True
    normalized_requester_name = requester_name.replace(' ', '').lower()
    for candidate in identity['requester_names']:
        normalized_candidate = str(candidate or '').replace(' ', '').lower()
        if not normalized_candidate:
            continue
        if normalized_candidate in normalized_requester_name or normalized_requester_name in normalized_candidate:
            return True
    return False


def _material_share_text(requests: list[dict]) -> str:
    lines = ['[구매자결산표]']
    for request in requests:
        created_date = str(request.get('created_at') or '')[:10]
        lines.append(f"- {created_date} | {request.get('requester_name', '')}")
        for item in request.get('items', []):
            qty = int(item.get('quantity') or 0)
            if qty <= 0:
                continue
            lines.append(f"  · {item.get('short_name') or item.get('name')}: {qty}")
        lines.append(f"  · 합계: {int(request.get('total_amount') or 0):,}원")
    return '\n'.join(lines)



def _pending_material_settlement_info(conn, user: dict) -> dict:
    if _grade_of(user) > 2:
        return {'count': 0, 'body': '', 'latest_date': ''}
    rows = conn.execute(
        "SELECT created_at FROM material_purchase_requests WHERE status = 'pending' ORDER BY created_at DESC, id DESC"
    ).fetchall()
    count = len(rows)
    if count <= 0:
        return {'count': 0, 'body': '', 'latest_date': ''}
    created_at = str(rows[0]['created_at'] or '')
    date_text = str(created_at)[:10]
    try:
        dt = datetime.strptime(date_text, '%Y-%m-%d')
        label = f"{dt.month}월 {dt.day}일 신청한 미결제한 자재결산이 있습니다."
    except Exception:
        label = '미결제한 자재결산이 있습니다.'
    return {'count': count, 'body': label, 'latest_date': date_text}

def _material_notification_lines(request_detail: dict) -> tuple[str, str]:
    actor_like_row = {
        'position_title': request_detail.get('requester_position_title'),
        'grade': request_detail.get('grade') or request_detail.get('requester_grade'),
        'branch_no': request_detail.get('requester_branch_no'),
        'branch_code': request_detail.get('requester_branch_code'),
        'name': request_detail.get('requester_user_name') or request_detail.get('requester_display_name') or request_detail.get('requester_name'),
        'nickname': request_detail.get('requester_nickname'),
        'login_id': request_detail.get('requester_login_id'),
        'email': request_detail.get('requester_email') or request_detail.get('requester_google_email'),
    }
    requester_prefix, requester_name = _notification_actor_prefix_and_name(actor_like_row)
    title = '자재구매 신청 접수'
    item_chunks: list[str] = []
    for item in request_detail.get('items', []) or []:
        item_name = str(item.get('short_name') or item.get('name') or '').strip()
        qty = int(item.get('quantity') or 0)
        if not item_name or qty <= 0:
            continue
        item_chunks.append(f'[{item_name} / {qty}개]')
    total_amount = int(request_detail.get('total_amount') or 0)
    body_parts = []
    if requester_prefix:
        body_parts.append(f'[{requester_prefix}]')
    if requester_name:
        body_parts.append(f'[{requester_name}]')
    body_parts.extend(item_chunks[:6])
    body_parts.append(f'[{total_amount:,}원]')
    return title, ' '.join(body_parts)


def _notify_material_purchase_request(conn, requester_user: dict, request_detail: dict) -> None:
    admin_rows = conn.execute(
        """
        SELECT id
        FROM users
        WHERE CAST(COALESCE(grade, '6') AS INTEGER) <= 2
          AND COALESCE(account_status, 'active') NOT IN ('pending', 'suspended', 'retired', 'deleted', '승인대기', '일시정지', '퇴사/종료', '계정삭제')
        ORDER BY CAST(COALESCE(grade, '6') AS INTEGER), id
        """
    ).fetchall()
    if not admin_rows:
        return
    title, body = _material_notification_lines(request_detail)
    requester_id = int(requester_user.get('id') or 0)
    for row in admin_rows:
        admin_id = int(row['id'] or 0)
        if admin_id <= 0 or admin_id == requester_id:
            continue
        insert_notification(conn, admin_id, 'material_purchase_request', title, body)



def _material_overview_payload(conn, user: dict) -> dict:
    today_key = datetime.now().date().isoformat()
    permissions = _material_permissions(user)
    request_rows = [
        _material_request_detail(conn, row_to_dict(row))
        for row in conn.execute(
            "SELECT * FROM material_purchase_requests ORDER BY created_at DESC, id DESC LIMIT 1000"
        ).fetchall()
    ]
    pending_requests = [row for row in request_rows if row.get('status') == 'pending']
    settled_requests = [row for row in request_rows if row.get('status') == 'settled']
    rejected_requests = [row for row in request_rows if row.get('status') == 'rejected']
    history_rows = settled_requests[:] + rejected_requests[:]
    history_rows.sort(key=lambda row: str(row.get('created_at') or ''), reverse=True)

    user_id = int(user.get('id') or 0)
    my_request_seen_ids: set[int] = set()
    my_request_rows: list[dict] = []

    if user_id > 0:
        identity = _material_request_identity_candidates(user)
        unique_keys = sorted(identity.get('unique_keys') or [])
        requester_names = sorted(identity.get('requester_names') or [])
        own_rows = [
            _material_request_detail(conn, row_to_dict(row))
            for row in conn.execute(
                "SELECT * FROM material_purchase_requests WHERE user_id = ? ORDER BY created_at DESC, id DESC LIMIT 1000",
                (user_id,),
            ).fetchall()
        ]
        extra_rows = []
        if unique_keys:
            placeholders = ','.join('?' for _ in unique_keys)
            extra_rows.extend([
                _material_request_detail(conn, row_to_dict(row))
                for row in conn.execute(
                    f"SELECT * FROM material_purchase_requests WHERE COALESCE(requester_unique_id, '') IN ({placeholders}) ORDER BY created_at DESC, id DESC LIMIT 1000",
                    tuple(unique_keys),
                ).fetchall()
            ])
        if requester_names:
            placeholders = ','.join('?' for _ in requester_names)
            extra_rows.extend([
                _material_request_detail(conn, row_to_dict(row))
                for row in conn.execute(
                    f"SELECT * FROM material_purchase_requests WHERE COALESCE(requester_name, '') IN ({placeholders}) ORDER BY created_at DESC, id DESC LIMIT 1000",
                    tuple(requester_names),
                ).fetchall()
            ])
        for row in own_rows + extra_rows:
            row_id = int(row.get('id') or 0)
            if row_id <= 0 or row_id in my_request_seen_ids:
                continue
            if not _material_request_belongs_to_user(row, user):
                continue
            my_request_seen_ids.add(row_id)
            my_request_rows.append(row)

    for row in request_rows:
        row_id = int(row.get('id') or 0)
        if row_id in my_request_seen_ids:
            continue
        if _material_request_belongs_to_user(row, user):
            my_request_seen_ids.add(row_id)
            my_request_rows.append(row)

    my_request_rows.sort(key=lambda row: (str(row.get('created_at') or ''), int(row.get('id') or 0)), reverse=True)
    products = _material_products(conn)
    inventory_rows = _material_today_inventory_rows(conn, today_key)
    inventory_map = {int(row['product_id']): row for row in inventory_rows}
    pending_qty_map = _pending_material_quantity_map(conn)
    effective_products = []
    for product in products:
        product_copy = dict(product)
        product_id = int(product_copy['id'])
        inventory_row = inventory_map.get(product_id, {})
        available_stock = int(inventory_row.get('current_stock', product_copy.get('current_stock') or 0) or 0)
        product_copy['base_current_stock'] = int(inventory_row.get('base_current_stock', product_copy.get('current_stock') or 0) or 0)
        product_copy['pending_qty'] = int(pending_qty_map.get(product_id, 0) or 0)
        product_copy['current_stock'] = max(0, available_stock)
        effective_products.append(product_copy)
    payload = {
        'today': today_key,
        'permissions': permissions,
        'products': effective_products,
        'pending_requests': pending_requests if permissions['can_view_requesters'] else [],
        'settled_requests': settled_requests if permissions['can_view_settlements'] else [],
        'history_requests': history_rows if permissions['can_view_history'] else [],
        'my_requests': my_request_rows if permissions['can_view_my_requests'] else [],
        'inventory_rows': inventory_rows if permissions['can_view_inventory'] else [],
        'share_text': _material_share_text(settled_requests[:30]) if permissions['can_view_settlements'] else '',
    }
    logger.info(
        'Materials overview built user_id=%s pending_count=%s settled_count=%s my_requests_count=%s can_view_requesters=%s',
        user_id,
        len(payload['pending_requests']),
        len(payload['settled_requests']),
        len(payload['my_requests']),
        bool(permissions.get('can_view_requesters')),
    )
    return payload

def _require_write_access(user: dict, area: str):
    grade = _grade_of(user)
    if grade >= 7:
        raise HTTPException(status_code=403, detail='현재 권한으로는 사용할 수 없습니다.')
    if grade == 6 and area in {'schedule', 'work_schedule'}:
        raise HTTPException(status_code=403, detail='일반 등급은 해당 기능을 관람만 할 수 있습니다.')
def get_optional_user(authorization: Optional[str] = Header(default=None)):
    token = _bearer_token(authorization)
    if not token:
        return None
    with get_conn() as conn:
        user = get_user_by_token(conn, token)
        return row_to_dict(user) if user else None

def require_user(authorization: Optional[str] = Header(default=None)):
    token = _bearer_token(authorization)
    if not token:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다.")
    with get_conn() as conn:
        user = get_user_by_token(conn, token)
        if not user:
            raise HTTPException(status_code=401, detail="유효하지 않은 세션입니다.")
        user_dict = row_to_dict(user)
        status = _normalize_account_status_value(user_dict.get('account_status'), user_dict.get('approved'), user_dict.get('grade'))
        if status in {'suspended', 'retired', 'deleted', 'pending'}:
            raise HTTPException(status_code=403, detail='현재 계정 상태로는 앱을 사용할 수 없습니다.')
        return user_dict
def require_admin(user=Depends(require_user)):
    if _grade_of(user) != 1:
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")
    return user
def require_admin_mode_user(user=Depends(require_user)):
    with get_conn() as conn:
        if not _can_access_admin_mode(user, conn):
            raise HTTPException(status_code=403, detail='권한이 없습니다.')
    return user
def require_admin_or_subadmin(user=Depends(require_user)):
    if _grade_of(user) > 2:
        raise HTTPException(status_code=403, detail='관리자 또는 부관리자 권한이 필요합니다.')
    return user
def can_manage_group_room(user: dict) -> bool:
    return _grade_of(user) <= 2

def is_blocked(conn, user_id: int, other_id: int) -> bool:
    row = conn.execute(
        """
        SELECT 1 FROM blocks
        WHERE (blocker_id = ? AND blocked_user_id = ?)
           OR (blocker_id = ? AND blocked_user_id = ?)
        LIMIT 1
        """,
        (user_id, other_id, other_id, user_id),
    ).fetchone()
    return bool(row)
def room_key(a: int, b: int) -> str:
    low, high = sorted([a, b])
    return f"{low}:{high}"
def user_basic(conn, user_id: int) -> dict:
    row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")
    return user_public_dict(row)
def get_room_setting(conn, user_id: int, room_type: str, room_ref: str) -> dict:
    row = conn.execute(
        "SELECT * FROM chat_room_settings WHERE user_id = ? AND room_type = ? AND room_ref = ?",
        (user_id, room_type, room_ref),
    ).fetchone()
    if row:
        data = row_to_dict(row)
        data.update({
            'pinned': bool(data.get('pinned', 0)),
            'favorite': bool(data.get('favorite', 0)),
            'muted': bool(data.get('muted', 0)),
            'hidden': bool(data.get('hidden', 0)),
        })
        return data
    return {'custom_name': '', 'pinned': False, 'favorite': False, 'muted': False, 'hidden': False}
def save_room_setting(conn, user_id: int, room_type: str, room_ref: str, payload: dict) -> dict:
    current = get_room_setting(conn, user_id, room_type, room_ref)
    next_data = {
        'custom_name': payload.get('custom_name', current.get('custom_name', '')),
        'pinned': int(payload.get('pinned', current.get('pinned', False))),
        'favorite': int(payload.get('favorite', current.get('favorite', False))),
        'muted': int(payload.get('muted', current.get('muted', False))),
        'hidden': int(payload.get('hidden', current.get('hidden', False))),
    }
    conn.execute(
        """
        INSERT INTO chat_room_settings(user_id, room_type, room_ref, custom_name, pinned, favorite, muted, hidden, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id, room_type, room_ref) DO UPDATE SET
            custom_name = excluded.custom_name,
            pinned = excluded.pinned,
            favorite = excluded.favorite,
            muted = excluded.muted,
            hidden = excluded.hidden,
            updated_at = excluded.updated_at
        """,
        (user_id, room_type, room_ref, next_data['custom_name'], next_data['pinned'], next_data['favorite'], next_data['muted'], next_data['hidden'], utcnow(), utcnow()),
    )
    return get_room_setting(conn, user_id, room_type, room_ref)
def build_reply_meta(conn, table: str, reply_to_id: Optional[int]):
    if not reply_to_id:
        return None
    row = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (reply_to_id,)).fetchone()
    if not row:
        return None
    item = row_to_dict(row)
    return {
        'id': item['id'],
        'message': item.get('message', ''),
        'sender': user_basic(conn, item['sender_id']),
    }
def reaction_summary(value: str):
    items = json_loads(value, [])
    summary = {}
    normalized = []
    for entry in items:
        if not isinstance(entry, dict):
            continue
        uid = entry.get('user_id')
        emoji = entry.get('emoji')
        if not uid or not emoji:
            continue
        normalized.append({'user_id': uid, 'emoji': emoji})
        if emoji not in summary:
            summary[emoji] = {'emoji': emoji, 'count': 0, 'user_ids': []}
        summary[emoji]['count'] += 1
        summary[emoji]['user_ids'].append(uid)
    return normalized, list(summary.values())
def enrich_chat_message(conn, row, table: str):
    item = row_to_dict(row)
    raw_reactions, grouped = reaction_summary(item.get('reactions', '[]'))
    return {
        **item,
        'sender': user_basic(conn, item['sender_id']),
        'reply_to': build_reply_meta(conn, table, item.get('reply_to_id')),
        'reactions': raw_reactions,
        'reaction_summary': grouped,
    }
def insert_chat_mention(conn, user_id: int, room_type: str, room_ref: str, message_id: int, sender_id: int) -> None:
    if not user_id or user_id == sender_id:
        return
    conn.execute(
        "INSERT INTO chat_mentions(user_id, room_type, room_ref, message_id, sender_id, seen, created_at) VALUES (?, ?, ?, ?, ?, 0, ?)",
        (user_id, room_type, room_ref, message_id, sender_id, utcnow()),
    )
def toggle_reaction(conn, table: str, message_id: int, user_id: int, emoji: str):
    row = conn.execute(f"SELECT reactions FROM {table} WHERE id = ?", (message_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail='메시지를 찾을 수 없습니다.')
    items = json_loads(row['reactions'], [])
    next_items = []
    replaced = False
    for entry in items:
        if not isinstance(entry, dict):
            continue
        if entry.get('user_id') == user_id:
            if entry.get('emoji') == emoji:
                replaced = True
                continue
            next_items.append({'user_id': user_id, 'emoji': emoji})
            replaced = True
        else:
            next_items.append({'user_id': entry.get('user_id'), 'emoji': entry.get('emoji')})
    if not replaced:
        next_items.append({'user_id': user_id, 'emoji': emoji})
    conn.execute(f"UPDATE {table} SET reactions = ? WHERE id = ?", (json.dumps(next_items, ensure_ascii=False), message_id))
def room_preview_text(message: dict) -> str:
    if message.get('attachment_type') == 'image':
        return '[사진]'
    if message.get('attachment_type') == 'file':
        return '[파일]'
    if message.get('attachment_type') == 'location':
        return '[위치]'
    return message.get('message', '')
def enrich_post(conn, post_row):
    post = row_to_dict(post_row)
    user = user_basic(conn, post["user_id"])
    likes = conn.execute("SELECT COUNT(*) FROM feed_likes WHERE post_id = ?", (post["id"],)).fetchone()[0]
    bookmarks = conn.execute("SELECT COUNT(*) FROM feed_bookmarks WHERE post_id = ?", (post["id"],)).fetchone()[0]
    comments = [
        {
            **row_to_dict(r),
            "user": user_basic(conn, r["user_id"]),
        }
        for r in conn.execute("SELECT * FROM feed_comments WHERE post_id = ? ORDER BY id DESC", (post["id"],)).fetchall()
    ]
    return {
        **post,
        "user": user,
        "likes_count": likes,
        "bookmarks_count": bookmarks,
        "comments": comments,
    }
@app.middleware("http")
async def request_logger(request: Request, call_next):
    started = datetime.utcnow()
    response = await call_next(request)
    elapsed_ms = int((datetime.utcnow() - started).total_seconds() * 1000)
    logger.info("%s %s -> %s (%sms)", request.method, request.url.path, response.status_code, elapsed_ms)
    return response


@app.middleware("http")
async def ensure_api_cors_headers(request: Request, call_next):
    origin = request.headers.get('origin', '')
    try:
        response = await call_next(request)
    except HTTPException as exc:
        response = JSONResponse(status_code=exc.status_code, content={'detail': exc.detail})
    except Exception as exc:
        logger.exception('Unhandled API error: %s %s', request.method, request.url.path)
        response = JSONResponse(status_code=500, content={'detail': '서버 내부 오류가 발생했습니다.'})
    if origin and _origin_allowed(origin):
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, PATCH, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = request.headers.get('access-control-request-headers', '*') or '*'
        vary = response.headers.get('Vary', '')
        response.headers['Vary'] = 'Origin' if not vary else (vary if 'Origin' in vary else f"{vary}, Origin")
    return response


@app.on_event("startup")
def startup():
    settings.upload_root.mkdir(parents=True, exist_ok=True)
    settings.settlement_runtime_dir.mkdir(parents=True, exist_ok=True)
    init_db()
    with get_conn() as conn:
        _sync_all_day_note_available_vehicle_counts(conn)
    settlement_sync_service.start()
    runtime = get_settings()
    cred = _credential_summary()
    logger.info("startup complete env=%s db_engine=%s policy=%s soomgo_email_env=%s soomgo_password_env=%s configured=%s", runtime.app_env, DB_ENGINE, runtime.policy_url, cred.get('email_env') or '없음', cred.get('password_env') or '없음', cred.get('configured'))


@app.get("/api/health")
def health():
    runtime = get_settings()
    safe_db_label = "postgresql" if DB_ENGINE == "postgresql" else "sqlite"
    return {
        "ok": True,
        "app_env": runtime.app_env,
        "db_engine": DB_ENGINE,
        "db_label": safe_db_label,
        "site_url": runtime.app_public_url,
        "api_url": runtime.api_public_url,
        "policy_url": runtime.policy_url,
        "r2_enabled": runtime.r2_enabled,
        "r2_public_base_url": runtime.r2_public_base_url,
        "soomgo_credentials_configured": _credential_summary().get('configured'),
        "soomgo_email_env": _credential_summary().get('email_env'),
        "soomgo_password_env": _credential_summary().get('password_env'),
    }


@app.get("/api/settlement/platform-sync-status")
def settlement_platform_sync_status(user=Depends(require_user)):
    return settlement_sync_service.status()


@app.get("/api/settlement/platform-credentials")
def settlement_platform_credentials(platform: str = Query('숨고'), user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    if platform not in ('숨고', '오늘'):
        raise HTTPException(status_code=400, detail='지원하지 않는 플랫폼입니다.')
    summary = _credential_summary(platform)
    return {
        'platform': platform,
        'configured': bool(summary.get('configured')),
        'email_source': summary.get('email_env') or '없음',
        'password_source': summary.get('password_env') or '없음',
        'email_preview': '저장됨' if summary.get('email_present') else '',
        'auth_state_present': bool(summary.get('auth_state_present')),
    }


@app.post("/api/settlement/platform-credentials")
def settlement_platform_credentials_save(payload: SettlementCredentialIn, user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    platform = (payload.platform or '숨고').strip()
    if platform not in ('숨고', '오늘'):
        raise HTTPException(status_code=400, detail='지원하지 않는 플랫폼입니다.')
    email = (payload.email or '').strip()
    password = (payload.password or '').strip()
    if not email or not password:
        raise HTTPException(status_code=400, detail=f'{platform} 아이디와 비밀번호를 모두 입력해 주세요.')
    email_key = 'soomgo_email' if platform == '숨고' else 'ohou_email'
    password_key = 'soomgo_password' if platform == '숨고' else 'ohou_password'
    now_iso = utcnow()
    with get_conn() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO app_secrets(secret_key, secret_value, updated_at) VALUES (?, ?, ?)",
            (email_key, email, now_iso),
        )
        conn.execute(
            "INSERT OR REPLACE INTO app_secrets(secret_key, secret_value, updated_at) VALUES (?, ?, ?)",
            (password_key, password, now_iso),
        )
    return settlement_platform_credentials(platform=platform, user=user)


@app.post("/api/settlement/platform-auth-state")
def settlement_platform_auth_state_save(payload: SettlementAuthStateIn, user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    platform = (payload.platform or '숨고').strip()
    if platform not in ('숨고', '오늘'):
        raise HTTPException(status_code=400, detail='지원하지 않는 플랫폼입니다.')
    try:
        return save_auth_state_json(payload.storage_state, platform=platform)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/settlement/platform-auth-guide")
def settlement_platform_auth_guide(platform: str = Query('숨고'), user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    if platform not in ('숨고', '오늘'):
        raise HTTPException(status_code=400, detail='지원하지 않는 플랫폼입니다.')
    return get_auth_session_guide(platform)


@app.post("/api/settlement/platform-sync/refresh")
def settlement_platform_sync_refresh(user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    try:
        return settlement_sync_service.run_once(trigger='manual')
    except RuntimeError as exc:
        message = str(exc)
        status_code = 409 if '이미 데이터 연동이 진행 중' in message else 400
        raise HTTPException(status_code=status_code, detail=message) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f'데이터 연동 중 오류가 발생했습니다: {exc}') from exc




def _normalize_settlement_date(value: str) -> str:
    raw = str(value or '').strip()
    if not raw:
        raise HTTPException(status_code=400, detail='결산 날짜가 비어 있습니다.')
    for fmt in ('%Y-%m-%d', '%y.%m.%d', '%y.%m.%d.', '%Y.%m.%d', '%Y.%m.%d.'):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    raise HTTPException(status_code=400, detail=f'지원하지 않는 결산 날짜 형식입니다: {raw}')


def _safe_settlement_block(block: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(block, dict):
        raise HTTPException(status_code=400, detail='결산 데이터 형식이 올바르지 않습니다.')
    return {
        'title': str(block.get('title') or '').strip(),
        'date': str(block.get('date') or '').strip(),
        'summaryHeaders': list(block.get('summaryHeaders') or []),
        'summaryRows': list(block.get('summaryRows') or []),
        'reviewHeaders': list(block.get('reviewHeaders') or []),
        'branchRows': list(block.get('branchRows') or []),
        'total': block.get('total') if isinstance(block.get('total'), dict) else {},
    }


def _settlement_metric_map(block: dict[str, Any]) -> dict[str, float]:
    rows = block.get('summaryRows') or []
    metrics: dict[str, float] = {
        '숨고': 0.0,
        '오늘': 0.0,
        '공홈': 0.0,
        '총견적': 0.0,
        '총계약': 0.0,
        '계약률': 0.0,
        '플랫폼리뷰': 0.0,
        '호점리뷰': 0.0,
        '이슈': 0.0,
    }
    for row in rows:
        source = str((row or {}).get('source') or '').strip()
        count_raw = str((row or {}).get('count') or '0').replace(',', '').strip()
        value_raw = str((row or {}).get('value') or '0').replace(',', '').strip()
        try:
            count_value = float(count_raw or 0)
        except ValueError:
            count_value = 0.0
        try:
            value_value = float(value_raw or 0)
        except ValueError:
            value_value = 0.0
        if source in ('숨고', '오늘', '공홈'):
            metrics[source] = count_value
        label = str((row or {}).get('label') or '')
        if '총 견적 발송 수' in label:
            metrics['총견적'] = value_value
        elif '총 계약 수' in label:
            metrics['총계약'] = value_value
        elif '계약률' in label:
            metrics['계약률'] = value_value
    total = block.get('total') if isinstance(block.get('total'), dict) else {}
    for key, metric_name in [('platformReview', '플랫폼리뷰'), ('branchReview', '호점리뷰'), ('issues', '이슈')]:
        raw = str(total.get(key) or '0').replace(',', '').strip()
        try:
            metrics[metric_name] = float(raw or 0)
        except ValueError:
            metrics[metric_name] = 0.0
    return metrics


def _settlement_period_labels(day: date) -> tuple[str, str]:
    week_start = day - timedelta(days=day.weekday())
    week_end = week_start + timedelta(days=6)
    week_key = f"{week_start.isoformat()}~{week_end.isoformat()}"
    month_key = day.strftime('%Y-%m')
    return week_key, month_key


def _aggregate_settlement_records(rows: list[dict[str, Any]], unit: str) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        settlement_day = datetime.strptime(row['settlement_date'], '%Y-%m-%d').date()
        week_key, month_key = _settlement_period_labels(settlement_day)
        group_key = week_key if unit == 'weekly' else month_key
        bucket = grouped.setdefault(group_key, {
            'period_key': group_key,
            'period_label': group_key,
            'record_count': 0,
            'dates': [],
            'metrics': {
                '숨고': 0.0,
                '오늘': 0.0,
                '공홈': 0.0,
                '총견적': 0.0,
                '총계약': 0.0,
                '플랫폼리뷰': 0.0,
                '호점리뷰': 0.0,
                '이슈': 0.0,
            },
            'last_reflected_at': '',
        })
        bucket['record_count'] += 1
        bucket['dates'].append(row['settlement_date'])
        metrics = _settlement_metric_map(row.get('block') or {})
        for key in bucket['metrics']:
            bucket['metrics'][key] += metrics.get(key, 0.0)
        reflected_at = row.get('reflected_at') or ''
        if reflected_at and reflected_at > bucket['last_reflected_at']:
            bucket['last_reflected_at'] = reflected_at
    result: list[dict[str, Any]] = []
    for _, bucket in sorted(grouped.items(), key=lambda item: item[0], reverse=True):
        total_quotes = bucket['metrics']['총견적']
        total_contracts = bucket['metrics']['총계약']
        result.append({
            'period_key': bucket['period_key'],
            'period_label': bucket['period_label'],
            'record_count': bucket['record_count'],
            'date_range': {
                'start': min(bucket['dates']) if bucket['dates'] else '',
                'end': max(bucket['dates']) if bucket['dates'] else '',
            },
            'summary': {
                '숨고': int(bucket['metrics']['숨고']),
                '오늘': int(bucket['metrics']['오늘']),
                '공홈': int(bucket['metrics']['공홈']),
                '총견적': int(bucket['metrics']['총견적']),
                '총계약': int(bucket['metrics']['총계약']),
                '계약률': round((total_contracts / total_quotes), 6) if total_quotes else 0,
                '플랫폼리뷰': int(bucket['metrics']['플랫폼리뷰']),
                '호점리뷰': int(bucket['metrics']['호점리뷰']),
                '이슈': int(bucket['metrics']['이슈']),
            },
            'last_reflected_at': bucket['last_reflected_at'],
        })
    return result


@app.get('/api/settlement/records')
def settlement_records(user=Depends(require_user)):
    rows: list[dict[str, Any]] = []
    with get_conn() as conn:
        fetched = conn.execute(
            "SELECT settlement_date, category, title, block_json, reflected_at, reflected_by_user_id, reflected_by_name FROM settlement_reflections WHERE category = 'daily' ORDER BY settlement_date DESC, reflected_at DESC"
        ).fetchall()
    for row in fetched:
        item = row_to_dict(row)
        item['block'] = json_loads(item.get('block_json'), {})
        rows.append(item)
    return {
        'daily_records': rows,
        'weekly_records': _aggregate_settlement_records(rows, 'weekly'),
        'monthly_records': _aggregate_settlement_records(rows, 'monthly'),
    }


@app.post('/api/settlement/records/reflect')
def settlement_records_reflect(payload: SettlementReflectIn, user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    category = (payload.category or 'daily').strip() or 'daily'
    if category != 'daily':
        raise HTTPException(status_code=400, detail='현재는 일일결산만 결산반영할 수 있습니다.')
    settlement_date = _normalize_settlement_date(payload.settlement_date)
    block = _safe_settlement_block(payload.block)
    reflected_at = utcnow()
    reflected_by_name = str(user.get('nickname') or user.get('name') or user.get('email') or '').strip()
    with get_conn() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO settlement_reflections(settlement_date, category, title, block_json, reflected_at, reflected_by_user_id, reflected_by_name) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                settlement_date,
                category,
                str(payload.title or block.get('title') or '').strip(),
                json.dumps(block, ensure_ascii=False),
                reflected_at,
                user.get('id'),
                reflected_by_name,
            ),
        )
    return {
        'ok': True,
        'settlement_date': settlement_date,
        'category': category,
        'reflected_at': reflected_at,
        'reflected_by_name': reflected_by_name,
        'block': block,
    }

@app.get("/api/deployment/meta")
def deployment_meta():
    return {
        "frontend_public_url": settings.app_public_url,
        "backend_public_url": settings.api_public_url,
        "policy_url": settings.policy_url,
        "account_deletion_url": settings.account_deletion_url,
        "recommended_frontend_hosting": "Cloudflare Pages",
        "recommended_backend_hosting": "Railway Hobby",
        "recommended_database": "Railway PostgreSQL",
        "recommended_storage": "Cloudflare R2",
        "recommended_dns": "Cloudflare",
    }


@app.post("/api/uploads/file")
async def upload_file(request: Request, file: UploadFile = File(...), category: str = Query("general"), user=Depends(require_user)):
    try:
        saved = save_upload(file, category=category)
    except StorageError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    url = saved["url"]
    if url.startswith("/"):
        base = str(request.base_url).rstrip("/")
        url = f"{base}{url}"
    return {**saved, "url": url, "uploaded_by": user["id"]}
@app.get("/api/demo-accounts")
def demo_accounts():
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT login_id, email, nickname, name, role, grade, group_number, group_number_text
            FROM users
            ORDER BY
                CASE
                    WHEN NULLIF(TRIM(COALESCE(group_number_text, '')), '') IS NULL THEN 1
                    ELSE 0
                END,
                LENGTH(COALESCE(NULLIF(TRIM(group_number_text), ''), CAST(COALESCE(group_number, 0) AS TEXT))),
                COALESCE(NULLIF(TRIM(group_number_text), ''), CAST(COALESCE(group_number, 0) AS TEXT)),
                id
            """
        ).fetchall()
        return [
            {
                "login_id": (r["login_id"] if "login_id" in r.keys() and r["login_id"] not in (None, '') else r["email"]),
                "email": r["email"],
                "nickname": r["nickname"],
                "name": r["name"],
                "role": r["role"],
                "grade": r["grade"],
                "grade_label": grade_label(r["grade"]),
                "group_number": str((r["group_number_text"] if r["group_number_text"] not in (None, '') else r["group_number"]) or '0'),
            }
            for r in rows
        ]
@app.post("/api/auth/signup")
def signup(payload: SignupIn):
    login_id = _validate_login_id_value(payload.login_id)
    account_email = _normalize_email_value(payload.email)
    password = payload.password.strip()
    nickname = payload.nickname.strip()
    gender = _validate_gender_value(payload.gender, allow_empty=False)
    region = payload.region.strip()
    phone = payload.phone.strip()
    recovery_email = _normalize_email_value(payload.recovery_email)
    actual_email = account_email
    google_email = _normalize_email_value(payload.google_email)
    vehicle_number = payload.vehicle_number.strip()

    required_fields = [
        ('아이디', login_id),
        ('비밀번호', password),
        ('닉네임', nickname),
        ('성별', gender),
        ('생년', str(payload.birth_year or '').strip()),
        ('지역', region),
        ('연락처', phone),
        ('복구 이메일', recovery_email),
    ]
    missing = [label for label, value in required_fields if not value]
    if missing:
        raise HTTPException(status_code=400, detail=f"다음 필수 항목을 입력해 주세요: {', '.join(missing)}")
    if payload.birth_year < 1900 or payload.birth_year > 2100:
        raise HTTPException(status_code=400, detail='생년 값이 올바르지 않습니다.')
    with get_conn() as conn:
        exists = _find_user_by_login_id_ci(conn, login_id)
        if exists:
            raise HTTPException(status_code=400, detail="이미 존재하는 아이디입니다.")
        generated_unique_id = generate_account_unique_id(conn, login_id)
        conn.execute(
            """
            INSERT INTO users(login_id, email, google_email, password_hash, name, nickname, role, grade, approved, account_status, permission_codes_json, account_type, branch_code, gender, birth_year, region, phone, recovery_email, vehicle_number, branch_no, account_unique_id, group_number, group_number_text, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 'user', 7, 0, 'pending', '[]', 'general', '', ?, ?, ?, ?, ?, ?, ?, ?, 0, '0', ?)
            """,
            (
                login_id,
                actual_email,
                google_email,
                hash_password(password),
                nickname,
                nickname,
                gender,
                payload.birth_year,
                region,
                phone,
                recovery_email,
                vehicle_number,
                payload.branch_no,
                payload.branch_no if payload.branch_no is not None else -1,
                generated_unique_id,
                utcnow(),
            ),
        )
        user_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.execute(
            "INSERT INTO preferences(user_id, data) VALUES (?, ?)",
            (user_id, json.dumps({"groupChatNotifications": True, "directChatNotifications": True, "likeNotifications": True, "theme": "dark"}, ensure_ascii=False)),
        )
        admin_rows = conn.execute(
            "SELECT id FROM users WHERE COALESCE(approved, 1) = 1 AND grade IN (1, 2)"
        ).fetchall()
        notification_title = '회원가입 승인 요청'
        notification_body = f"신규 회원가입 신청: {nickname} ({login_id})"
        for admin_row in admin_rows:
            insert_notification(conn, int(admin_row['id']), 'signup_request', notification_title, notification_body)
        return {
            'ok': True,
            'pending_approval': True,
            'message': '회원가입 신청이 완료되었습니다. 관리자 승인 후 일반 권한으로 로그인할 수 있습니다.',
            'user': {
                'id': user_id,
                'login_id': login_id,
                'email': actual_email,
                'nickname': nickname,
                'grade': 7,
                'grade_label': '기타',
                'approved': False,
            },
        }

@app.post('/api/auth/find-account')
def find_account(payload: AccountFindIn):
    nickname = payload.nickname.strip()
    phone = payload.phone.strip()
    recovery_email = payload.recovery_email.strip()
    if not nickname or not phone or not recovery_email:
        raise HTTPException(status_code=400, detail='닉네임, 연락처, 복구 이메일을 모두 입력해 주세요.')
    with get_conn() as conn:
        row = conn.execute(
            "SELECT email, nickname FROM users WHERE nickname = ? AND phone = ? AND recovery_email = ? ORDER BY id DESC LIMIT 1",
            (nickname, phone, recovery_email),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail='일치하는 계정을 찾을 수 없습니다.')
        return {'ok': True, 'account_id': row['email'], 'nickname': row['nickname'], 'message': '계정을 찾았습니다.'}

@app.post("/api/auth/login")
def login(payload: LoginIn):
    account_id = _validate_login_id_value(payload.login_id or payload.email)
    with get_conn() as conn:
        account = conn.execute(
            "SELECT * FROM users WHERE LOWER(TRIM(COALESCE(login_id, email, ''))) = ?",
            (account_id,),
        ).fetchone()
        if not account:
            raise HTTPException(status_code=404, detail='등록되지 않은 계정입니다.')
        if account['password_hash'] != hash_password(payload.password):
            raise HTTPException(status_code=401, detail='해당 계정의 비밀번호가 틀렸습니다.')
        grade = int(account['grade'] or 6)
        approved = int(account['approved'] if account['approved'] is not None else 1)
        account_status = _normalize_account_status_value(account['account_status'] if 'account_status' in account.keys() else '', approved, grade)
        if account_status == 'pending' or grade == 7 or not approved:
            raise HTTPException(status_code=403, detail="관리자 승인 후 로그인할 수 있습니다.")
        if account_status == 'suspended':
            raise HTTPException(status_code=403, detail='일시정지 상태 계정입니다.')
        if account_status == 'retired':
            raise HTTPException(status_code=403, detail='퇴사/종료 상태 계정입니다.')
        if account_status == 'deleted':
            raise HTTPException(status_code=403, detail='삭제된 계정입니다.')
        token = make_token()
        conn.execute("INSERT INTO auth_tokens(token, user_id, created_at) VALUES (?, ?, ?)", (token, account["id"], utcnow()))
        user_payload = user_public_dict(account)
        user_payload['permission_config'] = _get_permission_config(conn)
        return {'access_token': token, 'user': user_payload}

@app.post('/api/auth/verify-password')
def verify_current_password(payload: PasswordVerifyIn, user=Depends(require_user)):
    password = str(payload.password or '').strip()
    if not password:
        raise HTTPException(status_code=400, detail='비밀번호를 입력해 주세요.')
    with get_conn() as conn:
        account = conn.execute("SELECT password_hash FROM users WHERE id = ?", (user['id'],)).fetchone()
        if not account:
            raise HTTPException(status_code=404, detail='계정을 찾을 수 없습니다.')
        if str(account['password_hash'] or '') != hash_password(password):
            raise HTTPException(status_code=401, detail='비밀번호가 일치하지 않습니다.')
    return {'ok': True}

@app.post("/api/auth/logout")
def logout(user=Depends(require_user), authorization: Optional[str] = Header(default=None)):
    token = _bearer_token(authorization)
    with get_conn() as conn:
        conn.execute("DELETE FROM auth_tokens WHERE token = ?", (token,))
    return {"ok": True}

@app.delete('/api/account')
def delete_account(user=Depends(require_user)):
    with get_conn() as conn:
        exists = conn.execute('SELECT id FROM users WHERE id = ?', (user['id'],)).fetchone()
        if not exists:
            raise HTTPException(status_code=404, detail='사용자를 찾을 수 없습니다.')
        conn.execute('DELETE FROM users WHERE id = ?', (user['id'],))
    return {'ok': True, 'message': '계정이 삭제되었습니다.'}
@app.post("/api/auth/password-reset/request")
def password_reset_request(payload: PasswordResetRequestIn):
    today = datetime.utcnow().date().isoformat()
    with get_conn() as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM verification_codes WHERE recovery_email = ? AND substr(created_at,1,10) = ?",
            (payload.recovery_email, today),
        ).fetchone()[0]
        if count >= 2:
            raise HTTPException(status_code=429, detail="복구 이메일당 하루 최대 2회까지만 요청할 수 있습니다.")
        code = f"{random.randint(0, 999999):06d}"
        expires_at = (datetime.utcnow() + timedelta(minutes=10)).replace(microsecond=0).isoformat()
        conn.execute(
            "INSERT INTO verification_codes(recovery_email, code, purpose, expires_at, consumed, created_at) VALUES (?, ?, 'password_reset', ?, 0, ?)",
            (payload.recovery_email, code, expires_at, utcnow()),
        )
        response = {"ok": True, "message": "복구 코드가 발급되었습니다."}
        if EMAIL_DEMO_MODE:
            response["demo_code"] = code
        return response
@app.post("/api/auth/password-reset/confirm")
def password_reset_confirm(payload: PasswordResetConfirmIn):
    with get_conn() as conn:
        code_row = conn.execute(
            """
            SELECT * FROM verification_codes
            WHERE recovery_email = ? AND code = ? AND purpose = 'password_reset' AND consumed = 0
            ORDER BY id DESC LIMIT 1
            """,
            (payload.recovery_email, payload.code),
        ).fetchone()
        if not code_row:
            raise HTTPException(status_code=400, detail="복구 코드가 올바르지 않습니다.")
        if datetime.fromisoformat(code_row["expires_at"]) < datetime.utcnow():
            raise HTTPException(status_code=400, detail="복구 코드가 만료되었습니다.")
        user_row = conn.execute(
            "SELECT * FROM users WHERE email = ? AND recovery_email = ?",
            (payload.email, payload.recovery_email),
        ).fetchone()
        if not user_row:
            raise HTTPException(status_code=404, detail="이메일 또는 복구 이메일이 일치하지 않습니다.")
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (hash_password(payload.new_password), user_row["id"]))
        conn.execute("UPDATE verification_codes SET consumed = 1 WHERE id = ?", (code_row["id"],))
        return {"ok": True, "message": "비밀번호가 변경되었습니다."}
@app.get('/api/me')
def me(user=Depends(require_user)):
    with get_conn() as conn:
        cfg = _get_permission_config(conn)
    enriched = dict(user)
    enriched['permission_config'] = cfg
    return {'user': enriched}
@app.get("/api/home")
def home(user=Depends(require_user)):
    with get_conn() as conn:
        posts = conn.execute("SELECT * FROM feed_posts ORDER BY id DESC LIMIT 5").fetchall()
        return {
            "profile_completion": 80 if user["bio"] else 55,
            "recent_posts": [enrich_post(conn, p) for p in posts],
            "stats": {
                "friends": conn.execute("SELECT COUNT(*) FROM friends WHERE user_id = ?", (user["id"],)).fetchone()[0],
                "follows": conn.execute("SELECT COUNT(*) FROM follows WHERE from_user_id = ?", (user["id"],)).fetchone()[0],
                "notifications": conn.execute("SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0", (user["id"],)).fetchone()[0],
            }
        }
@app.get("/api/home-feed")
def home_feed(user=Depends(require_user)):
    with get_conn() as conn:
        posts = conn.execute("SELECT * FROM feed_posts ORDER BY id DESC").fetchall()
        return [enrich_post(conn, p) for p in posts]
@app.post("/api/feed")
def create_feed_post(payload: FeedPostIn, user=Depends(require_user)):
    if not payload.content.strip():
        raise HTTPException(status_code=400, detail="내용을 입력해 주세요.")
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO feed_posts(user_id, content, image_url, created_at) VALUES (?, ?, ?, ?)",
            (user["id"], payload.content.strip(), payload.image_url.strip(), utcnow()),
        )
        post_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        post = conn.execute("SELECT * FROM feed_posts WHERE id = ?", (post_id,)).fetchone()
        return enrich_post(conn, post)
@app.post("/api/feed/{post_id}/like")
def toggle_like(post_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        post = conn.execute("SELECT * FROM feed_posts WHERE id = ?", (post_id,)).fetchone()
        if not post:
            raise HTTPException(status_code=404, detail="피드를 찾을 수 없습니다.")
        exists = conn.execute("SELECT 1 FROM feed_likes WHERE post_id = ? AND user_id = ?", (post_id, user["id"])).fetchone()
        if exists:
            conn.execute("DELETE FROM feed_likes WHERE post_id = ? AND user_id = ?", (post_id, user["id"]))
            liked = False
        else:
            conn.execute("INSERT INTO feed_likes(post_id, user_id, created_at) VALUES (?, ?, ?)", (post_id, user["id"], utcnow()))
            liked = True
            if post["user_id"] != user["id"]:
                insert_notification(conn, post["user_id"], "feed_like", "피드 좋아요", f"{user['nickname']}님이 회원님의 피드에 좋아요를 눌렀습니다.")
        return {"liked": liked}
@app.post("/api/feed/{post_id}/bookmark")
def toggle_bookmark(post_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        exists = conn.execute("SELECT 1 FROM feed_bookmarks WHERE post_id = ? AND user_id = ?", (post_id, user["id"])).fetchone()
        if exists:
            conn.execute("DELETE FROM feed_bookmarks WHERE post_id = ? AND user_id = ?", (post_id, user["id"]))
            bookmarked = False
        else:
            conn.execute("INSERT INTO feed_bookmarks(post_id, user_id, created_at) VALUES (?, ?, ?)", (post_id, user["id"], utcnow()))
            bookmarked = True
        return {"bookmarked": bookmarked}
@app.post("/api/feed/{post_id}/comment")
def add_comment(post_id: int, payload: CommentIn, user=Depends(require_user)):
    with get_conn() as conn:
        post = conn.execute("SELECT * FROM feed_posts WHERE id = ?", (post_id,)).fetchone()
        if not post:
            raise HTTPException(status_code=404, detail="피드를 찾을 수 없습니다.")
        conn.execute(
            "INSERT INTO feed_comments(post_id, user_id, content, created_at) VALUES (?, ?, ?, ?)",
            (post_id, user["id"], payload.content.strip(), utcnow()),
        )
        if post["user_id"] != user["id"]:
            insert_notification(conn, post["user_id"], "feed_comment", "피드 댓글", f"{user['nickname']}님이 댓글을 남겼습니다.")
        return {"ok": True}
@app.get("/api/profile")
def get_profile(user=Depends(require_user)):
    return {"user": user}
@app.put("/api/profile")
def update_profile(payload: ProfileIn, user=Depends(require_user)):
    with get_conn() as conn:
        next_login_id = _validate_login_id_value(payload.login_id or user.get('login_id') or user.get('email'))
        next_email = _normalize_email_value(payload.email or next_login_id)
        existing = _find_user_by_email_ci(conn, next_email, user["id"])
        if existing:
            raise HTTPException(status_code=400, detail="이미 사용 중인 이메일입니다.")
        dup_login_id = _find_user_by_login_id_ci(conn, next_login_id, user["id"])
        if dup_login_id:
            raise HTTPException(status_code=400, detail=f"{next_login_id} 아이디는 이미 사용 중입니다.")
        if payload.branch_no != user.get('branch_no') and int(user.get('grade') or 6) != 1:
            raise HTTPException(status_code=403, detail='호점은 관리자 권한에서만 본인 프로필로 변경할 수 있습니다.')
        assignments = [
            ("login_id", next_login_id),
            ("email", next_email),
            ("nickname", payload.nickname.strip()),
            ("region", payload.region.strip() or "서울"),
            ("bio", payload.bio.strip()),
            ("one_liner", payload.one_liner.strip()),
            ("interests", json.dumps(payload.interests, ensure_ascii=False)),
            ("photo_url", payload.photo_url.strip()),
            ("phone", payload.phone.strip()),
            ("recovery_email", payload.recovery_email.strip()),
            ("gender", _validate_gender_value(payload.gender, allow_empty=True)),
            ("birth_year", int(payload.birth_year or 1990)),
            ("vehicle_number", payload.vehicle_number.strip()),
            ("branch_no", payload.branch_no if int(user.get('grade') or 6) == 1 else user.get('branch_no')),
            ("marital_status", payload.marital_status.strip()),
            ("resident_address", payload.resident_address.strip()),
            ("business_name", payload.business_name.strip()),
            ("business_number", payload.business_number.strip()),
            ("business_type", payload.business_type.strip()),
            ("business_item", payload.business_item.strip()),
            ("business_address", payload.business_address.strip()),
            ("bank_account", payload.bank_account.strip()),
            ("bank_name", payload.bank_name.strip()),
            ("mbti", payload.mbti.strip()),
            ("google_email", payload.google_email.strip()),
            ("resident_id", payload.resident_id.strip()),
        ]
        if payload.new_password.strip():
            assignments.append(("password_hash", hash_password(payload.new_password.strip())))
        set_sql = ", ".join(f"{col} = ?" for col, _ in assignments)
        values = [value for _, value in assignments] + [user["id"]]
        conn.execute(f"UPDATE users SET {set_sql} WHERE id = ?", values)
        updated = conn.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
        return {"user": user_public_dict(updated)}
@app.post("/api/profile/location")
def update_location(payload: LocationIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute(
            "UPDATE users SET latitude = ?, longitude = ?, region = ? WHERE id = ?",
            (payload.latitude, payload.longitude, payload.region, user["id"]),
        )
        updated = conn.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
        return {"user": user_public_dict(updated)}

@app.get("/api/location-sharing/status")
def get_location_sharing_status(user=Depends(require_user)):
    with get_conn() as conn:
        fresh = conn.execute("SELECT * FROM users WHERE id = ?", (user['id'],)).fetchone()
        status = _location_share_status(conn, user_public_dict(fresh))
        return status

@app.post("/api/location-sharing/consent")
def set_location_sharing_consent(payload: LocationShareConsentIn, user=Depends(require_user)):
    with get_conn() as conn:
        now = utcnow()
        conn.execute(
            "UPDATE users SET location_share_consent = ?, location_share_enabled = ?, location_share_updated_at = ? WHERE id = ?",
            (1 if payload.enabled else 0, 1 if payload.enabled else 0, now, user['id']),
        )
        updated = conn.execute("SELECT * FROM users WHERE id = ?", (user['id'],)).fetchone()
        updated_user = user_public_dict(updated)
        return {"user": updated_user, "status": _location_share_status(conn, updated_user)}
@app.get("/api/users")
def list_users(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM users WHERE id != ? ORDER BY id", (user["id"],)).fetchall()
        return [user_public_dict(r) for r in rows]
@app.get("/api/users/{user_id}/profile")
def user_profile(user_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")
        data = user_public_dict(row)
        data["is_following"] = bool(conn.execute("SELECT 1 FROM follows WHERE from_user_id = ? AND to_user_id = ?", (user["id"], user_id)).fetchone())
        data["is_friend"] = bool(conn.execute("SELECT 1 FROM friends WHERE user_id = ? AND friend_id = ?", (user["id"], user_id)).fetchone())
        return data
@app.post("/api/follows/{target_user_id}")
def toggle_follow(target_user_id: int, user=Depends(require_user)):
    if target_user_id == user["id"]:
        raise HTTPException(status_code=400, detail="본인을 팔로우할 수 없습니다.")
    with get_conn() as conn:
        if is_blocked(conn, user["id"], target_user_id):
            raise HTTPException(status_code=400, detail="차단 관계에서는 팔로우할 수 없습니다.")
        exists = conn.execute("SELECT 1 FROM follows WHERE from_user_id = ? AND to_user_id = ?", (user["id"], target_user_id)).fetchone()
        if exists:
            conn.execute("DELETE FROM follows WHERE from_user_id = ? AND to_user_id = ?", (user["id"], target_user_id))
            following = False
        else:
            conn.execute("INSERT INTO follows(from_user_id, to_user_id, created_at) VALUES (?, ?, ?)", (user["id"], target_user_id, utcnow()))
            following = True
            insert_notification(conn, target_user_id, "follow", "새 팔로우", f"{user['nickname']}님이 회원님을 팔로우했습니다.")
        return {"following": following}
@app.get("/api/follows")
def get_follows(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT u.* FROM follows f
            JOIN users u ON u.id = f.to_user_id
            WHERE f.from_user_id = ?
            ORDER BY u.nickname
            """,
            (user["id"],),
        ).fetchall()
        return [user_public_dict(r) for r in rows]
@app.get("/api/friends")
def get_friends(user=Depends(require_user)):
    with get_conn() as conn:
        friends = conn.execute(
            """
            SELECT u.* FROM friends f
            JOIN users u ON u.id = f.friend_id
            WHERE f.user_id = ?
            ORDER BY u.nickname
            """,
            (user["id"],),
        ).fetchall()
        inbound = conn.execute(
            """
            SELECT fr.*, u.nickname AS requester_nickname FROM friend_requests fr
            JOIN users u ON u.id = fr.requester_id
            WHERE fr.target_user_id = ? AND fr.status = 'pending'
            ORDER BY fr.id DESC
            """,
            (user["id"],),
        ).fetchall()
        outbound = conn.execute(
            """
            SELECT fr.*, u.nickname AS target_nickname FROM friend_requests fr
            JOIN users u ON u.id = fr.target_user_id
            WHERE fr.requester_id = ? AND fr.status = 'pending'
            ORDER BY fr.id DESC
            """,
            (user["id"],),
        ).fetchall()
        return {
            "friends": [user_public_dict(r) for r in friends],
            "received_requests": [row_to_dict(r) for r in inbound],
            "sent_requests": [row_to_dict(r) for r in outbound],
        }
@app.post("/api/friends/request/{target_user_id}")
def request_friend(target_user_id: int, user=Depends(require_user)):
    if target_user_id == user["id"]:
        raise HTTPException(status_code=400, detail="본인에게 요청할 수 없습니다.")
    with get_conn() as conn:
        if is_blocked(conn, user["id"], target_user_id):
            raise HTTPException(status_code=400, detail="차단 관계에서는 친구 요청이 불가합니다.")
        exists = conn.execute("SELECT 1 FROM friends WHERE user_id = ? AND friend_id = ?", (user["id"], target_user_id)).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="이미 친구입니다.")
        req = conn.execute("SELECT * FROM friend_requests WHERE requester_id = ? AND target_user_id = ?", (user["id"], target_user_id)).fetchone()
        if req:
            raise HTTPException(status_code=400, detail="이미 요청했습니다.")
        conn.execute(
            "INSERT INTO friend_requests(requester_id, target_user_id, status, created_at) VALUES (?, ?, 'pending', ?)",
            (user["id"], target_user_id, utcnow()),
        )
        actor_row = conn.execute(
            "SELECT id, login_id, email, name, nickname, grade, position_title, branch_no, branch_code FROM users WHERE id = ?",
            (user["id"],),
        ).fetchone()
        insert_notification(conn, target_user_id, "friend_request", "친구 요청", _friend_notification_body(actor_row, "친구 요청을 보냈습니다."))
        return {"ok": True}
@app.post("/api/friends/respond/{request_id}")
def respond_friend(request_id: int, payload: FriendRespondIn, user=Depends(require_user)):
    with get_conn() as conn:
        req = conn.execute("SELECT * FROM friend_requests WHERE id = ? AND target_user_id = ?", (request_id, user["id"])).fetchone()
        if not req:
            raise HTTPException(status_code=404, detail="친구 요청을 찾을 수 없습니다.")
        action = payload.action.lower()
        if action not in {"accepted", "rejected"}:
            raise HTTPException(status_code=400, detail="action 값은 accepted 또는 rejected 이어야 합니다.")
        conn.execute("UPDATE friend_requests SET status = ?, responded_at = ? WHERE id = ?", (action, utcnow(), request_id))
        if action == "accepted":
            conn.execute("INSERT OR IGNORE INTO friends(user_id, friend_id, created_at) VALUES (?, ?, ?)", (req["requester_id"], req["target_user_id"], utcnow()))
            conn.execute("INSERT OR IGNORE INTO friends(user_id, friend_id, created_at) VALUES (?, ?, ?)", (req["target_user_id"], req["requester_id"], utcnow()))
            actor_row = conn.execute(
                "SELECT id, login_id, email, name, nickname, grade, position_title, branch_no, branch_code FROM users WHERE id = ?",
                (user["id"],),
            ).fetchone()
            insert_notification(conn, req["requester_id"], "friend_accept", "친구 요청 수락", _friend_notification_body(actor_row, "친구 요청을 수락했습니다."))
        return {"ok": True, "status": action}
@app.post("/api/direct-chat-requests/{target_user_id}")
def direct_chat_request(target_user_id: int, user=Depends(require_user)):
    if target_user_id == user["id"]:
        raise HTTPException(status_code=400, detail="본인에게 요청할 수 없습니다.")
    with get_conn() as conn:
        if is_blocked(conn, user["id"], target_user_id):
            raise HTTPException(status_code=400, detail="차단 관계에서는 채팅 요청이 불가합니다.")
        req = conn.execute("SELECT * FROM direct_chat_requests WHERE requester_id = ? AND target_user_id = ?", (user["id"], target_user_id)).fetchone()
        if req:
            conn.execute("UPDATE direct_chat_requests SET status = 'pending', responded_at = '' WHERE id = ?", (req["id"],))
        else:
            conn.execute(
                "INSERT INTO direct_chat_requests(requester_id, target_user_id, status, created_at) VALUES (?, ?, 'pending', ?)",
                (user["id"], target_user_id, utcnow()),
            )
        insert_notification(conn, target_user_id, "direct_chat_request", "채팅 요청", f"{user['nickname']}님이 1:1 채팅을 요청했습니다.")
        return {"ok": True}
@app.get("/api/chat/{target_user_id}")
def get_direct_chat(target_user_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        if is_blocked(conn, user["id"], target_user_id):
            raise HTTPException(status_code=400, detail="차단 관계에서는 채팅을 볼 수 없습니다.")
        key = room_key(user["id"], target_user_id)
        rows = conn.execute("SELECT * FROM dm_messages WHERE room_key = ? ORDER BY id", (key,)).fetchall()
        room_ref = str(target_user_id)
        pending_mentions = conn.execute(
            "SELECT * FROM chat_mentions WHERE user_id = ? AND room_type = 'direct' AND room_ref = ? AND seen = 0 ORDER BY id",
            (user["id"], room_ref),
        ).fetchall()
        return {
            "room_key": key,
            "messages": [enrich_chat_message(conn, r, 'dm_messages') for r in rows],
            "target_user": user_basic(conn, target_user_id),
            "room_setting": get_room_setting(conn, user["id"], 'direct', room_ref),
            "pending_mentions": [{**row_to_dict(r), "sender": user_basic(conn, r["sender_id"])} for r in pending_mentions],
        }
@app.post("/api/chat/{target_user_id}")
def send_direct_chat(target_user_id: int, payload: MessageIn, user=Depends(require_user)):
    has_text = bool(payload.message.strip())
    has_attachment = bool(payload.attachment_url or payload.attachment_name)
    if not has_text and not has_attachment:
        raise HTTPException(status_code=400, detail="메시지를 입력해 주세요.")
    with get_conn() as conn:
        if is_blocked(conn, user["id"], target_user_id):
            raise HTTPException(status_code=400, detail="차단 관계에서는 채팅할 수 없습니다.")
        key = room_key(user["id"], target_user_id)
        message_text = payload.message.strip()
        if not message_text and payload.attachment_type == 'image':
            message_text = '사진을 보냈습니다.'
        elif not message_text and payload.attachment_type == 'file':
            message_text = '파일을 보냈습니다.'
        elif not message_text and payload.attachment_type == 'location':
            message_text = '위치를 공유했습니다.'
        conn.execute(
            """
            INSERT INTO dm_messages(room_key, sender_id, message, attachment_name, attachment_url, attachment_type, reply_to_id, mention_user_id, reactions, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (key, user["id"], message_text, payload.attachment_name, payload.attachment_url, payload.attachment_type, payload.reply_to_id, payload.mention_user_id, '[]', utcnow()),
        )
        message_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        insert_notification(conn, target_user_id, "direct_chat", "새 채팅", f"{user['nickname']}님이 새 메시지를 보냈습니다.")
        if payload.mention_user_id and payload.mention_user_id != user['id']:
            insert_notification(conn, payload.mention_user_id, "chat_mention", "채팅 태그", f"{user['nickname']}님이 나를 태그했습니다.")
            insert_chat_mention(conn, payload.mention_user_id, 'direct', str(user['id']), message_id, user['id'])
        return {"ok": True, "message_id": message_id}
@app.get("/api/chat/{target_user_id}/voice-room")
def get_direct_voice_room(target_user_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT * FROM voice_rooms
            WHERE room_type = 'direct' AND status = 'active'
              AND ((creator_id = ? AND target_user_id = ?) OR (creator_id = ? AND target_user_id = ?))
            ORDER BY id DESC LIMIT 1
            """,
            (user["id"], target_user_id, target_user_id, user["id"]),
        ).fetchone()
        return row_to_dict(row) if row else {"room": None}
@app.post("/api/chat/{target_user_id}/voice-room")
def create_direct_voice_room(target_user_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        if is_blocked(conn, user["id"], target_user_id):
            raise HTTPException(status_code=400, detail="차단 관계에서는 음성방을 만들 수 없습니다.")
        conn.execute(
            "INSERT INTO voice_rooms(room_type, creator_id, target_user_id, status, created_at) VALUES ('direct', ?, ?, 'active', ?)",
            (user["id"], target_user_id, utcnow()),
        )
        room_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        insert_notification(conn, target_user_id, "voice_invite", "음성통화 초대", f"{user['nickname']}님이 음성 통화를 시작했습니다.")
        return {"room_id": room_id}
@app.get("/api/group-rooms")
def list_group_rooms(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT gr.*,
                   (SELECT COUNT(*) FROM group_room_members m WHERE m.room_id = gr.id) AS member_count
            FROM group_rooms gr
            ORDER BY gr.id DESC
            """
        ).fetchall()
        return [{**row_to_dict(r), "creator": user_basic(conn, r["creator_id"])} for r in rows]
@app.post("/api/group-rooms")
def create_group_room(payload: GroupRoomIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO group_rooms(title, description, region, creator_id, created_at) VALUES (?, ?, ?, ?, ?)",
            (payload.title, payload.description, payload.region, user["id"], utcnow()),
        )
        room_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        member_ids = {int(user["id"])}
        for value in payload.member_ids or []:
            try:
                member_ids.add(int(value))
            except Exception:
                continue
        for member_id in member_ids:
            conn.execute("INSERT OR IGNORE INTO group_room_members(room_id, user_id, created_at) VALUES (?, ?, ?)", (room_id, member_id, utcnow()))
            if member_id != int(user["id"]):
                insert_notification(conn, member_id, 'group_invite', '단체방 초대', f"{user['nickname']}님이 단체방에 초대했습니다.")
        return {"room_id": room_id}
@app.post("/api/group-rooms/{room_id}/join")
def join_group_room(room_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        room = conn.execute("SELECT * FROM group_rooms WHERE id = ?", (room_id,)).fetchone()
        if not room:
            raise HTTPException(status_code=404, detail="그룹방을 찾을 수 없습니다.")
        conn.execute("INSERT OR IGNORE INTO group_room_members(room_id, user_id, created_at) VALUES (?, ?, ?)", (room_id, user["id"], utcnow()))
        return {"ok": True}
@app.get("/api/group-rooms/{room_id}/messages")
def get_group_messages(room_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        member = conn.execute("SELECT 1 FROM group_room_members WHERE room_id = ? AND user_id = ?", (room_id, user["id"])).fetchone()
        if not member:
            raise HTTPException(status_code=403, detail="그룹방 참가자만 조회할 수 있습니다.")
        rows = conn.execute("SELECT * FROM group_room_messages WHERE room_id = ? ORDER BY id", (room_id,)).fetchall()
        room = conn.execute("SELECT * FROM group_rooms WHERE id = ?", (room_id,)).fetchone()
        pending_mentions = conn.execute(
            "SELECT * FROM chat_mentions WHERE user_id = ? AND room_type = 'group' AND room_ref = ? AND seen = 0 ORDER BY id",
            (user["id"], str(room_id)),
        ).fetchall()
        member_rows = conn.execute("SELECT gm.user_id, gm.created_at FROM group_room_members gm WHERE gm.room_id = ? ORDER BY gm.created_at, gm.user_id", (room_id,)).fetchall()
        members = []
        for member_row in member_rows:
            target = conn.execute("SELECT * FROM users WHERE id = ?", (member_row["user_id"],)).fetchone()
            if target:
                members.append(user_public_dict(target))
        return {
            "room": {**row_to_dict(room), "creator": user_basic(conn, room["creator_id"]), "can_manage": can_manage_group_room(user)},
            "members": members,
            "messages": [enrich_chat_message(conn, r, 'group_room_messages') for r in rows],
            "room_setting": get_room_setting(conn, user["id"], 'group', str(room_id)),
            "pending_mentions": [{**row_to_dict(r), "sender": user_basic(conn, r["sender_id"])} for r in pending_mentions],
        }
@app.post("/api/group-rooms/{room_id}/messages")
def send_group_message(room_id: int, payload: MessageIn, user=Depends(require_user)):
    has_text = bool(payload.message.strip())
    has_attachment = bool(payload.attachment_url or payload.attachment_name)
    if not has_text and not has_attachment:
        raise HTTPException(status_code=400, detail="메시지를 입력해 주세요.")
    with get_conn() as conn:
        member = conn.execute("SELECT 1 FROM group_room_members WHERE room_id = ? AND user_id = ?", (room_id, user["id"])).fetchone()
        if not member:
            raise HTTPException(status_code=403, detail="그룹방 참가자만 작성할 수 있습니다.")
        message_text = payload.message.strip()
        if not message_text and payload.attachment_type == 'image':
            message_text = '사진을 보냈습니다.'
        elif not message_text and payload.attachment_type == 'file':
            message_text = '파일을 보냈습니다.'
        elif not message_text and payload.attachment_type == 'location':
            message_text = '위치를 공유했습니다.'
        conn.execute(
            "INSERT INTO group_room_messages(room_id, sender_id, message, attachment_name, attachment_url, attachment_type, reply_to_id, mention_user_id, reactions, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (room_id, user["id"], message_text, payload.attachment_name, payload.attachment_url, payload.attachment_type, payload.reply_to_id, payload.mention_user_id, '[]', utcnow()),
        )
        message_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        if payload.mention_user_id and payload.mention_user_id != user['id']:
            insert_notification(conn, payload.mention_user_id, "chat_mention", "채팅 태그", f"{user['nickname']}님이 나를 태그했습니다.")
            insert_chat_mention(conn, payload.mention_user_id, 'group', str(room_id), message_id, user['id'])
        return {"ok": True, "message_id": message_id}
@app.post("/api/voice/direct/{target_user_id}")
def start_voice_direct(target_user_id: int, user=Depends(require_user)):
    return create_direct_voice_room(target_user_id, user)
@app.post("/api/voice/group/{room_id}")
def start_voice_group(room_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        member = conn.execute("SELECT 1 FROM group_room_members WHERE room_id = ? AND user_id = ?", (room_id, user["id"])).fetchone()
        if not member:
            raise HTTPException(status_code=403, detail="그룹방 참가자만 음성방을 생성할 수 있습니다.")
        conn.execute(
            "INSERT INTO voice_rooms(room_type, creator_id, group_room_id, status, created_at) VALUES ('group', ?, ?, 'active', ?)",
            (user["id"], room_id, utcnow()),
        )
        room_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"room_id": room_id}
@app.get("/api/voice/rooms/{room_id}/signals")
def get_signals(room_id: int, user=Depends(require_user), since_id: int = Query(default=0)):
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM voice_signals WHERE room_id = ? AND id > ? ORDER BY id",
            (room_id, since_id),
        ).fetchall()
        return [{**row_to_dict(r), "payload": json.loads(r["payload"])} for r in rows]
@app.post("/api/voice/rooms/{room_id}/signals")
def post_signal(room_id: int, payload: VoiceSignalIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO voice_signals(room_id, sender_id, payload, created_at) VALUES (?, ?, ?, ?)",
            (room_id, user["id"], json.dumps(payload.payload, ensure_ascii=False), utcnow()),
        )
        return {"ok": True}
@app.get("/api/chat-list")
def chat_list(category: str = Query(default="general"), user=Depends(require_user)):
    with get_conn() as conn:
        setting_rows = conn.execute("SELECT * FROM chat_room_settings WHERE user_id = ?", (user["id"],)).fetchall()
        settings = {(r["room_type"], r["room_ref"]): row_to_dict(r) for r in setting_rows}
        mention_rows = conn.execute("SELECT * FROM chat_mentions WHERE user_id = ? AND seen = 0 ORDER BY id DESC", (user["id"],)).fetchall()
        mentions = {}
        for row in mention_rows:
            key = (row["room_type"], row["room_ref"])
            if key not in mentions:
                mentions[key] = row_to_dict(row)
        items = []
        direct_rows = conn.execute("SELECT * FROM dm_messages ORDER BY id DESC").fetchall()
        seen_direct = set()
        for row in direct_rows:
            a, b = [int(v) for v in row["room_key"].split(":")]
            if user["id"] not in (a, b):
                continue
            other_id = b if a == user["id"] else a
            if other_id in seen_direct:
                continue
            seen_direct.add(other_id)
            setting = settings.get(("direct", str(other_id)), {})
            if setting.get("hidden"):
                continue
            other = user_basic(conn, other_id)
            mention = mentions.get(("direct", str(other_id)))
            subtitle = f"{user_basic(conn, mention['sender_id'])['nickname']}님이 나를 태그했습니다." if mention else room_preview_text(row_to_dict(row))
            items.append({
                "id": f"direct-{other_id}",
                "room_type": "direct",
                "room_ref": str(other_id),
                "title": setting.get("custom_name") or other["nickname"],
                "subtitle": subtitle,
                "updated_at": row["created_at"],
                "pinned": bool(setting.get("pinned", 0)),
                "favorite": bool(setting.get("favorite", 0)),
                "muted": bool(setting.get("muted", 0)),
                "target_user": other,
                "unread_tag": bool(mention),
            })
        group_rows = conn.execute(
            """
            SELECT gr.* FROM group_rooms gr
            JOIN group_room_members gm ON gm.room_id = gr.id
            WHERE gm.user_id = ?
            ORDER BY gr.id DESC
            """,
            (user["id"],),
        ).fetchall()
        for room in group_rows:
            setting = settings.get(("group", str(room["id"])), {})
            if setting.get("hidden"):
                continue
            last_row = conn.execute("SELECT * FROM group_room_messages WHERE room_id = ? ORDER BY id DESC LIMIT 1", (room["id"],)).fetchone()
            mention = mentions.get(("group", str(room["id"])))
            subtitle = f"{user_basic(conn, mention['sender_id'])['nickname']}님이 나를 태그했습니다." if mention else (room_preview_text(row_to_dict(last_row)) if last_row else room["description"])
            items.append({
                "id": f"group-{room['id']}",
                "room_type": "group",
                "room_ref": str(room["id"]),
                "title": setting.get("custom_name") or room["title"],
                "subtitle": subtitle,
                "updated_at": last_row["created_at"] if last_row else room["created_at"],
                "pinned": bool(setting.get("pinned", 0)),
                "favorite": bool(setting.get("favorite", 0)),
                "muted": bool(setting.get("muted", 0)),
                "room": {**row_to_dict(room), "creator": user_basic(conn, room["creator_id"]), "can_manage": can_manage_group_room(user)},
                "unread_tag": bool(mention),
            })
        if category == 'general':
            items = [item for item in items if item['room_type'] == 'direct']
        elif category == 'group':
            items = [item for item in items if item['room_type'] == 'group']
        elif category == 'favorite':
            items = [item for item in items if item.get('favorite')]
        items.sort(key=lambda item: (not item.get('pinned', False), item.get('updated_at', ''), item.get('id')), reverse=False)
        items.sort(key=lambda item: (0 if item.get('pinned') else 1, item.get('updated_at', '')), reverse=False)
        items = sorted(items, key=lambda item: (0 if item.get('pinned') else 1, item.get('updated_at', '')), reverse=False)
        items.reverse()
        return items
@app.put("/api/chat-rooms/direct/{target_user_id}/settings")
def update_direct_chat_room_settings(target_user_id: int, payload: ChatRoomSettingIn, user=Depends(require_user)):
    with get_conn() as conn:
        setting = save_room_setting(conn, user['id'], 'direct', str(target_user_id), payload.model_dump(exclude_none=True))
        return {"ok": True, "setting": setting}
@app.put("/api/chat-rooms/group/{room_id}/settings")
def update_group_chat_room_settings(room_id: int, payload: ChatRoomSettingIn, user=Depends(require_user)):
    with get_conn() as conn:
        setting = save_room_setting(conn, user['id'], 'group', str(room_id), payload.model_dump(exclude_none=True))
        return {"ok": True, "setting": setting}
@app.post("/api/direct-chat/{target_user_id}/invite")
def invite_from_direct_room(target_user_id: int, payload: ChatInviteIn, user=Depends(require_user)):
    with get_conn() as conn:
        target = conn.execute("SELECT * FROM users WHERE id = ?", (target_user_id,)).fetchone()
        invited = conn.execute("SELECT * FROM users WHERE id = ?", (payload.user_id,)).fetchone()
        if not target or not invited:
            raise HTTPException(status_code=404, detail="초대할 사용자를 찾을 수 없습니다.")
        title = f"{user['nickname']}·{target['nickname']} 대화방"
        conn.execute("INSERT INTO group_rooms(title, description, region, creator_id, created_at) VALUES (?, ?, ?, ?, ?)", (title, '직접 채팅방에서 초대한 단체방', user.get('region', '서울'), user['id'], utcnow()))
        room_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        for uid in {user['id'], target_user_id, payload.user_id}:
            conn.execute("INSERT OR IGNORE INTO group_room_members(room_id, user_id, created_at) VALUES (?, ?, ?)", (room_id, uid, utcnow()))
        insert_notification(conn, payload.user_id, 'group_invite', '단체방 초대', f"{user['nickname']}님이 단체방에 초대했습니다.")
        return {"ok": True, "room_id": room_id}
@app.post("/api/group-rooms/{room_id}/invite")
def invite_to_group_room(room_id: int, payload: ChatInviteIn, user=Depends(require_user)):
    with get_conn() as conn:
        member = conn.execute("SELECT 1 FROM group_room_members WHERE room_id = ? AND user_id = ?", (room_id, user['id'])).fetchone()
        if not member:
            raise HTTPException(status_code=403, detail='그룹방 참가자만 초대할 수 있습니다.')
        if not can_manage_group_room(user):
            raise HTTPException(status_code=403, detail='관리자 또는 부관리자만 초대할 수 있습니다.')
        invited = conn.execute("SELECT * FROM users WHERE id = ?", (payload.user_id,)).fetchone()
        if not invited:
            raise HTTPException(status_code=404, detail='초대할 사용자를 찾을 수 없습니다.')
        conn.execute("INSERT OR IGNORE INTO group_room_members(room_id, user_id, created_at) VALUES (?, ?, ?)", (room_id, payload.user_id, utcnow()))
        insert_notification(conn, payload.user_id, 'group_invite', '단체방 초대', f"{user['nickname']}님이 단체방에 초대했습니다.")
        return {"ok": True}

@app.get("/api/group-rooms/{room_id}/members")
def list_group_room_members(room_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        member = conn.execute("SELECT 1 FROM group_room_members WHERE room_id = ? AND user_id = ?", (room_id, user['id'])).fetchone()
        if not member:
            raise HTTPException(status_code=403, detail='그룹방 참가자만 조회할 수 있습니다.')
        rows = conn.execute("SELECT gm.user_id FROM group_room_members gm WHERE gm.room_id = ? ORDER BY gm.created_at, gm.user_id", (room_id,)).fetchall()
        members = []
        for row in rows:
            target = conn.execute("SELECT * FROM users WHERE id = ?", (row['user_id'],)).fetchone()
            if target:
                members.append(user_public_dict(target))
        return {"can_manage": can_manage_group_room(user), "members": members}

@app.delete("/api/group-rooms/{room_id}/members/{member_user_id}")
def remove_group_room_member(room_id: int, member_user_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        member = conn.execute("SELECT 1 FROM group_room_members WHERE room_id = ? AND user_id = ?", (room_id, user['id'])).fetchone()
        if not member:
            raise HTTPException(status_code=403, detail='그룹방 참가자만 관리할 수 있습니다.')
        if not can_manage_group_room(user):
            raise HTTPException(status_code=403, detail='관리자 또는 부관리자만 추방할 수 있습니다.')
        target = conn.execute("SELECT * FROM users WHERE id = ?", (member_user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail='사용자를 찾을 수 없습니다.')
        conn.execute("DELETE FROM group_room_members WHERE room_id = ? AND user_id = ?", (room_id, member_user_id))
        save_room_setting(conn, member_user_id, 'group', str(room_id), {'hidden': True})
        insert_notification(conn, member_user_id, 'group_kick', '단체방 추방', f"{user['nickname']}님이 단체방에서 내보냈습니다.")
        return {"ok": True}

@app.post("/api/group-rooms/{room_id}/leave")
def leave_group_room(room_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute("DELETE FROM group_room_members WHERE room_id = ? AND user_id = ?", (room_id, user['id']))
        save_room_setting(conn, user['id'], 'group', str(room_id), {'hidden': True})
        return {"ok": True}
@app.get("/api/chat-mentions")
def list_chat_mentions(room_type: Optional[str] = None, room_ref: Optional[str] = None, user=Depends(require_user)):
    with get_conn() as conn:
        query = "SELECT * FROM chat_mentions WHERE user_id = ? AND seen = 0"
        params = [user['id']]
        if room_type:
            query += " AND room_type = ?"
            params.append(room_type)
        if room_ref:
            query += " AND room_ref = ?"
            params.append(room_ref)
        query += " ORDER BY id"
        rows = conn.execute(query, tuple(params)).fetchall()
        return [{**row_to_dict(r), 'sender': user_basic(conn, r['sender_id'])} for r in rows]
@app.post("/api/chat-mentions/{mention_id}/seen")
def mark_chat_mention_seen(mention_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute("UPDATE chat_mentions SET seen = 1 WHERE id = ? AND user_id = ?", (mention_id, user['id']))
        return {"ok": True}
@app.post("/api/dm-messages/{message_id}/reactions")
def react_direct_message(message_id: int, payload: ReactionIn, user=Depends(require_user)):
    with get_conn() as conn:
        toggle_reaction(conn, 'dm_messages', message_id, user['id'], payload.emoji)
        row = conn.execute("SELECT * FROM dm_messages WHERE id = ?", (message_id,)).fetchone()
        return enrich_chat_message(conn, row, 'dm_messages')
@app.post("/api/group-messages/{message_id}/reactions")
def react_group_message(message_id: int, payload: ReactionIn, user=Depends(require_user)):
    with get_conn() as conn:
        toggle_reaction(conn, 'group_room_messages', message_id, user['id'], payload.emoji)
        row = conn.execute("SELECT * FROM group_room_messages WHERE id = ?", (message_id,)).fetchone()
        return enrich_chat_message(conn, row, 'group_room_messages')
@app.get("/api/home/upcoming-schedules")
def home_upcoming_schedules(days: int = Query(default=5, ge=1, le=31), user=Depends(require_user)):
    start_date = datetime.now().date()
    end_date = start_date + timedelta(days=days - 1)
    with get_conn() as conn:
        items = _assigned_schedule_items(conn, user, start_date, end_date)
    grouped = []
    for key in sorted({item['schedule_date'] for item in items}):
        same_day = [item for item in items if item['schedule_date'] == key]
        grouped.append({
            'date': key,
            'label': datetime.strptime(key, '%Y-%m-%d').strftime('%m월 %d일'),
            'items': [
                {
                    'time_text': item['time_text'] or '미정',
                    'customer_name': item['customer_name'] or '-',
                    'representative_text': item['representative_text'] or '-',
                    'staff_text': item['staff_text'] or '-',
                    'start_address': item['start_address'] or '-',
                    'source': item['source'],
                } for item in same_day
            ],
        })
    return {'days': grouped}


@app.get('/api/workday/status')
def get_workday_status(user=Depends(require_user)):
    today = datetime.now().strftime('%Y-%m-%d')
    with get_conn() as conn:
        active_row = conn.execute(
            "SELECT * FROM workday_logs WHERE user_id = ? AND work_date = ? AND COALESCE(start_time, '') <> '' AND COALESCE(end_time, '') = '' ORDER BY id DESC LIMIT 1",
            (user['id'], today),
        ).fetchone()
        latest_row = conn.execute(
            "SELECT * FROM workday_logs WHERE user_id = ? AND work_date = ? ORDER BY id DESC LIMIT 1",
            (user['id'], today),
        ).fetchone()
        if not active_row and not latest_row:
            return {'active': False, 'today': None}
        data = row_to_dict(active_row or latest_row)
        return {
            'active': bool(active_row),
            'today': data,
        }

@app.get('/api/workday/logs')
def get_workday_logs(limit: int = Query(default=60, ge=1, le=365), user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM workday_logs WHERE user_id = ? ORDER BY work_date DESC, id DESC LIMIT ?",
            (user['id'], int(limit)),
        ).fetchall()
        return {'items': [row_to_dict(row) for row in rows]}

@app.post('/api/workday/toggle')
def toggle_workday(payload: WorkdayToggleIn, user=Depends(require_user)):
    now_dt = datetime.now()
    today = now_dt.strftime('%Y-%m-%d')
    current_time = now_dt.strftime('%H:%M')
    action = str(payload.action or '').strip().lower()
    if action not in {'start', 'end'}:
        raise HTTPException(status_code=400, detail='action 값은 start 또는 end 이어야 합니다.')
    with get_conn() as conn:
        open_row = conn.execute(
            "SELECT * FROM workday_logs WHERE user_id = ? AND work_date = ? AND COALESCE(start_time, '') <> '' AND COALESCE(end_time, '') = '' ORDER BY id DESC LIMIT 1",
            (user['id'], today),
        ).fetchone()
        if action == 'start':
            if open_row:
                raise HTTPException(status_code=409, detail='진행중인 일시작 기록이 있습니다. 먼저 일종료를 눌러주세요.')
            now = utcnow()
            conn.execute(
                "INSERT INTO workday_logs(user_id, work_date, start_time, end_time, started_at, ended_at, created_at, updated_at) VALUES (?, ?, ?, '', ?, '', ?, ?)",
                (user['id'], today, current_time, now, now, now),
            )
            updated = conn.execute(
                "SELECT * FROM workday_logs WHERE user_id = ? AND work_date = ? ORDER BY id DESC LIMIT 1",
                (user['id'], today),
            ).fetchone()
        else:
            if not open_row:
                raise HTTPException(status_code=409, detail='진행중인 일시작 기록이 없어 종료할 수 없습니다.')
            now = utcnow()
            conn.execute("UPDATE workday_logs SET end_time = ?, ended_at = ?, updated_at = ? WHERE id = ?", (current_time, now, now, open_row['id']))
            updated = conn.execute("SELECT * FROM workday_logs WHERE id = ?", (open_row['id'],)).fetchone()
        return {'ok': True, 'item': row_to_dict(updated)}

@app.get("/api/geocode")
@app.get("/api/geocode/")
def geocode_address(address: str = Query(..., min_length=2), user=Depends(require_user)):
    normalized = str(address or '').strip()
    if not normalized:
        raise HTTPException(status_code=400, detail='주소를 입력해 주세요.')
    route_normalized = _normalize_route_address(normalized)
    now_ts = time.time()
    cached = GEOCODE_CACHE.get(route_normalized)
    if cached and (now_ts - float(cached.get('stored_at') or 0)) < GEOCODE_CACHE_TTL_SECONDS:
        return {
            'lat': cached['lat'],
            'lng': cached['lng'],
            'label': route_normalized,
            'cached': True,
            'approximate': bool(cached.get('approximate')),
            'provider': str(cached.get('provider') or 'cache'),
        }

    try:
        kakao_point = _lookup_kakao_geocode(route_normalized)
        if kakao_point:
            GEOCODE_CACHE[route_normalized] = {
                'lat': float(kakao_point['lat']),
                'lng': float(kakao_point['lng']),
                'stored_at': now_ts,
                'approximate': False,
                'provider': 'kakao-local',
            }
            return kakao_point
    except Exception as exc:
        logger.warning('kakao geocode lookup failed for %s -> %s: %s', normalized, route_normalized, exc)

    try:
        naver_point = _lookup_naver_geocode(route_normalized)
        if naver_point:
            GEOCODE_CACHE[route_normalized] = {
                'lat': float(naver_point['lat']),
                'lng': float(naver_point['lng']),
                'stored_at': now_ts,
                'approximate': False,
                'provider': 'naver-geocode',
            }
            return naver_point
    except Exception as exc:
        logger.warning('naver geocode lookup failed for %s -> %s: %s', normalized, route_normalized, exc)

    url = 'https://nominatim.openstreetmap.org/search?' + urllib.parse.urlencode({
        'format': 'jsonv2',
        'limit': 1,
        'countrycodes': 'kr',
        'q': route_normalized,
    })
    request = urllib.request.Request(url, headers={
        'User-Agent': 'icj2424app-backend/1.0 (contact: admin@icj2424app.com)',
        'Accept': 'application/json',
    })
    try:
        with urllib.request.urlopen(request, timeout=8) as response:
            payload = json.loads(response.read().decode('utf-8'))
        first = payload[0] if isinstance(payload, list) and payload else None
        if first:
            point = {
                'lat': float(first.get('lat')),
                'lng': float(first.get('lon')),
                'stored_at': now_ts,
                'approximate': False,
                'provider': 'nominatim',
            }
            GEOCODE_CACHE[route_normalized] = point
            return {
                'lat': point['lat'],
                'lng': point['lng'],
                'label': route_normalized,
                'cached': False,
                'approximate': False,
                'provider': 'nominatim',
            }
    except Exception as exc:
        logger.warning('geocode lookup failed for %s -> %s: %s', normalized, route_normalized, exc)
    fallback = _derive_fallback_geocode(route_normalized)
    if fallback:
        GEOCODE_CACHE[route_normalized] = {
            'lat': float(fallback['lat']),
            'lng': float(fallback['lng']),
            'stored_at': now_ts,
            'approximate': True,
            'provider': 'fallback',
        }
        return {**fallback, 'provider': 'fallback'}
    raise HTTPException(status_code=404, detail='주소 좌표를 찾을 수 없습니다.')


def _resolve_geocode_point(address: str) -> dict[str, Any]:
    raw = str(address or '').strip()
    if not raw:
        raise HTTPException(status_code=400, detail='주소를 입력해 주세요.')
    normalized = _normalize_route_address(raw)
    result = geocode_address(normalized)
    return {
        'lat': float(result['lat']),
        'lng': float(result['lng']),
        'label': str(result.get('label') or normalized),
        'input_label': raw,
        'normalized_label': normalized,
        'provider': str(result.get('provider') or ''),
        'approximate': bool(result.get('approximate')),
    }


def _resolve_travel_geocode_point(address: str) -> dict[str, Any]:
    raw = str(address or '').strip()
    if not raw:
        raise HTTPException(status_code=400, detail='주소를 입력해 주세요.')
    normalized = _normalize_route_address(raw)
    now_ts = time.time()
    cached = GEOCODE_CACHE.get(normalized)
    cached_provider = str((cached or {}).get('provider') or '').strip().lower()
    if cached and cached_provider in {'kakao-local', 'naver-geocode'} and (now_ts - float(cached.get('stored_at') or 0)) < GEOCODE_CACHE_TTL_SECONDS:
        return {
            'lat': float(cached['lat']),
            'lng': float(cached['lng']),
            'label': normalized,
            'input_label': raw,
            'normalized_label': normalized,
            'provider': cached_provider,
            'approximate': False,
        }

    for provider_name, resolver in (('kakao-local', _lookup_kakao_geocode), ('naver-geocode', _lookup_naver_geocode)):
        try:
            point = resolver(normalized)
            if point:
                GEOCODE_CACHE[normalized] = {
                    'lat': float(point['lat']),
                    'lng': float(point['lng']),
                    'stored_at': now_ts,
                    'approximate': False,
                    'provider': provider_name,
                }
                return {
                    'lat': float(point['lat']),
                    'lng': float(point['lng']),
                    'label': str(point.get('label') or normalized),
                    'input_label': raw,
                    'normalized_label': normalized,
                    'provider': provider_name,
                    'approximate': False,
                }
        except Exception as exc:
            logger.warning('travel geocode lookup failed via %s for %s -> %s: %s', provider_name, raw, normalized, exc)

    raise HTTPException(status_code=404, detail='카카오맵 또는 네이버지도로 주소 좌표를 찾을 수 없습니다.')




def _normalize_route_address(address: str) -> str:
    raw = str(address or '').strip()
    if not raw:
        return ''

    normalized = raw.replace('\r', '\n')
    normalized = re.sub(r'\s*\n\s*', '\n', normalized)
    parts = [part.strip() for part in normalized.split('\n') if part.strip()]
    joined = ' '.join(parts)
    joined = re.sub(r'\s+', ' ', joined).strip()

    # Keep the road/jibun core address, but remove detailed residence info that often
    # shifts geocoding away from the representative building address.
    detail_patterns = [
        r'\b\d{1,3}층\b',
        r'\b지하\s*\d{1,2}층\b',
        r'\b옥탑\b',
        r'\b\d{1,4}호\b',
        r'\b\d{1,4}동\b',
        r'\b[A-Za-z]동\b',
        r'\b[A-Za-z]호\b',
    ]
    for pattern in detail_patterns:
        joined = re.sub(pattern, ' ', joined)

    # Remove trailing parenthetical unit/building details while preserving the core address.
    joined = re.sub(r'\(([^)]*(동|호|층)[^)]*)\)', ' ', joined)

    # If apartment/building name appears after a complete numeric address, trim the tail.
    tail_cut = re.search(r'^(.*?\d)(?:\s+[가-힣A-Za-z][^,]*)$', joined)
    if tail_cut:
        candidate = re.sub(r'\s+', ' ', tail_cut.group(1)).strip()
        if re.search(r'\d', candidate):
            joined = candidate

    joined = re.sub(r'\s+', ' ', joined).strip(' ,')
    return joined or raw

def _haversine_distance_m(start_lat: float, start_lng: float, end_lat: float, end_lng: float) -> float:
    radius_m = 6371000.0
    lat1 = math.radians(start_lat)
    lat2 = math.radians(end_lat)
    dlat = math.radians(end_lat - start_lat)
    dlng = math.radians(end_lng - start_lng)
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlng / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radius_m * c


def _format_duration_label(total_seconds: int) -> str:
    seconds = max(0, int(total_seconds or 0))
    hours, remainder = divmod(seconds, 3600)
    minutes = remainder // 60
    if hours and minutes:
        return f'{hours}시간 {minutes}분'
    if hours:
        return f'{hours}시간'
    if minutes:
        return f'{minutes}분'
    return '1분 미만'


def _fetch_json_request(url: str, headers: dict[str, str], timeout: int = 8) -> Any:
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=timeout) as response:
        charset = response.headers.get_content_charset() or 'utf-8'
        return json.loads(response.read().decode(charset))


def _lookup_kakao_geocode(address: str) -> dict[str, Any] | None:
    api_key = str(os.getenv('KAKAO_REST_API_KEY') or os.getenv('KAKAO_MOBILITY_REST_API_KEY') or '').strip()
    if not api_key:
        return None
    url = 'https://dapi.kakao.com/v2/local/search/address.json?' + urllib.parse.urlencode({
        'query': str(address or '').strip(),
        'analyze_type': 'similar',
        'size': 1,
    })
    payload = _fetch_json_request(url, {
        'Authorization': f'KakaoAK {api_key}',
        'Accept': 'application/json',
    })
    documents = payload.get('documents') if isinstance(payload, dict) else None
    first = documents[0] if isinstance(documents, list) and documents else None
    if not first:
        return None
    road = first.get('road_address') if isinstance(first.get('road_address'), dict) else None
    address_obj = first.get('address') if isinstance(first.get('address'), dict) else None
    target = road or address_obj or first
    x = target.get('x')
    y = target.get('y')
    if x in (None, '') or y in (None, ''):
        return None
    return {
        'lat': float(y),
        'lng': float(x),
        'label': str((road or address_obj or {}).get('address_name') or address or '').strip(),
        'cached': False,
        'approximate': False,
        'provider': 'kakao-local',
    }


def _is_route_duration_plausible(distance_m: int, duration_seconds: int) -> bool:
    distance_m = int(distance_m or 0)
    duration_seconds = int(duration_seconds or 0)
    if distance_m <= 0 or duration_seconds <= 0:
        return False
    speed_kmh = (distance_m / max(duration_seconds, 1)) * 3.6
    if speed_kmh > 130:
        return False
    if distance_m >= 3000 and speed_kmh < 3:
        return False
    return True


def _lookup_kakao_travel(start_point: dict[str, Any], end_point: dict[str, Any]) -> dict[str, Any] | None:
    api_key = str(os.getenv('KAKAO_MOBILITY_REST_API_KEY') or os.getenv('KAKAO_REST_API_KEY') or '').strip()
    if not api_key:
        return None
    url = 'https://apis-navi.kakaomobility.com/v1/directions?' + urllib.parse.urlencode({
        'origin': f"{start_point['lng']},{start_point['lat']}",
        'destination': f"{end_point['lng']},{end_point['lat']}",
        'priority': 'RECOMMEND',
    })
    payload = _fetch_json_request(url, {
        'Authorization': f'KakaoAK {api_key}',
        'Accept': 'application/json',
    })
    routes = payload.get('routes') if isinstance(payload, dict) else None
    summary = routes[0].get('summary') if isinstance(routes, list) and routes else None
    # Kakao Mobility Directions summary.duration is already returned in seconds.
    duration_seconds = int(round(float(summary.get('duration') or 0))) if summary else 0
    distance_m = int(summary.get('distance') or 0) if summary else 0
    if not _is_route_duration_plausible(distance_m, duration_seconds):
        return None
    return {
        'provider': 'kakao',
        'distance_m': distance_m,
        'duration_seconds': duration_seconds,
        'duration_text': _format_duration_label(duration_seconds),
        'approximate': bool(start_point.get('approximate') or end_point.get('approximate')),
    }


def _naver_maps_credentials() -> tuple[str, str]:
    client_id = str(
        os.getenv('NAVER_MAPS_CLIENT_ID')
        or os.getenv('NAVER_MAPS_KEY_ID')
        or os.getenv('NCP_MAPS_CLIENT_ID')
        or os.getenv('NCP_MAPS_KEY_ID')
        or ''
    ).strip()
    client_secret = str(
        os.getenv('NAVER_MAPS_CLIENT_SECRET')
        or os.getenv('NAVER_MAPS_KEY')
        or os.getenv('NCP_MAPS_CLIENT_SECRET')
        or os.getenv('NCP_MAPS_KEY')
        or ''
    ).strip()
    return client_id, client_secret


def _lookup_naver_geocode(address: str) -> dict[str, Any] | None:
    client_id, client_secret = _naver_maps_credentials()
    if not client_id or not client_secret:
        return None
    query = str(address or '').strip()
    if not query:
        return None
    url = 'https://maps.apigw.ntruss.com/map-geocode/v2/geocode?' + urllib.parse.urlencode({
        'query': query,
    })
    payload = _fetch_json_request(url, {
        'X-NCP-APIGW-API-KEY-ID': client_id,
        'X-NCP-APIGW-API-KEY': client_secret,
        'Accept': 'application/json',
    })
    addresses = payload.get('addresses') if isinstance(payload, dict) else None
    first = addresses[0] if isinstance(addresses, list) and addresses else None
    if not first:
        return None
    x = first.get('x')
    y = first.get('y')
    if x in (None, '') or y in (None, ''):
        return None
    label = str(first.get('roadAddress') or first.get('jibunAddress') or address or '').strip()
    return {
        'lat': float(y),
        'lng': float(x),
        'label': label,
        'cached': False,
        'approximate': False,
        'provider': 'naver-geocode',
    }


def _lookup_naver_travel(start_point: dict[str, Any], end_point: dict[str, Any]) -> dict[str, Any] | None:
    client_id, client_secret = _naver_maps_credentials()
    if not client_id or not client_secret:
        return None
    url = 'https://naveropenapi.apigw.ntruss.com/map-direction/v1/driving?' + urllib.parse.urlencode({
        'start': f"{start_point['lng']},{start_point['lat']}",
        'goal': f"{end_point['lng']},{end_point['lat']}",
        'option': 'trafast',
        'lang': 'ko',
    })
    payload = _fetch_json_request(url, {
        'X-NCP-APIGW-API-KEY-ID': client_id,
        'X-NCP-APIGW-API-KEY': client_secret,
        'Accept': 'application/json',
    })
    route = payload.get('route') if isinstance(payload, dict) else None
    candidates = None
    if isinstance(route, dict):
        candidates = route.get('trafast') or route.get('traoptimal')
    first = candidates[0].get('summary') if isinstance(candidates, list) and candidates else None
    duration_seconds = int(round(float(first.get('duration') or 0) / 1000.0)) if first else 0
    distance_m = int(first.get('distance') or 0) if first else 0
    if not _is_route_duration_plausible(distance_m, duration_seconds):
        return None
    return {
        'provider': 'naver',
        'distance_m': distance_m,
        'duration_seconds': duration_seconds,
        'duration_text': _format_duration_label(duration_seconds),
        'approximate': bool(start_point.get('approximate') or end_point.get('approximate')),
    }


def _estimate_travel_from_distance(start_point: dict[str, Any], end_point: dict[str, Any]) -> dict[str, Any]:
    straight_distance_m = _haversine_distance_m(start_point['lat'], start_point['lng'], end_point['lat'], end_point['lng'])
    road_distance_m = max(1000, int(round(straight_distance_m * 1.28)))
    avg_speed_mps = 28000 / 3600
    duration_seconds = max(60, int(round(road_distance_m / avg_speed_mps)))
    return {
        'provider': 'estimate',
        'distance_m': road_distance_m,
        'duration_seconds': duration_seconds,
        'duration_text': _format_duration_label(duration_seconds),
        'approximate': True,
    }


def _travel_provider_label(provider: str) -> str:
    code = str(provider or '').strip().lower()
    if code == 'kakao':
        return '카카오맵'
    if code == 'naver':
        return '네이버지도'
    if code == 'unavailable':
        return '측정불가, 직접 카카오맵 또는 네이버지도로 시간 확인'
    return '측정불가, 직접 카카오맵 또는 네이버지도로 시간 확인'


def _travel_route_mode(provider: str) -> str:
    return 'real' if str(provider or '').strip().lower() in {'kakao', 'naver'} else 'unavailable'


@app.get('/api/travel-time')
def travel_time_lookup(start_address: str = Query(..., min_length=2), end_address: str = Query(..., min_length=2), user=Depends(require_user)):
    start_point = _resolve_travel_geocode_point(start_address)
    end_point = _resolve_travel_geocode_point(end_address)
    attempts: list[str] = []
    errors: list[str] = []
    for provider_name, resolver in (('kakao', _lookup_kakao_travel), ('naver', _lookup_naver_travel)):
        attempts.append(provider_name)
        try:
            result = resolver(start_point, end_point)
            if result:
                provider_code = str(result['provider'])
                return {
                    'start': start_point,
                    'end': end_point,
                    'provider': provider_code,
                    'provider_label': _travel_provider_label(provider_code),
                    'route_mode': _travel_route_mode(provider_code),
                    'distance_m': result['distance_m'],
                    'duration_seconds': result['duration_seconds'],
                    'duration_text': result['duration_text'],
                    'approximate': bool(result.get('approximate')),
                    'attempts': attempts,
                    'start_geocode_provider': str(start_point.get('provider') or ''),
                    'end_geocode_provider': str(end_point.get('provider') or ''),
                    'normalized_start_address': start_point.get('normalized_label') or start_point.get('label') or start_address,
                    'normalized_end_address': end_point.get('normalized_label') or end_point.get('label') or end_address,
                }
        except Exception as exc:
            logger.warning('travel time lookup failed via %s: %s', provider_name, exc)
            errors.append(f'{provider_name}:{exc}')
    return {
        'start': start_point,
        'end': end_point,
        'provider': 'unavailable',
        'provider_label': _travel_provider_label('unavailable'),
        'route_mode': _travel_route_mode('unavailable'),
        'distance_m': 0,
        'duration_seconds': 0,
        'duration_text': '',
        'approximate': True,
        'attempts': attempts,
        'fallback_reason': 'real-route-unavailable',
        'errors': errors,
        'start_geocode_provider': str(start_point.get('provider') or ''),
        'end_geocode_provider': str(end_point.get('provider') or ''),
        'normalized_start_address': start_point.get('normalized_label') or start_point.get('label') or start_address,
        'normalized_end_address': end_point.get('normalized_label') or end_point.get('label') or end_address,
    }

@app.get("/api/map-users")
def map_users(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM users WHERE latitude IS NOT NULL AND longitude IS NOT NULL AND COALESCE(vehicle_number, '') != '' AND branch_no IS NOT NULL AND COALESCE(location_share_consent, 0) = 1 AND COALESCE(location_share_enabled, 0) = 1 ORDER BY branch_no, id").fetchall()
        visible = []
        today = datetime.now().date()
        now_dt = datetime.now()
        for row in rows:
            public = user_public_dict(row)
            status = _location_share_status(conn, public)
            if not status['active_now']:
                continue
            assignments = _assigned_schedule_items(conn, public, today, today)
            active_item = None
            upcoming_item = None
            for item in assignments:
                start_dt, end_dt = item.get('window_start') or (None, None)
                if start_dt and end_dt and start_dt <= now_dt <= end_dt:
                    active_item = item
                    break
                if start_dt and start_dt > now_dt and upcoming_item is None:
                    upcoming_item = item
            display_item = active_item or upcoming_item
            current_location = str(public.get('region') or public.get('resident_address') or '위치 확인중').strip()
            destination_address = str((active_item or {}).get('end_address') or '').strip()
            is_moving = bool(active_item and destination_address)
            status_text = f"현위치 {current_location}에 있고 정차 중"
            if is_moving:
                status_text = f"현위치 {current_location}에 있고, {destination_address}로 이동중"
            public['map_status'] = {
                'current_location': current_location,
                'destination_address': destination_address,
                'is_moving': is_moving,
                'status_text': status_text,
                'assignment_time': (display_item or {}).get('time_text') or '',
                'customer_name': (display_item or {}).get('customer_name') or '',
                'travel_eta_text': '',
                'estimated_arrival_text': '',
            }
            visible.append(public)
        return visible
@app.get("/api/map-region-boundaries")
def map_region_boundaries(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM region_boundaries").fetchall()
        return [{**row_to_dict(r), "geojson": json.loads(r["geojson"])} for r in rows]
@app.get("/api/meetup-schedules")
def list_meetups(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM meetup_schedules ORDER BY meetup_date, start_time").fetchall()
        return [{**row_to_dict(r), "creator": user_basic(conn, r["creator_id"])} for r in rows]
@app.post("/api/meetup-schedules")
def create_meetup(payload: MeetupIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO meetup_schedules(creator_id, title, place, meetup_date, start_time, end_time, content, cautions, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (user["id"], payload.title, payload.place, payload.meetup_date, payload.start_time, payload.end_time, payload.content, payload.cautions, payload.notes, utcnow()),
        )
        return {"ok": True}
@app.put("/api/meetup-schedules/{schedule_id}")
def update_meetup(schedule_id: int, payload: MeetupIn, user=Depends(require_user)):
    with get_conn() as conn:
        schedule = conn.execute("SELECT * FROM meetup_schedules WHERE id = ?", (schedule_id,)).fetchone()
        if not schedule:
            raise HTTPException(status_code=404, detail="모임 일정을 찾을 수 없습니다.")
        if schedule["creator_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="작성자만 수정할 수 있습니다.")
        conn.execute(
            """
            UPDATE meetup_schedules SET title = ?, place = ?, meetup_date = ?, start_time = ?, end_time = ?, content = ?, cautions = ?, notes = ?
            WHERE id = ?
            """,
            (payload.title, payload.place, payload.meetup_date, payload.start_time, payload.end_time, payload.content, payload.cautions, payload.notes, schedule_id),
        )
        return {"ok": True}
@app.delete("/api/meetup-schedules/{schedule_id}")
def delete_meetup(schedule_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        schedule = conn.execute("SELECT * FROM meetup_schedules WHERE id = ?", (schedule_id,)).fetchone()
        if not schedule:
            raise HTTPException(status_code=404, detail="모임 일정을 찾을 수 없습니다.")
        if schedule["creator_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="작성자만 삭제할 수 있습니다.")
        conn.execute("DELETE FROM meetup_schedules WHERE id = ?", (schedule_id,))
        return {"ok": True}
@app.get("/api/meetup-reviews")
def list_meetup_reviews(user=Depends(require_user), schedule_id: Optional[int] = None):
    with get_conn() as conn:
        if schedule_id:
            rows = conn.execute("SELECT * FROM meetup_reviews WHERE schedule_id = ? ORDER BY id DESC", (schedule_id,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM meetup_reviews ORDER BY id DESC").fetchall()
        return [{**row_to_dict(r), "user": user_basic(conn, r["user_id"])} for r in rows]
@app.post("/api/meetup-reviews")
def create_meetup_review(payload: MeetupReviewIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute("INSERT INTO meetup_reviews(schedule_id, user_id, content, created_at) VALUES (?, ?, ?, ?)", (payload.schedule_id, user["id"], payload.content, utcnow()))
        return {"ok": True}
@app.get("/api/boards/{category}")
def list_board_posts(category: str, user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM board_posts WHERE category = ? ORDER BY id DESC", (category,)).fetchall()
        return [{**row_to_dict(r), "user": user_basic(conn, r["user_id"])} for r in rows]
@app.post("/api/boards/{category}")
def create_board_post(category: str, payload: BoardPostIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO board_posts(category, user_id, title, content, image_url, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (category, user["id"], payload.title, payload.content, payload.image_url, utcnow()),
        )
        return {"ok": True}
@app.get("/api/boards/{category}/{post_id}")
def get_board_post(category: str, post_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM board_posts WHERE category = ? AND id = ?", (category, post_id)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="게시글을 찾을 수 없습니다.")
        comments = conn.execute("SELECT * FROM board_comments WHERE post_id = ? ORDER BY id", (post_id,)).fetchall()
        return {
            **row_to_dict(row),
            "user": user_basic(conn, row["user_id"]),
            "comments": [{**row_to_dict(c), "user": user_basic(conn, c["user_id"])} for c in comments],
        }
@app.post("/api/boards/{category}/{post_id}/comments")
def add_board_comment(category: str, post_id: int, payload: CommentIn, user=Depends(require_user)):
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM board_posts WHERE category = ? AND id = ?", (category, post_id)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="게시글을 찾을 수 없습니다.")
        conn.execute("INSERT INTO board_comments(post_id, user_id, content, created_at) VALUES (?, ?, ?, ?)", (post_id, user["id"], payload.content.strip(), utcnow()))
        return {"ok": True}
@app.get("/api/notifications")
def list_notifications(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM notifications WHERE user_id = ? ORDER BY id DESC LIMIT 50", (user["id"],)).fetchall()
        return [row_to_dict(r) for r in rows]

@app.get("/api/notifications/unread-count")
def notifications_unread_count(user=Depends(require_user)):
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0", (user["id"],)).fetchone()[0]
        return {"count": int(count or 0)}


@app.get("/api/badges-summary")
def badges_summary(user=Depends(require_user)):
    with get_conn() as conn:
        notification_count = conn.execute("SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0", (user['id'],)).fetchone()[0] or 0
        chat_count = conn.execute(
            "SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0 AND type IN ('direct_chat', 'direct_chat_request', 'group_invite', 'chat_mention', 'work_schedule_assignment')",
            (user['id'],),
        ).fetchone()[0] or 0
        friend_request_count = conn.execute(
            "SELECT COUNT(*) FROM friend_requests WHERE target_user_id = ? AND status = 'pending'",
            (user['id'],),
        ).fetchone()[0] or 0
        return {
            'notification_count': int(notification_count),
            'chat_count': int(chat_count),
            'friend_request_count': int(friend_request_count),
            'menu_count': int(friend_request_count),
        }

@app.post("/api/notifications/{notification_id}/read")
def mark_notification_read(notification_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute("UPDATE notifications SET is_read = 1 WHERE id = ? AND user_id = ?", (notification_id, user["id"]))
        row = conn.execute("SELECT * FROM notifications WHERE id = ? AND user_id = ?", (notification_id, user["id"])).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail='알림을 찾을 수 없습니다.')
        return row_to_dict(row)
@app.get("/api/feed-like-notifications")
def feed_like_notifications(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM notifications WHERE user_id = ? AND type = 'feed_like' ORDER BY id DESC", (user["id"],)).fetchall()
        return [row_to_dict(r) for r in rows]
def _calendar_event_out(conn, row):
    item = row_to_dict(row)
    item["created_by_nickname"] = user_basic(conn, row["user_id"])["nickname"]
    item["schedule_type"] = str(item.get("schedule_type") or ('B' if int(item.get("status_b_count") or 0) > 0 else 'C' if int(item.get("status_c_count") or 0) > 0 else 'A'))
    item["status_a_count"] = int(item.get("status_a_count") or 0)
    item["status_b_count"] = int(item.get("status_b_count") or 0)
    item["status_c_count"] = int(item.get("status_c_count") or 0)
    item["image_list"] = _calendar_event_image_list(item.get('image_data'))
    item["sync_group_id"] = str(item.get('sync_group_id') or '')
    item["sync_role"] = str(item.get('sync_role') or '')
    return item
def _can_access_shared_schedule(user: dict | None) -> bool:
    try:
        grade = int((user or {}).get('grade') or 6)
    except Exception:
        return False
    if grade > 5:
        return False
    role_name = str((user or {}).get('role') or '').strip()
    position_title = str((user or {}).get('position_title') or '').strip()
    if role_name in {'general', 'other'} or position_title in {'일반', '기타'}:
        return False
    return True


def _shared_schedule_user_ids(conn) -> list[int]:
    rows = conn.execute(
        """
        SELECT id, role, position_title
        FROM users
        WHERE COALESCE(approved, 1) = 1 AND CAST(COALESCE(grade, '6') AS INTEGER) <= 5
        ORDER BY CAST(COALESCE(grade, '6') AS INTEGER), id
        """
    ).fetchall()
    output: list[int] = []
    for row in rows:
        data = row_to_dict(row)
        role_name = str(data.get('role') or '').strip()
        position_title = str(data.get('position_title') or '').strip()
        if role_name in {'general', 'other'} or position_title in {'일반', '기타'}:
            continue
        try:
            value = int(data.get('id') or 0)
        except Exception:
            continue
        if value > 0 and value not in output:
            output.append(value)
    return output


def _calendar_edit_log_summaries(before: dict[str, Any], after: dict[str, Any]) -> list[str]:
    fields = [
        ('title', '제목'), ('content', '메모'), ('event_date', '일정일자'), ('start_time', '시작시각'), ('end_time', '종료시각'),
        ('visit_time', '방문시각'), ('move_start_date', '시작일'), ('move_end_date', '종료일'), ('start_address', '출발지'), ('end_address', '도착지'),
        ('platform', '플랫폼'), ('customer_name', '고객명'), ('department_info', '부서/인원'), ('amount1', '이사금액'), ('deposit_method', '계약방법'),
        ('deposit_amount', '계약금액'), ('deposit_datetime', '예약금 입금일시'), ('reservation_name', '예약자명'), ('reservation_phone', '연락처'), ('representative1', '담당대표1'), ('representative2', '담당대표2'), ('representative3', '담당대표3'),
        ('staff1', '담당직원1'), ('staff2', '담당직원2'), ('staff3', '담당직원3'), ('image_data', '첨부파일')
    ]
    changes: list[str] = []
    for key, label in fields:
        before_value = str(before.get(key) or '').strip()
        after_value = str(after.get(key) or '').strip()
        if before_value != after_value:
            changes.append(label)
    return changes


STORAGE_SCHEDULE_DEPARTMENTS = {'짐보관이사 2인 업무', '짐보관이사 3인 이상업무'}


def _is_storage_schedule_department(value: Any) -> bool:
    return str(value or '').strip() in STORAGE_SCHEDULE_DEPARTMENTS


def _storage_schedule_group_id(event_id: int) -> str:
    return f"storage-schedule-{int(event_id)}"


def _storage_manager_name_from_event(event_data: dict[str, Any]) -> str:
    reps, _ = _calendar_assignment_names(event_data)
    if reps and reps != '-':
        return reps
    return str(event_data.get('representative1') or '').strip()


def _sync_storage_status_with_calendar_group(conn, event_data: dict[str, Any], remove_only: bool = False):
    if not event_data:
        return
    group_id = str(event_data.get('sync_group_id') or _storage_schedule_group_id(int(event_data.get('id') or 0))).strip()
    if not group_id:
        return
    state = get_storage_status_state(conn)
    rows = list((state or {}).get('rows') or [])
    matched = [row for row in rows if str((row or {}).get('source_group_id') or '').strip() == group_id]
    preserved_scale = str((matched[0] or {}).get('scale') or '').strip() if matched else ''
    next_rows = [row for row in rows if str((row or {}).get('source_group_id') or '').strip() != group_id]
    if not remove_only and _is_storage_schedule_department(event_data.get('department_info')):
        next_rows.append({
            'id': f'storage-row-{group_id}',
            'customer_name': str(event_data.get('customer_name') or '').strip(),
            'manager_name': _storage_manager_name_from_event(event_data),
            'start_date': str(event_data.get('move_start_date') or event_data.get('event_date') or '').strip(),
            'end_date': str(event_data.get('move_end_date') or event_data.get('move_start_date') or event_data.get('event_date') or '').strip(),
            'scale': preserved_scale,
            'source_type': 'calendar_storage_schedule',
            'source_group_id': group_id,
            'source_event_id': int(event_data.get('id') or 0),
            'source_locked': 1,
        })
    save_storage_status_state(conn, {'rows': next_rows})


def _upsert_storage_schedule_group(conn, base_event: dict[str, Any]):
    if not base_event:
        return None
    base_id = int(base_event.get('id') or 0)
    if base_id <= 0:
        return None
    existing_group_id = str(base_event.get('sync_group_id') or '').strip()
    group_id = existing_group_id or _storage_schedule_group_id(base_id)
    primary_row = conn.execute("SELECT * FROM calendar_events WHERE sync_group_id = ? AND sync_role = 'storage_start' AND user_id = ? ORDER BY id LIMIT 1", (group_id, base_event.get('user_id'))).fetchone() if existing_group_id else None
    primary_id = int(primary_row['id']) if primary_row else base_id
    is_storage = _is_storage_schedule_department(base_event.get('department_info'))
    start_date = str(base_event.get('move_start_date') or base_event.get('event_date') or '').strip()
    end_date = str(base_event.get('move_end_date') or start_date).strip() or start_date
    conn.execute("UPDATE calendar_events SET title = ?, content = ?, event_date = ?, start_time = ?, end_time = ?, location = ?, color = ?, visit_time = ?, move_start_date = ?, move_end_date = ?, start_address = ?, end_address = ?, platform = ?, customer_name = ?, department_info = ?, schedule_type = ?, status_a_count = ?, status_b_count = ?, status_c_count = ?, amount1 = ?, amount2 = ?, amount_item = ?, deposit_method = ?, deposit_amount = ?, deposit_datetime = ?, reservation_name = ?, reservation_phone = ?, representative1 = ?, representative2 = ?, representative3 = ?, staff1 = ?, staff2 = ?, staff3 = ?, image_data = ?, sync_group_id = ?, sync_role = ? WHERE id = ?", (base_event.get('title') or '', base_event.get('content') or '', start_date or base_event.get('event_date') or '', base_event.get('start_time') or '미정', base_event.get('end_time') or '미정', base_event.get('location') or '', base_event.get('color') or '#2563eb', base_event.get('visit_time') or '미정', start_date, end_date, base_event.get('start_address') or '', base_event.get('end_address') or '', base_event.get('platform') or '', base_event.get('customer_name') or '', base_event.get('department_info') or '', base_event.get('schedule_type') or 'A', int(base_event.get('status_a_count') or 0), int(base_event.get('status_b_count') or 0), int(base_event.get('status_c_count') or 0), base_event.get('amount1') or '', base_event.get('amount2') or '', base_event.get('amount_item') or '', base_event.get('deposit_method') or '', base_event.get('deposit_amount') or '', base_event.get('deposit_datetime') or '', base_event.get('reservation_name') or '', base_event.get('reservation_phone') or '', base_event.get('representative1') or '', base_event.get('representative2') or '', base_event.get('representative3') or '', base_event.get('staff1') or '', base_event.get('staff2') or '', base_event.get('staff3') or '', base_event.get('image_data') or '', group_id if is_storage else '', 'storage_start' if is_storage else '', primary_id))
    mirror_row = conn.execute("SELECT * FROM calendar_events WHERE sync_group_id = ? AND sync_role = 'storage_end' AND user_id = ? ORDER BY id LIMIT 1", (group_id, base_event.get('user_id'))).fetchone()
    if not is_storage:
        if mirror_row:
            conn.execute("DELETE FROM calendar_events WHERE id = ?", (mirror_row['id'],))
        _sync_storage_status_with_calendar_group(conn, {**base_event, 'sync_group_id': group_id}, remove_only=True)
        return None
    if end_date and end_date != start_date:
        values = (
            base_event.get('user_id'), base_event.get('title') or '', base_event.get('content') or '', end_date,
            base_event.get('move_end_start_time') or base_event.get('start_time') or '미정',
            base_event.get('move_end_end_time') or base_event.get('end_time') or '미정',
            base_event.get('end_address') or base_event.get('location') or '',
            base_event.get('color') or '#2563eb',
            base_event.get('move_end_start_time') or base_event.get('visit_time') or '미정',
            start_date, end_date, base_event.get('start_address') or '', base_event.get('end_address') or '',
            base_event.get('platform') or '', base_event.get('customer_name') or '', base_event.get('department_info') or '', base_event.get('schedule_type') or 'A',
            int(base_event.get('status_a_count') or 0), int(base_event.get('status_b_count') or 0), int(base_event.get('status_c_count') or 0),
            base_event.get('amount1') or '', base_event.get('amount2') or '', base_event.get('amount_item') or '', base_event.get('deposit_method') or '', base_event.get('deposit_amount') or '', base_event.get('deposit_datetime') or '',
            base_event.get('reservation_name') or '', base_event.get('reservation_phone') or '',
            base_event.get('representative1') or '', base_event.get('representative2') or '', base_event.get('representative3') or '', base_event.get('staff1') or '', base_event.get('staff2') or '', base_event.get('staff3') or '',
            base_event.get('image_data') or '', group_id, 'storage_end'
        )
        if mirror_row:
            conn.execute(
                "UPDATE calendar_events SET user_id=?, title=?, content=?, event_date=?, start_time=?, end_time=?, location=?, color=?, visit_time=?, move_start_date=?, move_end_date=?, start_address=?, end_address=?, platform=?, customer_name=?, department_info=?, schedule_type=?, status_a_count=?, status_b_count=?, status_c_count=?, amount1=?, amount2=?, amount_item=?, deposit_method=?, deposit_amount=?, deposit_datetime=?, reservation_name=?, reservation_phone=?, representative1=?, representative2=?, representative3=?, staff1=?, staff2=?, staff3=?, image_data=?, sync_group_id=?, sync_role=? WHERE id = ?",
                values + (mirror_row['id'],),
            )
        else:
            conn.execute(
                "INSERT INTO calendar_events(user_id, title, content, event_date, start_time, end_time, location, color, visit_time, move_start_date, move_end_date, start_address, end_address, platform, customer_name, department_info, schedule_type, status_a_count, status_b_count, status_c_count, amount1, amount2, amount_item, deposit_method, deposit_amount, deposit_datetime, reservation_name, reservation_phone, representative1, representative2, representative3, staff1, staff2, staff3, image_data, created_at, sync_group_id, sync_role) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                values + (utcnow(),),
            )
    else:
        if mirror_row:
            conn.execute("DELETE FROM calendar_events WHERE id = ?", (mirror_row['id'],))
    next_base = conn.execute("SELECT * FROM calendar_events WHERE id = ?", (primary_id,)).fetchone()
    if next_base:
        _sync_storage_status_with_calendar_group(conn, row_to_dict(next_base), remove_only=False)
    return group_id


def _calendar_event_image_list(value: Any) -> list[str]:
    raw = value
    if isinstance(raw, list):
        return [str(item).strip() for item in raw if str(item).strip()]
    text = str(raw or '').strip()
    if not text:
        return []
    if text.startswith('['):
        try:
            data = json.loads(text)
            if isinstance(data, list):
                return [str(item).strip() for item in data if str(item).strip()]
        except Exception:
            pass
    if '\n' in text:
        return [part.strip() for part in text.splitlines() if part.strip()]
    if ',' in text and 'data:' not in text:
        parts = [part.strip() for part in text.split(',') if part.strip()]
        if len(parts) > 1:
            return parts
    return [text]

def _calendar_event_accessible(conn, event_id: int, user: dict) -> dict[str, Any] | None:
    row = conn.execute("SELECT * FROM calendar_events WHERE id = ?", (event_id,)).fetchone()
    if not row:
        return None
    item = row_to_dict(row)
    owner_id = int(item.get('user_id') or 0)
    if owner_id == int(user.get('id') or 0):
        return item
    if not _can_access_shared_schedule(user):
        return None
    owner_row = conn.execute("SELECT grade, approved FROM users WHERE id = ?", (owner_id,)).fetchone()
    if not owner_row:
        return None
    owner = row_to_dict(owner_row)
    try:
        owner_grade = int(owner.get('grade') or 6)
    except Exception:
        owner_grade = 6
    if owner_grade <= 5 and int(owner.get('approved') if owner.get('approved') is not None else 1) == 1:
        return item
    return None


@app.get("/api/calendar/events")
def get_calendar_events(start_date: str | None = None, end_date: str | None = None, user=Depends(require_user)):
    with get_conn() as conn:
        if _can_access_shared_schedule(user):
            shared_ids = _shared_schedule_user_ids(conn)
        else:
            shared_ids = [int(user["id"])]
        if not shared_ids:
            shared_ids = [int(user["id"])]
        placeholders = ','.join('?' for _ in shared_ids)
        query = f"SELECT * FROM calendar_events WHERE user_id IN ({placeholders})"
        params = list(shared_ids)
        if start_date:
            query += " AND event_date >= ?"
            params.append(start_date)
        if end_date:
            query += " AND event_date <= ?"
            params.append(end_date)
        query += " ORDER BY event_date, CASE WHEN start_time = '미정' THEN '99:99' ELSE start_time END, id"
        rows = conn.execute(query, tuple(params)).fetchall()
        return [_calendar_event_out(conn, r) for r in rows]
@app.get("/api/calendar/events/{event_id}")
def get_calendar_event(event_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        item = _calendar_event_accessible(conn, event_id, user)
        if not item:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        return _calendar_event_out(conn, item)
@app.get("/api/calendar/events/{event_id}/edit-logs")
@app.get("/api/calendar/events/{event_id}/edit-logs/")
def get_calendar_event_edit_logs(event_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        item = _calendar_event_accessible(conn, event_id, user)
        if not item:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        rows = conn.execute("SELECT * FROM calendar_event_edit_logs WHERE event_id = ? ORDER BY id DESC", (event_id,)).fetchall()
        output = []
        for row in rows:
            data = row_to_dict(row)
            actor = user_basic(conn, row['user_id']) if row['user_id'] else {'nickname': '시스템'}
            output.append({**data, 'account_name': actor.get('nickname') or actor.get('name') or '알 수 없음'})
        return output

@app.get("/api/calendar/events/{event_id}/comments")
@app.get("/api/calendar/events/{event_id}/comments/")
def get_calendar_event_comments(event_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        item = _calendar_event_accessible(conn, event_id, user)
        if not item:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        rows = conn.execute("SELECT * FROM calendar_event_comments WHERE event_id = ? ORDER BY id DESC", (event_id,)).fetchall()
        output = []
        for row in rows:
            data = row_to_dict(row)
            author = user_basic(conn, row['user_id'])
            output.append({**data, 'user': author, 'image_list': _calendar_event_image_list(data.get('image_data'))})
        return output

@app.post("/api/calendar/events/{event_id}/comments")
@app.post("/api/calendar/events/{event_id}/comments/")
def create_calendar_event_comment(event_id: int, payload: CalendarEventCommentIn, user=Depends(require_user)):
    with get_conn() as conn:
        item = _calendar_event_accessible(conn, event_id, user)
        if not item:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        content = str(payload.content or '').strip()
        image_data = str(payload.image_data or '').strip()
        if not content and not image_data:
            raise HTTPException(status_code=400, detail='댓글 내용을 입력해 주세요.')
        now_value = utcnow()
        conn.execute(
            "INSERT INTO calendar_event_comments(event_id, user_id, content, image_data, created_at) VALUES (?, ?, ?, ?, ?)",
            (event_id, user['id'], content, image_data, now_value),
        )
        return {'ok': True}

@app.put("/api/calendar/events/{event_id}/comments/{comment_id}")
@app.put("/api/calendar/events/{event_id}/comments/{comment_id}/")
def update_calendar_event_comment(event_id: int, comment_id: int, payload: CalendarEventCommentUpdateIn, user=Depends(require_user)):
    with get_conn() as conn:
        item = _calendar_event_accessible(conn, event_id, user)
        if not item:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        row = conn.execute("SELECT * FROM calendar_event_comments WHERE id = ? AND event_id = ?", (comment_id, event_id)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="댓글을 찾을 수 없습니다.")
        if int(row['user_id'] or 0) != int(user['id']) and int(user.get('grade', 9) or 9) > 1:
            raise HTTPException(status_code=403, detail="댓글 수정 권한이 없습니다.")
        content = str(payload.content or '').strip()
        image_data = str(payload.image_data or '').strip()
        if not content and not image_data:
            raise HTTPException(status_code=400, detail='댓글 내용을 입력해 주세요.')
        conn.execute("UPDATE calendar_event_comments SET content = ?, image_data = ? WHERE id = ? AND event_id = ?", (content, image_data, comment_id, event_id))
        return {'ok': True}

@app.delete("/api/calendar/events/{event_id}/comments/{comment_id}")
@app.delete("/api/calendar/events/{event_id}/comments/{comment_id}/")
def delete_calendar_event_comment(event_id: int, comment_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        item = _calendar_event_accessible(conn, event_id, user)
        if not item:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        row = conn.execute("SELECT * FROM calendar_event_comments WHERE id = ? AND event_id = ?", (comment_id, event_id)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="댓글을 찾을 수 없습니다.")
        if int(row['user_id'] or 0) != int(user['id']) and int(user.get('grade', 9) or 9) > 1:
            raise HTTPException(status_code=403, detail="댓글 삭제 권한이 없습니다.")
        conn.execute("DELETE FROM calendar_event_comments WHERE id = ? AND event_id = ?", (comment_id, event_id))
        return {'ok': True}

@app.post("/api/calendar/events")
def create_calendar_event(payload: CalendarEventIn, user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO calendar_events(
                user_id, title, content, event_date, start_time, end_time, location, color, visit_time, move_start_date, move_end_date, start_address, end_address,
                platform, customer_name, department_info, schedule_type, status_a_count, status_b_count, status_c_count, amount1, amount2, amount_item, deposit_method, deposit_amount, deposit_datetime, reservation_name, reservation_phone,
                representative1, representative2, representative3, staff1, staff2, staff3, image_data, created_at, sync_group_id, sync_role
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user["id"], payload.title, payload.content, payload.event_date, payload.start_time, payload.end_time,
                payload.location, payload.color, payload.visit_time, payload.move_start_date, payload.move_end_date, payload.start_address, payload.end_address,
                payload.platform, payload.customer_name, payload.department_info, payload.schedule_type, payload.status_a_count, payload.status_b_count, payload.status_c_count,
                payload.amount1, payload.amount2, payload.amount_item, payload.deposit_method, payload.deposit_amount, payload.deposit_datetime, payload.reservation_name, payload.reservation_phone,
                payload.representative1, payload.representative2, payload.representative3, payload.staff1, payload.staff2, payload.staff3, payload.image_data, utcnow(), '', ''
            ),
        )
        _sync_work_schedule_day_note_counts(conn, user["id"], payload.event_date)
        next_row = conn.execute("SELECT * FROM calendar_events WHERE user_id = ? ORDER BY id DESC LIMIT 1", (user["id"],)).fetchone()
        if next_row:
            next_data = row_to_dict(next_row)
            _upsert_storage_schedule_group(conn, next_data)
            next_row = conn.execute("SELECT * FROM calendar_events WHERE id = ?", (next_row['id'],)).fetchone()
            next_data = row_to_dict(next_row)
            reps, staffs = _calendar_assignment_names(next_data)
            _notify_work_schedule_assignments(
                conn,
                actor=user,
                schedule_date=next_data.get('event_date') or '',
                schedule_time=next_data.get('start_time') or next_data.get('visit_time') or '',
                customer_name=next_data.get('customer_name') or next_data.get('title') or '',
                representative_names=reps,
                staff_names=staffs,
            )
        return {"ok": True}
@app.put("/api/calendar/events/{event_id}")
def update_calendar_event(event_id: int, payload: CalendarEventIn, user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    if not _can_edit_calendar_event(user):
        raise HTTPException(status_code=403, detail="해당 직급만 일정을 수정할 수 있습니다.")
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM calendar_events WHERE id = ?", (event_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        row_data = row_to_dict(row)
        update_target_id = event_id
        group_id = str(row_data.get('sync_group_id') or '').strip()
        if group_id:
            primary_row = conn.execute("SELECT * FROM calendar_events WHERE sync_group_id = ? AND sync_role = 'storage_start' ORDER BY id LIMIT 1", (group_id,)).fetchone()
            if primary_row:
                update_target_id = int(primary_row['id'])
                row = primary_row
                row_data = row_to_dict(primary_row)
        previous_event_date = row["event_date"]
        conn.execute(
            """
            UPDATE calendar_events
            SET title = ?, content = ?, event_date = ?, start_time = ?, end_time = ?, location = ?, color = ?, visit_time = ?, move_start_date = ?, move_end_date = ?, start_address = ?, end_address = ?,
                platform = ?, customer_name = ?, department_info = ?, schedule_type = ?, status_a_count = ?, status_b_count = ?, status_c_count = ?, amount1 = ?, amount2 = ?, amount_item = ?, deposit_method = ?, deposit_amount = ?, deposit_datetime = ?, reservation_name = ?, reservation_phone = ?,
                representative1 = ?, representative2 = ?, representative3 = ?, staff1 = ?, staff2 = ?, staff3 = ?, image_data = ?
            WHERE id = ?
            """,
            (
                payload.title, payload.content, payload.event_date, payload.start_time, payload.end_time, payload.location,
                payload.color, payload.visit_time, payload.move_start_date, payload.move_end_date, payload.start_address, payload.end_address,
                payload.platform, payload.customer_name, payload.department_info, payload.schedule_type, payload.status_a_count, payload.status_b_count, payload.status_c_count,
                payload.amount1, payload.amount2, payload.amount_item, payload.deposit_method, payload.deposit_amount, payload.deposit_datetime, payload.reservation_name, payload.reservation_phone,
                payload.representative1, payload.representative2, payload.representative3, payload.staff1, payload.staff2, payload.staff3, payload.image_data, update_target_id
            ),
        )
        previous_data = row_to_dict(row)
        change_labels = _calendar_edit_log_summaries(previous_data, payload.model_dump())
        if change_labels:
            conn.execute(
                "INSERT INTO calendar_event_edit_logs(event_id, user_id, change_summary, created_at) VALUES (?, ?, ?, ?)",
                (event_id, user['id'], ', '.join(change_labels), utcnow()),
            )
        _sync_work_schedule_day_note_counts(conn, user["id"], previous_event_date)
        _sync_work_schedule_day_note_counts(conn, user["id"], payload.event_date)
        next_row = conn.execute("SELECT * FROM calendar_events WHERE id = ?", (update_target_id,)).fetchone()
        if next_row:
            _upsert_storage_schedule_group(conn, row_to_dict(next_row))
            next_row = conn.execute("SELECT * FROM calendar_events WHERE id = ?", (event_id,)).fetchone()
        if next_row:
            next_data = row_to_dict(next_row)
            reps, staffs = _calendar_assignment_names(next_data)
            _notify_work_schedule_assignments(
                conn,
                actor=user,
                schedule_date=next_data.get('event_date') or '',
                schedule_time=next_data.get('start_time') or next_data.get('visit_time') or '',
                customer_name=next_data.get('customer_name') or next_data.get('title') or '',
                representative_names=reps,
                staff_names=staffs,
                previous_ids=set(_work_assignment_target_ids(conn, *_calendar_assignment_names(previous_data), user.get('id'))),
            )
            _notify_calendar_event_changes(conn, user, previous_data, next_data)
        return {"ok": True}
@app.post("/api/calendar/events/department-replace")
def replace_calendar_event_department(payload: CalendarDepartmentReplaceIn, user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    if not _can_edit_calendar_event(user):
        raise HTTPException(status_code=403, detail="해당 직급만 부서/인원 값을 변경할 수 있습니다.")
    from_values = [str(value or '').strip() for value in (payload.from_values or []) if str(value or '').strip()]
    if not from_values:
        return {"ok": True, "updated": 0}
    to_value = str(payload.to_value or '미정').strip() or '미정'
    to_color = str(payload.to_color or '#000000').strip() or '#000000'
    placeholders = ','.join('?' for _ in from_values)
    with get_conn() as conn:
        rows = conn.execute(f"SELECT * FROM calendar_events WHERE department_info IN ({placeholders})", tuple(from_values)).fetchall()
        if not rows:
            return {"ok": True, "updated": 0}
        now_value = utcnow()
        updated_count = 0
        touched_dates = set()
        for row in rows:
            row_data = row_to_dict(row)
            touched_dates.add(str(row_data.get('event_date') or ''))
            conn.execute(
                "UPDATE calendar_events SET department_info = ?, color = ? WHERE id = ?",
                (to_value, to_color, row['id']),
            )
            conn.execute(
                "INSERT INTO calendar_event_edit_logs(event_id, user_id, change_summary, created_at) VALUES (?, ?, ?, ?)",
                (row['id'], user['id'], f"부서/인원 → {to_value}", now_value),
            )
            updated_count += 1
        for event_date in touched_dates:
            if event_date:
                _sync_work_schedule_day_note_counts(conn, user.get('id'), event_date)
        return {"ok": True, "updated": updated_count}

@app.delete("/api/calendar/events/{event_id}")
def delete_calendar_event(event_id: int, user=Depends(require_user)):
    _require_write_access(user, 'schedule')
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM calendar_events WHERE id = ? AND user_id = ?", (event_id, user["id"])).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="본인이 등록한 일정만 삭제할 수 있습니다.")
        row_data = row_to_dict(row)
        group_id = str(row_data.get('sync_group_id') or '').strip()
        touched_dates = {str(row_data.get('event_date') or '')}
        if group_id:
            group_rows = conn.execute("SELECT id, event_date FROM calendar_events WHERE user_id = ? AND sync_group_id = ?", (user['id'], group_id)).fetchall()
            for item in group_rows:
                touched_dates.add(str(item['event_date'] or ''))
            conn.execute("DELETE FROM calendar_events WHERE user_id = ? AND sync_group_id = ?", (user['id'], group_id))
            _sync_storage_status_with_calendar_group(conn, row_data, remove_only=True)
        else:
            conn.execute("DELETE FROM calendar_events WHERE id = ? AND user_id = ?", (event_id, user["id"]))
            if _is_storage_schedule_department(row_data.get('department_info')):
                _sync_storage_status_with_calendar_group(conn, {**row_data, 'sync_group_id': _storage_schedule_group_id(event_id)}, remove_only=True)
        for event_date in touched_dates:
            if event_date:
                _sync_work_schedule_day_note_counts(conn, user["id"], event_date)
        return {"ok": True}
def _schedule_day_title(base_date: date, target_date: date) -> str:
    delta = (target_date - base_date).days
    if delta == 0:
        return '당일일정'
    if delta == 1:
        return '내일일정'
    if delta == 2:
        return '모레일정'
    return f'{delta}일후일정'
def _parse_branch_exclusions(value: str) -> list[int]:
    if not value:
        return []
    tokens = []
    try:
        parsed = json.loads(value)
        if isinstance(parsed, list):
            tokens = parsed
    except Exception:
        tokens = []
    if not tokens:
        tokens = re.split(r'[\n,]', value)
    output = []
    for token in tokens:
        match = re.search(r'(\d{1,2})', str(token))
        if not match:
            continue
        branch_no = int(match.group(1))
        if 1 <= branch_no <= 50 and branch_no not in output:
            output.append(branch_no)
    return output
def _column_exists(conn, table_name: str, column_name: str) -> bool:
    try:
        if DB_ENGINE == 'postgresql':
            rows = conn.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_schema = 'public' AND table_name = ? AND column_name = ?",
                (table_name, column_name),
            ).fetchall()
            return bool(rows)
        rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
        return any(str(row[1]) == column_name for row in rows)
    except Exception:
        return False



def _get_admin_total_vehicle_count(conn):
    """일정 화면의 가용차량수는 관리자모드 > 계정권한의 차량가용여부=가용 계정 수만 사용한다."""
    vehicle_available_exists = _column_exists(conn, 'users', 'vehicle_available')
    try:
        where_clauses = ['branch_no IS NOT NULL']
        if vehicle_available_exists:
            where_clauses.append('COALESCE(vehicle_available, 1) = 1')
        where_sql = ' WHERE ' + ' AND '.join(where_clauses)
        row = conn.execute(f"SELECT COUNT(*) AS cnt FROM users{where_sql}").fetchone()
        if not row:
            return 0
        value = row['cnt'] if isinstance(row, dict) or hasattr(row, 'keys') else row[0]
        return max(int(value or 0), 0)
    except Exception:
        # 마지막 안전장치: users 전체 건수라도 반환해서 월별로 0/8이 갈라지는 문제를 막는다.
        try:
            fallback = conn.execute('SELECT COUNT(*) AS cnt FROM users').fetchone()
            value = fallback['cnt'] if isinstance(fallback, dict) or hasattr(fallback, 'keys') else fallback[0]
            return max(int(value or 0), 0)
        except Exception:
            return 0



def _sync_all_day_note_available_vehicle_counts(conn) -> None:
    auto_count = _get_admin_total_vehicle_count(conn)
    try:
        rows = conn.execute("SELECT id, schedule_date FROM work_schedule_day_notes").fetchall()
        for row in rows:
            item = row_to_dict(row)
            date_key = _normalize_date_key(item.get('schedule_date'))
            excluded_count = 0
            if date_key:
                _, auto_unavailable_map, _ = _get_vehicle_base_and_auto_unavailable(conn, [date_key])
                excluded_count = len({int(entry.get('user_id') or 0) for entry in auto_unavailable_map.get(date_key, []) if int(entry.get('user_id') or 0) > 0})
            conn.execute('UPDATE work_schedule_day_notes SET available_vehicle_count = ? WHERE id = ?', (max(auto_count - excluded_count, 0), int(item.get('id'))))
    except Exception:
        return


def _normalize_date_key(value: Any) -> str:
    raw = str(value or '').strip()
    if not raw:
        return ''
    raw = raw.split('T', 1)[0].strip()
    raw = raw.replace('.', '-').replace('/', '-').replace(' ', '')
    match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', raw)
    if not match:
        return ''
    year, month, day = match.groups()
    try:
        return date(int(year), int(month), int(day)).isoformat()
    except ValueError:
        return ''



def _vehicle_account_display_name(row: dict[str, Any] | None, default: str = '') -> str:
    item = row or {}
    display_name = str(item.get('name') or item.get('nickname') or item.get('position_title') or item.get('email') or '').strip()
    if display_name:
        return display_name
    branch_value = item.get('branch_no')
    if branch_value not in (None, ''):
        return f'{branch_value}호점'
    return default.strip() or '미지정'


SCHEDULE_EXCLUSION_REASON_PREFIX = '[스케줄열외]'


def _get_shared_schedule_note_owner_id(conn) -> int:
    row = conn.execute(
        """
        SELECT id
        FROM users
        WHERE COALESCE(approved, 1) = 1 AND grade IN (1, 2)
        ORDER BY CASE WHEN grade = 1 THEN 0 ELSE 1 END, id
        LIMIT 1
        """
    ).fetchone()
    if row:
        return int(row['id']) if hasattr(row, 'keys') else int(row[0])
    fallback = conn.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()
    if fallback:
        return int(fallback['id']) if hasattr(fallback, 'keys') else int(fallback[0])
    return 1


def _get_schedule_note_owner_id_for_user(conn, user: dict | None) -> int:
    if not user:
        return _get_shared_schedule_note_owner_id(conn)
    try:
        grade = int(user.get('grade') or 6)
    except Exception:
        grade = 6
    if grade <= 5:
        return _get_shared_schedule_note_owner_id(conn)
    try:
        return int(user.get('id') or _get_shared_schedule_note_owner_id(conn))
    except Exception:
        return _get_shared_schedule_note_owner_id(conn)


def _get_schedule_note_owner_id_by_user_id(conn, user_id: int) -> int:
    row = conn.execute("SELECT id, grade FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        return _get_shared_schedule_note_owner_id(conn)
    user = row_to_dict(row)
    return _get_schedule_note_owner_id_for_user(conn, user)


def _is_schedule_sync_exclusion_reason(reason: Any) -> bool:
    return str(reason or '').strip().startswith(SCHEDULE_EXCLUSION_REASON_PREFIX)


def _sync_schedule_business_exclusions(conn, schedule_date: str, excluded_business: str = '', excluded_business_details: list[dict] | None = None) -> None:
    normalized_date = _normalize_date_key(schedule_date)
    if not normalized_date:
        return
    details = excluded_business_details or []
    rows = conn.execute(
        """
        SELECT id, branch_no, nickname, name, email, show_in_branch_status
        FROM users
        WHERE branch_no IS NOT NULL
        ORDER BY COALESCE(branch_no, 9999), nickname, name, email
        """
    ).fetchall()
    user_rows = [row_to_dict(row) for row in rows]
    branch_map: dict[int, dict[str, Any]] = {}
    user_id_map: dict[int, dict[str, Any]] = {}
    for item in user_rows:
        try:
            user_id = int(item.get('id') or 0)
        except Exception:
            user_id = 0
        try:
            branch_no = int(item.get('branch_no') or 0)
        except Exception:
            branch_no = 0
        if user_id > 0:
            user_id_map[user_id] = item
        if branch_no > 0 and bool(item.get('show_in_branch_status', True)):
            branch_map[branch_no] = item

    desired_user_ids: set[int] = set()
    for entry in details:
        try:
            entry_user_id = int(entry.get('user_id') or 0)
        except Exception:
            entry_user_id = 0
        try:
            entry_branch_no = int(entry.get('branch_no') or 0)
        except Exception:
            entry_branch_no = 0
        if entry_user_id > 0 and entry_user_id in user_id_map:
            desired_user_ids.add(entry_user_id)
            continue
        if entry_branch_no > 0 and entry_branch_no in branch_map:
            desired_user_ids.add(int(branch_map[entry_branch_no]['id']))

    if not desired_user_ids:
        for branch_no in _parse_branch_exclusions(excluded_business):
            if branch_no in branch_map:
                desired_user_ids.add(int(branch_map[branch_no]['id']))

    existing_rows = conn.execute(
        "SELECT id, user_id, reason FROM vehicle_exclusions WHERE start_date = ? AND end_date = ?",
        (normalized_date, normalized_date),
    ).fetchall()
    existing_sync_by_user: dict[int, list[int]] = {}
    for row in existing_rows:
        item = row_to_dict(row)
        if not _is_schedule_sync_exclusion_reason(item.get('reason')):
            continue
        try:
            target_user_id = int(item.get('user_id') or 0)
        except Exception:
            target_user_id = 0
        if target_user_id <= 0:
            continue
        existing_sync_by_user.setdefault(target_user_id, []).append(int(item.get('id')))

    for user_id, exclusion_ids in existing_sync_by_user.items():
        if user_id not in desired_user_ids:
            for exclusion_id in exclusion_ids:
                conn.execute("DELETE FROM vehicle_exclusions WHERE id = ?", (exclusion_id,))

    now = utcnow()
    for user_id in desired_user_ids:
        if user_id in existing_sync_by_user:
            continue
        conn.execute(
            "INSERT INTO vehicle_exclusions(user_id, start_date, end_date, reason, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            (user_id, normalized_date, normalized_date, f'{SCHEDULE_EXCLUSION_REASON_PREFIX} 스케줄 열외', now, now),
        )


def _serialize_schedule_business_details(conn, raw_text: str) -> list[dict[str, Any]]:
    details: list[dict[str, Any]] = []
    branch_ids = _parse_branch_exclusions(raw_text)
    if not branch_ids:
        return details
    placeholders = ','.join('?' for _ in branch_ids)
    rows = conn.execute(
        f"SELECT id, branch_no, nickname, name, email FROM users WHERE branch_no IN ({placeholders}) ORDER BY COALESCE(branch_no, 9999), nickname, name, email",
        tuple(branch_ids),
    ).fetchall()
    row_map = {int(row['branch_no']): row_to_dict(row) for row in rows if row['branch_no'] is not None}
    for branch_no in branch_ids:
        row = row_map.get(int(branch_no))
        if row:
            label = _vehicle_account_display_name(row, default=f'{branch_no}호점')
            details.append({'user_id': int(row.get('id') or 0), 'branch_no': int(branch_no), 'name': label, 'reason': '스케줄 열외'})
        else:
            details.append({'branch_no': int(branch_no), 'name': f'{branch_no}호점', 'reason': '스케줄 열외'})
    return details

def _build_available_vehicle_accounts(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        branch_no = row.get('branch_no')
        display_name = _vehicle_account_display_name(row, default=(f"{branch_no}호점" if branch_no not in (None, '') else '미지정'))
        output.append({
            'branch_no': branch_no,
            'label': f"[{branch_no}호점 차량] {display_name}" if branch_no not in (None, '') else f"[미지정 차량] {display_name}",
            'display_name': display_name,
        })
    return output


def _get_vehicle_base_and_auto_unavailable(conn, date_keys: list[str]) -> tuple[int, dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    if not date_keys:
        return 0, {}, []
    rows = conn.execute(
        """
        SELECT id, branch_no, name, nickname, email, position_title, COALESCE(vehicle_available, 1) AS vehicle_available
        FROM users
        WHERE branch_no IS NOT NULL
        ORDER BY COALESCE(branch_no, 9999), nickname, name, email
        """
    ).fetchall()
    users = [row_to_dict(row) for row in rows]
    active_users = [row for row in users if bool(row.get('vehicle_available', 1))]
    active_user_map = {int(row.get('id') or 0): row for row in active_users if int(row.get('id') or 0) > 0}
    available_vehicle_accounts = _build_available_vehicle_accounts(active_users)

    exclusion_rows = conn.execute("SELECT id, user_id, start_date, end_date, reason FROM vehicle_exclusions").fetchall()
    exclusion_map: dict[int, list[dict[str, Any]]] = {}
    for row in exclusion_rows:
        entry = row_to_dict(row)
        user_id = int(entry.get('user_id') or 0)
        if user_id <= 0 or user_id not in active_user_map:
            continue
        start_key = _normalize_date_key(entry.get('start_date'))
        end_key = _normalize_date_key(entry.get('end_date'))
        if not start_key or not end_key or end_key < start_key:
            continue
        exclusion_map.setdefault(user_id, []).append({
            **entry,
            'start_date': start_key,
            'end_date': end_key,
        })

    result: dict[str, list[dict[str, Any]]] = {key: [] for key in date_keys}
    for key in date_keys:
        normalized_key = _normalize_date_key(key)
        if not normalized_key:
            continue
        for user_id, user_row in active_user_map.items():
            for exclusion in exclusion_map.get(user_id, []):
                if exclusion['start_date'] <= normalized_key <= exclusion['end_date']:
                    display_name = _vehicle_account_display_name(user_row, default=f'계정 {user_id}')
                    result[key].append({
                        'user_id': user_id,
                        'exclusion_id': entry.get('id') if (entry := exclusion) else None,
                        'branch_no': user_row.get('branch_no'),
                        'name': display_name,
                        'reason': str(exclusion.get('reason') or '').strip() or '차량열외',
                        'start_date': exclusion.get('start_date'),
                        'end_date': exclusion.get('end_date'),
                    })
                    break
    return len(active_users), result, available_vehicle_accounts


@app.api_route('/api/admin/accounts/{user_id}/vehicle-exclusions', methods=['GET', 'POST'])
@app.api_route('/api/admin/accounts/{user_id}/vehicle-exclusions/', methods=['GET', 'POST'])
@app.api_route('/api/admin/accounts/{user_id}/vehicle_exclusions', methods=['GET', 'POST'])
def vehicle_exclusions_collection(user_id: int, payload: Optional[VehicleExclusionIn] = None, request: Request = None, admin=Depends(require_admin_mode_user)):
    method = (request.method if request else 'GET').upper()
    if method == 'GET':
        return _list_vehicle_exclusions_response(user_id)
    if method == 'POST':
        if payload is None:
            raise HTTPException(status_code=400, detail='열외 일정 입력값이 없습니다.')
        return _create_vehicle_exclusion_response(user_id, payload)
    raise HTTPException(status_code=405, detail='허용되지 않는 요청입니다.')


@app.api_route('/api/admin/vehicle-exclusions/{user_id}', methods=['GET', 'POST'])
@app.api_route('/api/admin/vehicle_exclusions/{user_id}', methods=['GET', 'POST'])
def vehicle_exclusions_collection_alias(user_id: int, payload: Optional[VehicleExclusionIn] = None, request: Request = None, admin=Depends(require_admin_mode_user)):
    return vehicle_exclusions_collection(user_id=user_id, payload=payload, request=request, admin=admin)


def _list_vehicle_exclusions_response(user_id: int):
    with get_conn() as conn:
        account = conn.execute("SELECT id, name, nickname, branch_no FROM users WHERE id = ?", (user_id,)).fetchone()
        if not account:
            raise HTTPException(status_code=404, detail='계정을 찾을 수 없습니다.')
        rows = conn.execute("SELECT * FROM vehicle_exclusions WHERE user_id = ? ORDER BY start_date DESC, end_date DESC, id DESC", (user_id,)).fetchall()
    return {'account': row_to_dict(account), 'items': [row_to_dict(row) for row in rows]}


def _create_vehicle_exclusion_response(user_id: int, payload: VehicleExclusionIn):
    start_date = str(payload.start_date or '').strip()
    end_date = str(payload.end_date or '').strip()
    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', start_date) or not re.fullmatch(r'\d{4}-\d{2}-\d{2}', end_date):
        raise HTTPException(status_code=400, detail='열외 기간 날짜 형식이 올바르지 않습니다.')
    if end_date < start_date:
        raise HTTPException(status_code=400, detail='열외 종료일은 시작일보다 빠를 수 없습니다.')
    with get_conn() as conn:
        if not conn.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone():
            raise HTTPException(status_code=404, detail='계정을 찾을 수 없습니다.')
        now = utcnow()
        conn.execute("INSERT INTO vehicle_exclusions(user_id, start_date, end_date, reason, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)", (user_id, start_date, end_date, str(payload.reason or '').strip(), now, now))
        _sync_all_day_note_available_vehicle_counts(conn)
        rows = conn.execute("SELECT * FROM vehicle_exclusions WHERE user_id = ? ORDER BY start_date DESC, end_date DESC, id DESC", (user_id,)).fetchall()
    return {'ok': True, 'items': [row_to_dict(row) for row in rows]}




def _update_vehicle_exclusion_response(user_id: int, exclusion_id: int, payload: VehicleExclusionIn):
    start_date = str(payload.start_date or '').strip()
    end_date = str(payload.end_date or '').strip()
    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', start_date) or not re.fullmatch(r'\d{4}-\d{2}-\d{2}', end_date):
        raise HTTPException(status_code=400, detail='열외 기간 날짜 형식이 올바르지 않습니다.')
    if end_date < start_date:
        raise HTTPException(status_code=400, detail='열외 종료일은 시작일보다 빠를 수 없습니다.')
    with get_conn() as conn:
        if not conn.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone():
            raise HTTPException(status_code=404, detail='계정을 찾을 수 없습니다.')
        existing = conn.execute("SELECT id FROM vehicle_exclusions WHERE id = ? AND user_id = ?", (exclusion_id, user_id)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail='열외 일정을 찾을 수 없습니다.')
        now = utcnow()
        conn.execute(
            "UPDATE vehicle_exclusions SET start_date = ?, end_date = ?, reason = ?, updated_at = ? WHERE id = ? AND user_id = ?",
            (start_date, end_date, str(payload.reason or '').strip(), now, exclusion_id, user_id),
        )
        _sync_all_day_note_available_vehicle_counts(conn)
        rows = conn.execute("SELECT * FROM vehicle_exclusions WHERE user_id = ? ORDER BY start_date DESC, end_date DESC, id DESC", (user_id,)).fetchall()
    return {'ok': True, 'items': [row_to_dict(row) for row in rows]}


@app.api_route('/api/admin/accounts/{user_id}/vehicle-exclusions/{exclusion_id}', methods=['PUT'])
@app.api_route('/api/admin/accounts/{user_id}/vehicle-exclusions/{exclusion_id}/', methods=['PUT'])
@app.api_route('/api/admin/accounts/{user_id}/vehicle_exclusions/{exclusion_id}', methods=['PUT'])
def vehicle_exclusions_item_update(user_id: int, exclusion_id: int, payload: VehicleExclusionIn, request: Request = None, admin=Depends(require_admin_mode_user)):
    return _update_vehicle_exclusion_response(user_id, exclusion_id, payload)


@app.api_route('/api/admin/vehicle-exclusions/{user_id}/{exclusion_id}', methods=['PUT'])
@app.api_route('/api/admin/vehicle_exclusions/{user_id}/{exclusion_id}', methods=['PUT'])
def vehicle_exclusions_item_update_alias(user_id: int, exclusion_id: int, payload: VehicleExclusionIn, request: Request = None, admin=Depends(require_admin_mode_user)):
    return _update_vehicle_exclusion_response(user_id, exclusion_id, payload)

@app.api_route('/api/admin/accounts/{user_id}/vehicle-exclusions/{exclusion_id}', methods=['DELETE'])
@app.api_route('/api/admin/accounts/{user_id}/vehicle-exclusions/{exclusion_id}/', methods=['DELETE'])
@app.api_route('/api/admin/accounts/{user_id}/vehicle_exclusions/{exclusion_id}', methods=['DELETE'])
def vehicle_exclusions_item(user_id: int, exclusion_id: int, request: Request = None, admin=Depends(require_admin_mode_user)):
    return _delete_vehicle_exclusion_response(user_id, exclusion_id)


@app.api_route('/api/admin/vehicle-exclusions/{user_id}/{exclusion_id}', methods=['DELETE'])
@app.api_route('/api/admin/vehicle_exclusions/{user_id}/{exclusion_id}', methods=['DELETE'])
def vehicle_exclusions_item_alias(user_id: int, exclusion_id: int, request: Request = None, admin=Depends(require_admin_mode_user)):
    return _delete_vehicle_exclusion_response(user_id, exclusion_id)


def _delete_vehicle_exclusion_response(user_id: int, exclusion_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM vehicle_exclusions WHERE id = ? AND user_id = ?", (exclusion_id, user_id))
        _sync_all_day_note_available_vehicle_counts(conn)
        rows = conn.execute("SELECT * FROM vehicle_exclusions WHERE user_id = ? ORDER BY start_date DESC, end_date DESC, id DESC", (user_id,)).fetchall()
    return {'ok': True, 'items': [row_to_dict(row) for row in rows]}


def _sync_work_schedule_day_note_counts(conn, user_id: int, schedule_date: str):
    owner_user_id = _get_schedule_note_owner_id_by_user_id(conn, user_id)
    event_rows = conn.execute(
        """
        SELECT
            COALESCE(SUM(COALESCE(status_a_count, 0)), 0) AS status_a_count,
            COALESCE(SUM(COALESCE(status_b_count, 0)), 0) AS status_b_count,
            COALESCE(SUM(COALESCE(status_c_count, 0)), 0) AS status_c_count
        FROM calendar_events
        WHERE user_id = ? AND event_date = ?
        """,
        (user_id, schedule_date),
    ).fetchone()
    now = utcnow()
    total_a = int(event_rows['status_a_count'] or 0) if event_rows else 0
    total_b = int(event_rows['status_b_count'] or 0) if event_rows else 0
    total_c = int(event_rows['status_c_count'] or 0) if event_rows else 0
    _, auto_unavailable_map, _ = _get_vehicle_base_and_auto_unavailable(conn, [schedule_date])
    excluded_count = len({int(entry.get('user_id') or 0) for entry in auto_unavailable_map.get(schedule_date, []) if int(entry.get('user_id') or 0) > 0})
    next_available_count = max(_get_admin_total_vehicle_count(conn) - excluded_count, 0)
    existing = conn.execute(
        "SELECT id FROM work_schedule_day_notes WHERE user_id = ? AND schedule_date = ?",
        (owner_user_id, schedule_date),
    ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE work_schedule_day_notes
            SET available_vehicle_count = ?, status_a_count = ?, status_b_count = ?, status_c_count = ?, updated_at = ?
            WHERE user_id = ? AND schedule_date = ?
            """,
            (next_available_count, total_a, total_b, total_c, now, owner_user_id, schedule_date),
        )
    else:
        conn.execute(
            """
            INSERT INTO work_schedule_day_notes(
                user_id, schedule_date, excluded_business, excluded_staff, excluded_business_details, excluded_staff_details,
                available_vehicle_count, status_a_count, status_b_count, status_c_count, day_memo, is_handless_day, created_at, updated_at
            )
            VALUES (?, ?, '', '', '[]', '[]', ?, ?, ?, ?, '', 0, ?, ?)
            """,
            (owner_user_id, schedule_date, next_available_count, total_a, total_b, total_c, now, now),
        )

def _collect_auto_unavailable_business(conn, date_keys: list[str]) -> dict[str, list[dict[str, Any]]]:
    _, result, _ = _get_vehicle_base_and_auto_unavailable(conn, date_keys)
    return result

@app.get('/api/admin/accounts/{user_id}/vehicle-exclusions')
@app.get('/api/admin/accounts/{user_id}/vehicle-exclusions/')
@app.get('/api/admin/accounts/{user_id}/vehicle_exclusions')
def list_vehicle_exclusions(user_id: int, admin=Depends(require_admin_mode_user)):
    with get_conn() as conn:
        account = conn.execute("SELECT id, name, nickname, branch_no FROM users WHERE id = ?", (user_id,)).fetchone()
        if not account:
            raise HTTPException(status_code=404, detail='계정을 찾을 수 없습니다.')
        rows = conn.execute("SELECT * FROM vehicle_exclusions WHERE user_id = ? ORDER BY start_date DESC, end_date DESC, id DESC", (user_id,)).fetchall()
    return {'account': row_to_dict(account), 'items': [row_to_dict(row) for row in rows]}

@app.post('/api/admin/accounts/{user_id}/vehicle-exclusions')
@app.post('/api/admin/accounts/{user_id}/vehicle-exclusions/')
@app.post('/api/admin/accounts/{user_id}/vehicle_exclusions')
def create_vehicle_exclusion(user_id: int, payload: VehicleExclusionIn, admin=Depends(require_admin_mode_user)):
    start_date = str(payload.start_date or '').strip()
    end_date = str(payload.end_date or '').strip()
    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', start_date) or not re.fullmatch(r'\d{4}-\d{2}-\d{2}', end_date):
        raise HTTPException(status_code=400, detail='열외 기간 날짜 형식이 올바르지 않습니다.')
    if end_date < start_date:
        raise HTTPException(status_code=400, detail='열외 종료일은 시작일보다 빠를 수 없습니다.')
    with get_conn() as conn:
        if not conn.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone():
            raise HTTPException(status_code=404, detail='계정을 찾을 수 없습니다.')
        now = utcnow()
        conn.execute("INSERT INTO vehicle_exclusions(user_id, start_date, end_date, reason, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)", (user_id, start_date, end_date, str(payload.reason or '').strip(), now, now))
        _sync_all_day_note_available_vehicle_counts(conn)
        rows = conn.execute("SELECT * FROM vehicle_exclusions WHERE user_id = ? ORDER BY start_date DESC, end_date DESC, id DESC", (user_id,)).fetchall()
    return {'ok': True, 'items': [row_to_dict(row) for row in rows]}

@app.delete('/api/admin/accounts/{user_id}/vehicle-exclusions/{exclusion_id}')
@app.delete('/api/admin/accounts/{user_id}/vehicle-exclusions/{exclusion_id}/')
@app.delete('/api/admin/accounts/{user_id}/vehicle_exclusions/{exclusion_id}')
def delete_vehicle_exclusion(user_id: int, exclusion_id: int, admin=Depends(require_admin_mode_user)):
    with get_conn() as conn:
        conn.execute("DELETE FROM vehicle_exclusions WHERE id = ? AND user_id = ?", (exclusion_id, user_id))
        _sync_all_day_note_available_vehicle_counts(conn)
        rows = conn.execute("SELECT * FROM vehicle_exclusions WHERE user_id = ? ORDER BY start_date DESC, end_date DESC, id DESC", (user_id,)).fetchall()
    return {'ok': True, 'items': [row_to_dict(row) for row in rows]}

@app.get('/api/work-schedule')
def get_work_schedule(start_date: Optional[str] = Query(default=None), days: int = Query(default=7, ge=1, le=62), user=Depends(require_user)):
    base_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else datetime.now().date()
    date_keys = [(base_date + timedelta(days=index)).isoformat() for index in range(days)]
    with get_conn() as conn:
        placeholders = ','.join('?' for _ in date_keys)
        if _can_access_shared_schedule(user):
            shared_ids = _shared_schedule_user_ids(conn)
        else:
            shared_ids = [int(user['id'])]
        if not shared_ids:
            shared_ids = [int(user['id'])]
        shared_placeholders = ','.join('?' for _ in shared_ids)
        work_rows = conn.execute(
            f"""
            SELECT * FROM work_schedule_entries
            WHERE user_id IN ({shared_placeholders}) AND schedule_date IN ({placeholders})
            ORDER BY schedule_date, CASE WHEN COALESCE(schedule_time, '') = '' THEN '99:99' ELSE schedule_time END, id
            """,
            (*shared_ids, *date_keys),
        ).fetchall()
        event_rows = conn.execute(
            f"""
            SELECT * FROM calendar_events
            WHERE user_id IN ({shared_placeholders}) AND event_date IN ({placeholders})
            ORDER BY event_date, CASE WHEN COALESCE(start_time, '') IN ('', '미정') THEN '99:99' ELSE start_time END, id
            """,
            (*shared_ids, *date_keys),
        ).fetchall()
        notes_owner_id = _get_schedule_note_owner_id_for_user(conn, user)
        notes_rows = conn.execute(
            f"""
            SELECT * FROM work_schedule_day_notes
            WHERE user_id = ? AND schedule_date IN ({placeholders})
            """,
            (notes_owner_id, *date_keys),
        ).fetchall()
        branch_rows = conn.execute("SELECT branch_no, name, nickname FROM users WHERE branch_no IS NOT NULL").fetchall()
        dynamic_total_vehicle_count, auto_unavailable_by_date, available_vehicle_accounts = _get_vehicle_base_and_auto_unavailable(conn, date_keys)
    branch_name_map = {int(r['branch_no']): (str(r['name'] or r['nickname'] or '').strip() or f"{int(r['branch_no'])}호점") for r in branch_rows if r['branch_no'] is not None}
    entries_by_date = {key: [] for key in date_keys}
    for row in event_rows:
        item = row_to_dict(row)
        representative_names = ' / '.join([name for name in [row['representative1'], row['representative2'], row['representative3']] if str(name or '').strip()])
        staff_names = ' / '.join([name for name in [row['staff1'], row['staff2'], row['staff3']] if str(name or '').strip()])
        entries_by_date[row['event_date']].append({
            'id': f"calendar-{row['id']}",
            'entry_type': 'calendar',
            'event_id': row['id'],
            'schedule_date': row['event_date'],
            'schedule_time': '' if row['start_time'] in ('', '미정') else row['start_time'],
            'customer_name': row['customer_name'] or '고객명',
            'representative_names': representative_names,
            'staff_names': staff_names,
            'memo': row['content'] or row['location'] or '',
            'title': row['title'],
            'color': row['color'] or '#2563eb',
            'platform': row['platform'] or '',
            'department_info': row['department_info'] or '',
            'amount1': row['amount1'] or '',
            'amount2': row['amount2'] or '',
            'deposit_method': row['deposit_method'] or '',
            'deposit_amount': row['deposit_amount'] or '',
            'source_summary': row['title'],
            'created_by': user.get('nickname') or '',
            'status_a_count': int(row['status_a_count'] or 0),
            'status_b_count': int(row['status_b_count'] or 0),
            'status_c_count': int(row['status_c_count'] or 0),
        })
    for row in work_rows:
        item = row_to_dict(row)
        item['entry_type'] = 'manual'
        item['event_id'] = None
        item['color'] = '#334155'
        item['title'] = ''
        item['platform'] = ''
        item['department_info'] = ''
        item['amount1'] = ''
        item['amount2'] = ''
        item['deposit_method'] = ''
        item['deposit_amount'] = ''
        item['source_summary'] = ''
        item['created_by'] = user.get('nickname') or ''
        entries_by_date[row['schedule_date']].append(item)
    notes_by_date = {row['schedule_date']: row_to_dict(row) for row in notes_rows}
    output = []
    for key in date_keys:
        target = datetime.strptime(key, '%Y-%m-%d').date()
        note = notes_by_date.get(key, {})
        excluded_business = note.get('excluded_business', '')
        excluded_staff = note.get('excluded_staff', '')
        excluded_business_details = json.loads(note.get('excluded_business_details') or '[]')
        excluded_staff_details = json.loads(note.get('excluded_staff_details') or '[]')
        branch_ids = _parse_branch_exclusions(excluded_business)
        auto_unavailable = auto_unavailable_by_date.get(key, [])
        auto_user_ids = {int(item.get('user_id') or 0) for item in auto_unavailable if int(item.get('user_id') or 0) > 0}
        excluded_vehicle_count = len(auto_user_ids)
        excluded_business_names = []
        if excluded_business_details:
            for entry in excluded_business_details:
                name = str(entry.get('name') or entry.get('label') or '').strip() or '사업자'
                reason = str(entry.get('reason') or '').strip()
                branch_no = entry.get('branch_no')
                branch_label = '본점' if str(branch_no).strip() in {'0', '본점'} else (f'{int(branch_no)}호점' if str(branch_no).strip().isdigit() else str(branch_no or '-').strip())
                excluded_business_names.append(f'[{branch_label}] | [{name}] | [(사유 : {reason or "-"})]')
        else:
            for branch_no in branch_ids:
                display_name = branch_name_map.get(branch_no, f'{branch_no}호점')
                excluded_business_names.append(f'{display_name}-열외')
        for entry in auto_unavailable:
            text_value = f"{str(entry.get('name') or '').strip() or f'{entry.get('branch_no')}호점'} (사유 : {str(entry.get('reason') or '').strip() or '차량열외'})"
            if text_value not in excluded_business_names:
                excluded_business_names.append(text_value)
        if excluded_staff_details:
            staff_display = [f"{str(entry.get('name') or '직원').strip()} (사유 : {str(entry.get('reason') or '-').strip() or '-'})" for entry in excluded_staff_details]
        else:
            staff_tokens = [token.strip() for token in re.split(r'[\n,/]+', excluded_staff or '') if token.strip()]
            staff_display = [token if '-' in token else f'{token}-열외' for token in staff_tokens]
        day_entries = sorted(entries_by_date[key], key=lambda item: ((item.get('schedule_time') or '99:99') if (item.get('schedule_time') or '') not in ('', '미정') else '99:99', str(item.get('customer_name') or item.get('title') or ''), str(item.get('id'))))
        try:
            base_available_count = max(int(dynamic_total_vehicle_count or 0), 0)
        except (TypeError, ValueError):
            base_available_count = 0
        available_vehicle_count = max(base_available_count - excluded_vehicle_count, 0)
        def _safe_count(value):
            try:
                return max(int(value or 0), 0)
            except (TypeError, ValueError):
                return 0
        event_status_a_count = sum(_safe_count(item.get('status_a_count')) for item in day_entries if item.get('entry_type') == 'calendar')
        event_status_b_count = sum(_safe_count(item.get('status_b_count')) for item in day_entries if item.get('entry_type') == 'calendar')
        event_status_c_count = sum(_safe_count(item.get('status_c_count')) for item in day_entries if item.get('entry_type') == 'calendar')
        output.append({
            'date': key,
            'title': _schedule_day_title(base_date, target),
            'entries': day_entries,
            'excluded_business': excluded_business,
            'excluded_business_names': excluded_business_names,
            'auto_unavailable_business': auto_unavailable,
            'excluded_staff': excluded_staff,
            'excluded_staff_names': staff_display,
            'excluded_vehicle_count': excluded_vehicle_count,
            'available_vehicle_count': available_vehicle_count,
            'base_vehicle_count': base_available_count,
            'available_vehicle_accounts': available_vehicle_accounts,
            'status_a_count': event_status_a_count if event_status_a_count else _safe_count(note.get('status_a_count')),
            'status_b_count': event_status_b_count if event_status_b_count else _safe_count(note.get('status_b_count')),
            'status_c_count': event_status_c_count if event_status_c_count else _safe_count(note.get('status_c_count')),
            'day_memo': note.get('day_memo', '') or '',
            'excluded_business_details': excluded_business_details,
            'excluded_staff_details': excluded_staff_details,
            'is_handless_day': bool(note.get('is_handless_day')),
        })
    return {'days': output}
@app.post("/api/work-schedule/entries")
def create_work_schedule_entry(payload: WorkScheduleEntryIn, user=Depends(require_user)):
    _require_write_access(user, 'work_schedule')
    now = utcnow()
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO work_schedule_entries(user_id, schedule_date, schedule_time, customer_name, representative_names, staff_names, memo, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (user['id'], payload.schedule_date, payload.schedule_time, payload.customer_name, payload.representative_names, payload.staff_names, payload.memo, now, now),
        )
        row = conn.execute(
            "SELECT * FROM work_schedule_entries WHERE user_id = ? AND created_at = ? ORDER BY id DESC LIMIT 1",
            (user['id'], now),
        ).fetchone()
        _notify_work_schedule_assignments(
            conn,
            actor=user,
            schedule_date=payload.schedule_date,
            schedule_time=payload.schedule_time,
            customer_name=payload.customer_name,
            representative_names=payload.representative_names,
            staff_names=payload.staff_names,
        )
        return row_to_dict(row)
@app.put("/api/work-schedule/entries/{entry_id}")
def update_work_schedule_entry(entry_id: int, payload: WorkScheduleEntryIn, user=Depends(require_user)):
    _require_write_access(user, 'work_schedule')
    with get_conn() as conn:
        existing = conn.execute("SELECT * FROM work_schedule_entries WHERE id = ? AND user_id = ?", (entry_id, user['id'])).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail='스케줄 항목을 찾을 수 없습니다.')
        previous_ids = set(_work_assignment_target_ids(conn, existing['representative_names'], existing['staff_names'], user.get('id')))
        conn.execute(
            """
            UPDATE work_schedule_entries
            SET schedule_date = ?, schedule_time = ?, customer_name = ?, representative_names = ?, staff_names = ?, memo = ?, updated_at = ?
            WHERE id = ? AND user_id = ?
            """,
            (payload.schedule_date, payload.schedule_time, payload.customer_name, payload.representative_names, payload.staff_names, payload.memo, utcnow(), entry_id, user['id']),
        )
        row = conn.execute("SELECT * FROM work_schedule_entries WHERE id = ?", (entry_id,)).fetchone()
        next_data = row_to_dict(row) if row else {}
        _notify_work_schedule_assignments(
            conn,
            actor=user,
            schedule_date=payload.schedule_date,
            schedule_time=payload.schedule_time,
            customer_name=payload.customer_name,
            representative_names=payload.representative_names,
            staff_names=payload.staff_names,
            previous_ids=previous_ids,
        )
        _notify_work_schedule_entry_changes(conn, user, row_to_dict(existing), next_data)
        return next_data
@app.delete("/api/work-schedule/entries/{entry_id}")
def delete_work_schedule_entry(entry_id: int, user=Depends(require_user)):
    _require_write_access(user, 'work_schedule')
    with get_conn() as conn:
        conn.execute("DELETE FROM work_schedule_entries WHERE id = ? AND user_id = ?", (entry_id, user['id']))
        return {'ok': True}
@app.put("/api/work-schedule/day-note")
def upsert_work_schedule_day_note(payload: WorkScheduleDayNoteIn, user=Depends(require_user)):
    _require_write_access(user, 'work_schedule')
    with get_conn() as conn:
        owner_user_id = _get_schedule_note_owner_id_for_user(conn, user)
        normalized_details = payload.excluded_business_details or _serialize_schedule_business_details(conn, payload.excluded_business)
        _sync_schedule_business_exclusions(conn, payload.schedule_date, payload.excluded_business, normalized_details)
        _, auto_unavailable_map, _ = _get_vehicle_base_and_auto_unavailable(conn, [payload.schedule_date])
        excluded_count = len({int(entry.get('user_id') or 0) for entry in auto_unavailable_map.get(payload.schedule_date, []) if int(entry.get('user_id') or 0) > 0})
        existing = conn.execute("SELECT id FROM work_schedule_day_notes WHERE user_id = ? AND schedule_date = ?", (owner_user_id, payload.schedule_date)).fetchone()
        now = utcnow()
        computed_available_vehicle_count = max(_get_admin_total_vehicle_count(conn) - excluded_count, 0)
        if existing:
            conn.execute(
                """
                UPDATE work_schedule_day_notes
                SET excluded_business = ?, excluded_staff = ?, excluded_business_details = ?, excluded_staff_details = ?, available_vehicle_count = ?, status_a_count = ?, status_b_count = ?, status_c_count = ?, day_memo = ?, is_handless_day = ?, updated_at = ?
                WHERE user_id = ? AND schedule_date = ?
                """,
                (payload.excluded_business, payload.excluded_staff, json.dumps(normalized_details, ensure_ascii=False), json.dumps(payload.excluded_staff_details, ensure_ascii=False), computed_available_vehicle_count, payload.status_a_count, payload.status_b_count, payload.status_c_count, payload.day_memo, 1 if payload.is_handless_day else 0, now, owner_user_id, payload.schedule_date),
            )
        else:
            conn.execute(
                """
                INSERT INTO work_schedule_day_notes(user_id, schedule_date, excluded_business, excluded_staff, excluded_business_details, excluded_staff_details, available_vehicle_count, status_a_count, status_b_count, status_c_count, day_memo, is_handless_day, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (owner_user_id, payload.schedule_date, payload.excluded_business, payload.excluded_staff, json.dumps(normalized_details, ensure_ascii=False), json.dumps(payload.excluded_staff_details, ensure_ascii=False), computed_available_vehicle_count, payload.status_a_count, payload.status_b_count, payload.status_c_count, payload.day_memo, 1 if payload.is_handless_day else 0, now, now),
            )
        row = conn.execute("SELECT * FROM work_schedule_day_notes WHERE user_id = ? AND schedule_date = ?", (owner_user_id, payload.schedule_date)).fetchone()
        return row_to_dict(row)

@app.post("/api/work-schedule/handless-bulk")
def save_handless_bulk(payload: HandlessBulkIn, user=Depends(require_user)):
    _require_write_access(user, 'work_schedule')
    visible_dates = [str(item).strip() for item in payload.visible_dates if str(item).strip()]
    selected_dates = set(str(item).strip() for item in payload.selected_dates if str(item).strip())
    now = utcnow()
    with get_conn() as conn:
        owner_user_id = _get_schedule_note_owner_id_for_user(conn, user)
        existing_rows = conn.execute(
            f"SELECT * FROM work_schedule_day_notes WHERE user_id = ? AND schedule_date IN ({','.join('?' for _ in visible_dates)})",
            (owner_user_id, *visible_dates),
        ).fetchall() if visible_dates else []
        existing_map = {row['schedule_date']: row_to_dict(row) for row in existing_rows}
        for schedule_date in visible_dates:
            current = existing_map.get(schedule_date, {})
            is_handless = schedule_date in selected_dates
            _, auto_unavailable_map, _ = _get_vehicle_base_and_auto_unavailable(conn, [schedule_date])
            excluded_count = len({int(entry.get('user_id') or 0) for entry in auto_unavailable_map.get(schedule_date, []) if int(entry.get('user_id') or 0) > 0})
            payload_row = {
                'excluded_business': current.get('excluded_business', '') or '',
                'excluded_staff': current.get('excluded_staff', '') or '',
                'excluded_business_details': json.loads(current.get('excluded_business_details') or '[]') if isinstance(current.get('excluded_business_details'), str) else current.get('excluded_business_details', []) or [],
                'excluded_staff_details': json.loads(current.get('excluded_staff_details') or '[]') if isinstance(current.get('excluded_staff_details'), str) else current.get('excluded_staff_details', []) or [],
                'available_vehicle_count': max(_get_admin_total_vehicle_count(conn) - excluded_count, 0),
                'status_a_count': int(current.get('status_a_count') or 0),
                'status_b_count': int(current.get('status_b_count') or 0),
                'status_c_count': int(current.get('status_c_count') or 0),
                'day_memo': current.get('day_memo', '') or '',
                'is_handless_day': is_handless,
            }
            existing = conn.execute("SELECT id FROM work_schedule_day_notes WHERE user_id = ? AND schedule_date = ?", (owner_user_id, schedule_date)).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE work_schedule_day_notes
                    SET excluded_business = ?, excluded_staff = ?, excluded_business_details = ?, excluded_staff_details = ?, available_vehicle_count = ?, status_a_count = ?, status_b_count = ?, status_c_count = ?, day_memo = ?, is_handless_day = ?, updated_at = ?
                    WHERE user_id = ? AND schedule_date = ?
                    """,
                    (payload_row['excluded_business'], payload_row['excluded_staff'], json.dumps(payload_row['excluded_business_details'], ensure_ascii=False), json.dumps(payload_row['excluded_staff_details'], ensure_ascii=False), payload_row['available_vehicle_count'], payload_row['status_a_count'], payload_row['status_b_count'], payload_row['status_c_count'], payload_row['day_memo'], 1 if payload_row['is_handless_day'] else 0, now, owner_user_id, schedule_date),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO work_schedule_day_notes(user_id, schedule_date, excluded_business, excluded_staff, excluded_business_details, excluded_staff_details, available_vehicle_count, status_a_count, status_b_count, status_c_count, day_memo, is_handless_day, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (owner_user_id, schedule_date, payload_row['excluded_business'], payload_row['excluded_staff'], json.dumps(payload_row['excluded_business_details'], ensure_ascii=False), json.dumps(payload_row['excluded_staff_details'], ensure_ascii=False), payload_row['available_vehicle_count'], payload_row['status_a_count'], payload_row['status_b_count'], payload_row['status_c_count'], payload_row['day_memo'], 1 if payload_row['is_handless_day'] else 0, now, now),
                )
        return {'ok': True, 'saved_count': len(visible_dates)}



def _digits_only(value: str | None) -> str:
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except Exception:
        return default


def _quote_desired_date(item: dict, payload: dict) -> str:
    if item.get('form_type') == 'storage':
        start_date = str(payload.get('storage_start_date') or item.get('desired_date') or '').strip()
        end_date = str(payload.get('storage_end_date') or '').strip()
        return ' ~ '.join([part for part in [start_date, end_date] if part]) or str(item.get('desired_date') or '').strip()
    return str(payload.get('move_date') or item.get('desired_date') or '').strip()


def _quote_primary_date(item: dict, payload: dict) -> str:
    if item.get('form_type') == 'storage':
        return str(payload.get('storage_start_date') or item.get('desired_date') or '').strip()
    return str(payload.get('move_date') or item.get('desired_date') or '').strip()


def _compute_quote_estimate(item: dict) -> dict:
    payload = item.get('payload') or {}
    area = str(payload.get('area') or '')
    household = str(payload.get('household') or '')
    move_types = payload.get('move_types') or []
    premium_options = payload.get('premium_options') or []
    furniture_types = payload.get('furniture_types') or []
    disassembly_types = payload.get('disassembly_types') or []
    large_item_types = payload.get('large_item_types') or []
    via_exists = any(str(payload.get(key) or '').strip() for key in ['via_address', 'via_address_detail', 'via_pickup_items', 'via_drop_items'])
    origin_elevator = str(payload.get('origin_elevator') or '')
    destination_elevator = str(payload.get('destination_elevator') or '')
    form_type = 'storage' if item.get('form_type') == 'storage' else 'same_day'
    score = 0
    score += {'1인': 0, '2인': 2, '3인': 4, '4인 이상': 6}.get(household, 1)
    score += {'10평 미만': 0, '10평대': 2, '20평대': 4, '30평대 이상': 7}.get(area, 1)
    score += len(move_types) * 2 + len(premium_options) + len(furniture_types) + len(disassembly_types) + len(large_item_types)
    if form_type == 'storage':
        score += 4
    if via_exists:
        score += 2
    if origin_elevator == '없음':
        score += 2
    if destination_elevator == '없음':
        score += 2
    if any('포장이사' in str(v) for v in move_types):
        score += 3
    if any('반포장' in str(v) for v in move_types):
        score += 2

    if score <= 5:
        crew = 1
        vehicles = 1
        low, high = 120000, 220000
        grade = '소형'
    elif score <= 10:
        crew = 2
        vehicles = 1
        low, high = 220000, 380000
        grade = '일반'
    elif score <= 16:
        crew = 3
        vehicles = 1
        low, high = 380000, 650000
        grade = '중형'
    else:
        crew = 4
        vehicles = 2
        low, high = 650000, 1200000
        grade = '대형'

    if form_type == 'storage':
        low += 100000
        high += 220000
    if via_exists:
        low += 50000
        high += 120000
    if any('피아노' in str(v) or '냉장고' in str(v) or '세탁기' in str(v) for v in large_item_types):
        low += 40000
        high += 100000

    lines = [
        f"기본 난이도: {grade}",
        f"가구원/평수/옵션 반영 점수: {score}",
        f"추천 인원: {crew}명",
        f"추천 차량: {vehicles}대",
    ]
    if form_type == 'storage':
        lines.append('짐보관이사 가산 금액이 포함되었습니다.')
    if via_exists:
        lines.append('경유지 정보가 있어 추가 금액이 반영되었습니다.')
    if origin_elevator == '없음' or destination_elevator == '없음':
        lines.append('엘리베이터 없음 조건이 반영되었습니다.')
    return {
        'recommended_crew': crew,
        'recommended_vehicle_count': vehicles,
        'estimated_low': low,
        'estimated_high': high,
        'difficulty_grade': grade,
        'score': score,
        'explanation_lines': lines,
        'move_date_label': _quote_desired_date(item, payload),
    }


def _find_repeat_customer(conn, item: dict) -> list[dict]:
    phone_digits = _digits_only(item.get('contact_phone'))
    if not phone_digits:
        return []
    rows = conn.execute(
        "SELECT id, requester_name, desired_date, summary_title, created_at, payload_json FROM quote_form_submissions WHERE REPLACE(REPLACE(REPLACE(contact_phone, '-', ''), ' ', ''), '.', '') = ? AND id <> ? ORDER BY created_at DESC LIMIT 10",
        (phone_digits, int(item.get('id') or 0)),
    ).fetchall()
    results = []
    for row in rows:
        data = row_to_dict(row)
        payload = json_loads(data.get('payload_json'), {})
        results.append({
            'id': data.get('id'),
            'customer_name': data.get('requester_name') or '',
            'desired_date': _quote_desired_date(data, payload),
            'summary_title': data.get('summary_title') or '',
            'created_at': data.get('created_at') or '',
        })
    return results


def _schedule_conflict_analysis(conn, item: dict, estimate: dict) -> dict:
    payload = item.get('payload') or {}
    target_date = _quote_primary_date(item, payload)
    if not target_date:
        return {'target_date': '', 'available_vehicle_count': None, 'conflicts': [], 'conflict_level': '확인불가', 'recommended_action': '이사 희망 날짜가 없어 충돌 분석을 생략했습니다.'}
    event_rows = conn.execute(
        "SELECT id, customer_name, start_time, representative1, representative2, representative3, staff1, staff2, staff3, status_a_count, status_b_count, status_c_count FROM calendar_events WHERE event_date = ? ORDER BY CASE WHEN COALESCE(start_time, '') = '' THEN '99:99' ELSE start_time END, id",
        (target_date,),
    ).fetchall()
    note_row = conn.execute("SELECT available_vehicle_count, excluded_business, excluded_staff, day_memo FROM work_schedule_day_notes WHERE schedule_date = ? ORDER BY id DESC LIMIT 1", (target_date,)).fetchone()
    available_vehicle_count = None
    if note_row:
        available_vehicle_count = _to_int(note_row['available_vehicle_count'], 0)
    conflicts = []
    used_vehicle_slots = 0
    for row in event_rows:
        assigned_people = [str(row[k] or '').strip() for k in ['representative1','representative2','representative3','staff1','staff2','staff3'] if str(row[k] or '').strip()]
        vehicle_need = max(1, _to_int(row['status_a_count']) + _to_int(row['status_b_count']) + _to_int(row['status_c_count']))
        used_vehicle_slots += vehicle_need
        conflicts.append({
            'event_id': row['id'],
            'customer_name': row['customer_name'] or '-',
            'start_time': row['start_time'] or '미정',
            'assigned_people': assigned_people,
            'vehicle_need': vehicle_need,
        })
    post_use_remaining = None if available_vehicle_count is None else available_vehicle_count - used_vehicle_slots - int(estimate.get('recommended_vehicle_count') or 0)
    if available_vehicle_count is None:
        level = '확인필요'
        action = '일정 메모의 가용차량수가 없어서 수동 확인이 필요합니다.'
    elif post_use_remaining < 0:
        level = '충돌위험'
        action = f"추천 차량 {estimate.get('recommended_vehicle_count')}대 기준으로 {abs(post_use_remaining)}대 부족합니다. 날짜 변경 또는 차량 재배치가 필요합니다."
    elif post_use_remaining == 0:
        level = '주의'
        action = '가용차량이 정확히 소진됩니다. 열외차량/인원 배치 재확인이 필요합니다.'
    else:
        level = '가능'
        action = f"추가 투입 후에도 차량 {post_use_remaining}대 여유가 있습니다."
    return {
        'target_date': target_date,
        'available_vehicle_count': available_vehicle_count,
        'scheduled_vehicle_count': used_vehicle_slots,
        'remaining_vehicle_count_after_assignment': post_use_remaining,
        'conflicts': conflicts,
        'conflict_level': level,
        'recommended_action': action,
    }


def _deposit_alert_summary(item: dict) -> dict:
    payload = item.get('payload') or {}
    target_date = _quote_primary_date(item, payload)
    if not target_date:
        return {'target_date': '', 'days_until_move': None, 'should_alert': False, 'message': '이사일이 없어 계약금 알림 판정을 생략했습니다.'}
    try:
        move_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        days_until = (move_date - date.today()).days
    except Exception:
        return {'target_date': target_date, 'days_until_move': None, 'should_alert': False, 'message': '날짜 형식 확인이 필요합니다.'}
    should_alert = days_until <= 3
    message = '관리자 알림 대상입니다. 이사일 3일 전 이내입니다.' if should_alert else f'현재 기준 D-{days_until} 입니다. 3일 전부터 관리자 알림 대상으로 전환됩니다.'
    return {'target_date': target_date, 'days_until_move': days_until, 'should_alert': should_alert, 'message': message}


def _recommended_checklist(item: dict) -> dict:
    payload = item.get('payload') or {}
    move_types = payload.get('move_types') or []
    base_items = [
        {'label': '고객 연락 및 당일 일정 재확인', 'checked': False},
        {'label': '출발지/도착지 주소 및 주차 가능 여부 확인', 'checked': False},
        {'label': '엘리베이터 / 계단 / 사다리 필요 여부 확인', 'checked': False},
        {'label': '파손 우려 물품 사전 체크', 'checked': False},
    ]
    if item.get('form_type') == 'storage':
        base_items.extend([
            {'label': '보관 시작/종료일 및 창고 위치 재확인', 'checked': False},
            {'label': '보관 중 파손방지 포장 강화', 'checked': False},
        ])
    if any('포장' in str(v) for v in move_types):
        base_items.append({'label': '포장 자재(박스/테이프/커버) 사전 준비', 'checked': False})
    if str(payload.get('waste_service') or ''):
        base_items.append({'label': '폐기물 신고 서비스 접수 여부 재확인', 'checked': False})
    if any(str(payload.get(k) or '').strip() for k in ['via_address', 'via_pickup_items', 'via_drop_items']):
        base_items.append({'label': '경유지 상하차 물품 분리 라벨링', 'checked': False})
    return {'name': 'AI 추천 체크리스트', 'items': base_items}


def _quote_operations_preview(conn, item: dict) -> dict:
    estimate = _compute_quote_estimate(item)
    return {
        'estimate': estimate,
        'crm_matches': _find_repeat_customer(conn, item),
        'schedule_analysis': _schedule_conflict_analysis(conn, item, estimate),
        'deposit_alert': _deposit_alert_summary(item),
        'recommended_checklist': _recommended_checklist(item),
        'vehicle_tracking_summary': {
            'status': '준비완료',
            'message': '차량 위치 공유 / 지오펜스 자동 상태 변경을 연결할 준비가 완료되었습니다. 지도 화면 위치 데이터와 연동하면 도착/작업중/완료 자동 전환이 가능합니다.',
        },
        'review_automation_summary': {
            'status': '준비완료',
            'message': '숨고 / 오늘의집 리뷰 수집 이후 AI 답변 초안 자동 생성 흐름으로 연결할 수 있습니다.',
        },
        'evidence_summary': {
            'status': '준비완료',
            'message': '전/후 사진 업로드 및 증빙 관리 테이블이 추가되었습니다. 추후 R2/S3 업로드 API 연결 시 즉시 확장 가능합니다.',
        },
        'attendance_summary': {
            'status': '준비완료',
            'message': 'GPS 출퇴근 및 일정 기반 급여 계산용 요약 테이블이 추가되었습니다.',
        },
    }


def _build_quote_estimate_workbook(item: dict, preview: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = '견적추출'
    header_fill = PatternFill(fill_type='solid', fgColor='DDEBF7')
    section_fill = PatternFill(fill_type='solid', fgColor='E2F0D9')
    bold = Font(bold=True)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 48
    rows = [
        ('고객명', item.get('requester_name') or ''),
        ('연락처', item.get('contact_phone') or ''),
        ('희망일', preview['estimate'].get('move_date_label') or ''),
        ('예상 견적 하한', int(preview['estimate'].get('estimated_low') or 0)),
        ('예상 견적 상한', int(preview['estimate'].get('estimated_high') or 0)),
        ('추천 인원', f"{preview['estimate'].get('recommended_crew')}명"),
        ('추천 차량', f"{preview['estimate'].get('recommended_vehicle_count')}대"),
        ('충돌 분석', preview['schedule_analysis'].get('conflict_level') or ''),
        ('권장 조치', preview['schedule_analysis'].get('recommended_action') or ''),
        ('재방문 고객 여부', '있음' if preview['crm_matches'] else '없음'),
    ]
    ws['A1'] = '이청잘 자동 견적 추출 결과'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    ws['A3'] = '기본 요약'
    ws['A3'].font = bold
    ws['A3'].fill = section_fill
    row_idx = 4
    for label, value in rows:
        ws[f'A{row_idx}'] = label
        ws[f'B{row_idx}'] = value
        ws[f'A{row_idx}'].font = bold
        ws[f'A{row_idx}'].fill = header_fill
        row_idx += 1
    ws[f'A{row_idx+1}'] = 'AI 계산 근거'
    ws[f'A{row_idx+1}'].font = bold
    ws[f'A{row_idx+1}'].fill = section_fill
    for idx, line in enumerate(preview['estimate'].get('explanation_lines') or [], start=row_idx+2):
        ws[f'A{idx}'] = f'근거 {idx-(row_idx+1)}'
        ws[f'B{idx}'] = line
    start = row_idx + 4 + len(preview['estimate'].get('explanation_lines') or [])
    ws[f'A{start}'] = '추천 체크리스트'
    ws[f'A{start}'].font = bold
    ws[f'A{start}'].fill = section_fill
    for idx, item_row in enumerate(preview['recommended_checklist'].get('items') or [], start=start+1):
        ws[f'A{idx}'] = f'체크 {idx-start}'
        ws[f'B{idx}'] = item_row.get('label') or ''
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/api/quote-forms/submit")
def submit_quote_form(payload: QuoteFormSubmitIn, user=Depends(get_optional_user)):
    if not payload.privacy_agreed:
        raise HTTPException(status_code=400, detail='개인정보 수집 및 이용 동의가 필요합니다.')
    requester_name = str(payload.requester_name or '').strip()
    if not requester_name:
        raise HTTPException(status_code=400, detail='고객 성함을 입력해 주세요.')
    contact_phone = str(payload.contact_phone or '').strip()
    if not contact_phone:
        raise HTTPException(status_code=400, detail='견적 받으실 연락처를 입력해 주세요.')
    desired_date = str(payload.desired_date or '').strip()
    now = utcnow()
    form_type = 'storage' if str(payload.form_type or '').strip() == 'storage' else 'same_day'
    summary_title = str(payload.summary_title or '').strip() or f"{'짐보관이사' if form_type == 'storage' else '당일이사'} · {requester_name}"
    payload_json = json.dumps({**(payload.payload or {}), 'privacy_agreed': bool(payload.privacy_agreed)}, ensure_ascii=False)
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO quote_form_submissions(form_type, requester_user_id, requester_name, contact_phone, desired_date, summary_title, status, payload_json, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, 'received', ?, ?, ?)",
            (form_type, user['id'] if user else None, requester_name, contact_phone, desired_date, summary_title, payload_json, now, now),
        )
        return {'ok': True}

@app.get('/api/admin/quote-forms')
def admin_quote_forms(admin=Depends(require_admin_mode_user)):
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, form_type, requester_user_id, requester_name, contact_phone, desired_date, summary_title, status, payload_json, created_at, updated_at FROM quote_form_submissions ORDER BY created_at DESC, id DESC"
        ).fetchall()
    items = []
    for row in rows:
        item = row_to_dict(row)
        item['payload'] = json_loads(item.get('payload_json'), {})
        items.append(item)
    return {'items': items}

@app.get('/api/admin/quote-forms/{submission_id}')
def admin_quote_form_detail(submission_id: int, admin=Depends(require_admin_mode_user)):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, form_type, requester_user_id, requester_name, contact_phone, desired_date, summary_title, status, payload_json, created_at, updated_at FROM quote_form_submissions WHERE id = ?",
            (submission_id,),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail='해당 양식 접수를 찾을 수 없습니다.')
    item = row_to_dict(row)
    item['payload'] = json_loads(item.get('payload_json'), {})
    return {'item': item}


@app.get('/api/admin/quote-forms/{submission_id}/operations-preview')
def admin_quote_form_operations_preview(submission_id: int, admin=Depends(require_admin_mode_user)):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, form_type, requester_user_id, requester_name, contact_phone, desired_date, summary_title, status, payload_json, created_at, updated_at FROM quote_form_submissions WHERE id = ?",
            (submission_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail='해당 양식 접수를 찾을 수 없습니다.')
        item = row_to_dict(row)
        item['payload'] = json_loads(item.get('payload_json'), {})
        preview = _quote_operations_preview(conn, item)
    return {'item_id': submission_id, 'preview': preview}


@app.get('/api/admin/quote-forms/{submission_id}/estimate-excel')
def admin_quote_form_estimate_excel(submission_id: int, admin=Depends(require_admin_mode_user)):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, form_type, requester_user_id, requester_name, contact_phone, desired_date, summary_title, status, payload_json, created_at, updated_at FROM quote_form_submissions WHERE id = ?",
            (submission_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail='해당 양식 접수를 찾을 수 없습니다.')
        item = row_to_dict(row)
        item['payload'] = json_loads(item.get('payload_json'), {})
        preview = _quote_operations_preview(conn, item)
    filename = f"estimate_{submission_id}.xlsx"
    data = _build_quote_estimate_workbook(item, preview)
    return StreamingResponse(io.BytesIO(data), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={filename}'})


@app.get('/api/operations/dashboard')
def operations_dashboard(user=Depends(require_user)):
    with get_conn() as conn:
        today = date.today().isoformat()
        rows_today = conn.execute("SELECT id, customer_name, amount1, amount2, deposit_amount, representative1, representative2, representative3, staff1, staff2, staff3 FROM calendar_events WHERE event_date = ? ORDER BY id DESC", (today,)).fetchall()
        rows_month = conn.execute("SELECT event_date, amount1, amount2, deposit_amount FROM calendar_events WHERE event_date >= ? ORDER BY event_date DESC", ((date.today() - timedelta(days=30)).isoformat(),)).fetchall()
        quote_count = conn.execute("SELECT COUNT(*) AS cnt FROM quote_form_submissions WHERE created_at >= ?", ((datetime.utcnow() - timedelta(days=30)).isoformat(),)).fetchone()['cnt']
        repeat_candidates = conn.execute("SELECT contact_phone, COUNT(*) AS cnt FROM quote_form_submissions GROUP BY contact_phone HAVING COUNT(*) > 1 ORDER BY cnt DESC LIMIT 5").fetchall()
        latest_locations = conn.execute("SELECT COUNT(*) AS cnt FROM vehicle_live_locations WHERE updated_at >= ?", ((datetime.utcnow() - timedelta(hours=1)).isoformat(),)).fetchone()['cnt']
        evidence_count = conn.execute("SELECT COUNT(*) AS cnt FROM work_media_evidence").fetchone()['cnt']
        checklist_count = conn.execute("SELECT COUNT(*) AS cnt FROM work_checklists").fetchone()['cnt']
    def money_sum(rows, *keys):
        total = 0
        for row in rows:
            for key in keys:
                total += _to_int(row[key], 0)
        return total
    staff_names = set()
    for row in rows_today:
        for key in ['representative1','representative2','representative3','staff1','staff2','staff3']:
            name = str(row[key] or '').strip()
            if name:
                staff_names.add(name)
    return {
        'today': {
            'schedule_count': len(rows_today),
            'assigned_people_count': len(staff_names),
            'sales_amount': money_sum(rows_today, 'amount1', 'amount2'),
            'deposit_amount': money_sum(rows_today, 'deposit_amount'),
        },
        'month': {
            'quote_count': int(quote_count or 0),
            'sales_amount': money_sum(rows_month, 'amount1', 'amount2'),
            'deposit_amount': money_sum(rows_month, 'deposit_amount'),
        },
        'operations': {
            'repeat_customer_candidates': [{'contact_phone': row['contact_phone'], 'count': row['cnt']} for row in repeat_candidates],
            'live_vehicle_count': int(latest_locations or 0),
            'evidence_count': int(evidence_count or 0),
            'checklist_count': int(checklist_count or 0),
        },
        'feature_status': [
            {'key': 'quote_ai', 'label': '자동 견적 생성', 'status': '활성'},
            {'key': 'schedule_conflict', 'label': '일정 충돌 분석', 'status': '활성'},
            {'key': 'vehicle_tracking', 'label': '차량 위치 / 자동 상태 체크', 'status': '준비완료'},
            {'key': 'deposit_alert', 'label': '계약금 / 잔금 알림', 'status': '준비완료'},
            {'key': 'crm', 'label': '고객 CRM 누적', 'status': '활성'},
            {'key': 'reviews_ai', 'label': '리뷰 자동 수집 / AI 답변', 'status': '준비완료'},
            {'key': 'dashboard', 'label': '대시보드', 'status': '활성'},
            {'key': 'checklist', 'label': '작업 체크리스트', 'status': '준비완료'},
            {'key': 'evidence', 'label': '사진/영상 증빙', 'status': '준비완료'},
            {'key': 'attendance', 'label': '출퇴근 / 급여 계산', 'status': '준비완료'},
        ],
    }

@app.post("/api/inquiries")
def create_inquiry(payload: InquiryIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO inquiries(user_id, category, title, content, status, created_at) VALUES (?, ?, ?, ?, 'received', ?)",
            (user["id"], payload.category, payload.title, payload.content, utcnow()),
        )
        return {"ok": True}
@app.post("/api/report/{target_user_id}")
def create_report(target_user_id: int, payload: ReportIn, user=Depends(require_user)):
    if target_user_id == user["id"]:
        raise HTTPException(status_code=400, detail="본인을 신고할 수 없습니다.")
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO reports(reporter_id, target_user_id, reason, detail, status, created_at) VALUES (?, ?, ?, ?, 'open', ?)",
            (user["id"], target_user_id, payload.reason, payload.detail, utcnow()),
        )
        return {"ok": True}
@app.get('/api/admin-mode')
def get_admin_mode(admin=Depends(require_admin_mode_user)):
    with get_conn() as conn:
        total_vehicle_count = _get_admin_total_vehicle_count(conn)
        branch_count_override = _get_branch_count_override(conn)
        permission_config = _get_permission_config(conn)
        users = conn.execute(
            """
            SELECT id, email, name, nickname, role, grade, approved, gender, birth_year, region, phone, recovery_email, vehicle_number, branch_no, account_unique_id, group_number, group_number_text, created_at,
                   marital_status, resident_address, business_name, business_number, business_type, business_item, business_address,
                   bank_account, bank_name, mbti, google_email, account_status, permission_codes_json, account_type, branch_code, resident_id, position_title, vehicle_available, show_in_branch_status, show_in_employee_status, show_in_field_employee_status, show_in_hq_status, archived_in_branch_status
            FROM users
            ORDER BY COALESCE(branch_no, 9999), nickname
            """
        ).fetchall()
    user_dicts = [_serialize_admin_user_row(row) for row in users]
    branches = [item for item in user_dicts if item.get('show_in_branch_status')]
    active_branches = [item for item in branches if not item.get('archived_in_branch_status')]
    employees = [item for item in user_dicts if item.get('show_in_field_employee_status') or (item.get('show_in_employee_status') and not item.get('show_in_hq_status'))]
    head_office_staff = [item for item in user_dicts if item.get('show_in_hq_status') or _is_head_office_staff(item)]
    return {
        'config': {'total_vehicle_count': total_vehicle_count, 'branch_count_override': branch_count_override},
        'permission_config': permission_config,
        'branch_count': len(active_branches),
        'branches': branches,
        'employee_count': len(employees),
        'employees': employees,
        'head_office_count': len(head_office_staff),
        'head_office_staff': head_office_staff,
        'accounts': user_dicts,
    }
@app.post('/api/admin-mode/config')
def save_admin_mode_config(payload: AdminModeConfigIn, admin=Depends(require_admin_mode_user)):
    with get_conn() as conn:
        existing_menu_permissions = _get_admin_setting(conn, 'menu_permissions_json', '')
        if int(admin.get('grade') or 6) != 1 and str(payload.menu_permissions_json or '').strip() != str(existing_menu_permissions or '').strip():
            raise HTTPException(status_code=403, detail='메뉴권한은 관리자만 변경할 수 있습니다.')
    settings_to_save = {
        'total_vehicle_count': str(payload.total_vehicle_count or '').strip(),
        'branch_count_override': str(payload.branch_count_override or '').strip(),
        'admin_mode_access_grade': str(payload.admin_mode_access_grade or 2),
        'role_assign_actor_max_grade': str(payload.role_assign_actor_max_grade),
        'role_assign_target_min_grade': str(payload.role_assign_target_min_grade),
        'account_suspend_actor_max_grade': str(payload.account_suspend_actor_max_grade),
        'account_suspend_target_min_grade': str(payload.account_suspend_target_min_grade),
        'signup_approve_actor_max_grade': str(payload.signup_approve_actor_max_grade),
        'signup_approve_target_min_grade': str(payload.signup_approve_target_min_grade),
        'menu_permissions_json': str(payload.menu_permissions_json or '').strip(),
        'menu_locks_json': str(payload.menu_locks_json or '').strip(),
    }
    with get_conn() as conn:
        now = utcnow()
        for key, value in settings_to_save.items():
            conn.execute(
                "INSERT INTO admin_settings(key, value, updated_at) VALUES (?, ?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at",
                (key, value, now),
            )
    return {'ok': True}
@app.put("/api/admin/accounts/{user_id}")
def update_admin_account(user_id: int, payload: AdminAccountUpdateIn, admin=Depends(require_admin_mode_user)):
    if payload.grade not in {1, 2, 3, 4, 5, 6, 7}:
        raise HTTPException(status_code=400, detail='허용되지 않는 권한입니다.')
    with get_conn() as conn:
        existing = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail='계정을 찾을 수 없습니다.')
        target_grade = int(payload.grade or 6)
        if int(admin.get('grade') or 6) != 1 and not _can_manage_grade(admin, target_grade, conn):
            raise HTTPException(status_code=403, detail='해당 권한을 수정할 수 없습니다.')
        approved = int(payload.approved) if payload.approved is not None else int(existing['approved'] if existing['approved'] is not None else 1)
        next_position_title = str((payload.position_title if payload.position_title is not None else (existing['position_title'] if 'position_title' in existing.keys() else '')) or '').strip()
        if not next_position_title and existing['branch_no'] not in (None, '') and int(existing['branch_no']) > 0:
            next_position_title = '호점대표'
        vehicle_available_value = 0 if _is_staff_grade(target_grade) else (1 if payload.vehicle_available else 0)
        conn.execute("UPDATE users SET grade = ?, approved = ?, position_title = ?, vehicle_available = ? WHERE id = ?", (target_grade, approved, next_position_title, vehicle_available_value, user_id))
        _sync_all_day_note_available_vehicle_counts(conn)
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return user_public_dict(row)
@app.post("/api/admin/accounts/bulk")
def update_admin_accounts_bulk(payload: AdminAccountsBulkUpdateIn, admin=Depends(require_admin_mode_user)):
    updated = []
    with get_conn() as conn:
        for item in payload.accounts:
            user_id = int(item.id or 0)
            target_grade = int(item.grade or 6)
            existing = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
            if not existing:
                continue
            if int(admin.get('grade') or 6) != 1 and not _can_manage_grade(admin, target_grade, conn):
                continue
            approved = int(item.approved) if item.approved is not None else int(existing['approved'] if existing['approved'] is not None else 1)
            next_position_title = str((item.position_title if item.position_title is not None else (existing['position_title'] if 'position_title' in existing.keys() else '')) or '').strip()
            if not next_position_title and existing['branch_no'] not in (None, '') and int(existing['branch_no']) > 0:
                next_position_title = '호점대표'
            vehicle_available_value = 0 if _is_staff_grade(target_grade) else (1 if item.vehicle_available else 0)
            conn.execute("UPDATE users SET grade = ?, approved = ?, position_title = ?, vehicle_available = ? WHERE id = ?", (target_grade, approved, next_position_title, vehicle_available_value, user_id))
            row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
            updated.append(user_public_dict(row))
        _sync_all_day_note_available_vehicle_counts(conn)
    return {'ok': True, 'accounts': updated}

@app.post("/api/admin/accounts/switch-type")
def switch_admin_account_type(payload: AdminAccountTypeSwitchIn, admin=Depends(require_admin_or_subadmin)):
    target_type = str(payload.target_type or '').strip().lower()
    if target_type not in {'business', 'employee'}:
        raise HTTPException(status_code=400, detail='전환 유형이 올바르지 않습니다.')
    with get_conn() as conn:
        existing = conn.execute("SELECT * FROM users WHERE id = ?", (int(payload.user_id),)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail='계정을 찾을 수 없습니다.')
        if int(admin.get('grade') or 6) == 2 and int(existing['grade'] or 6) <= 2:
            raise HTTPException(status_code=403, detail='관리자 및 부관리자 계정은 전환할 수 없습니다.')
        current_type = _normalize_account_type(existing)
        if current_type == target_type:
            return {'ok': True, 'user': _serialize_admin_user_row(existing)}
        role_value = 'business' if target_type == 'business' else 'user'
        position_title = str((existing['position_title'] if 'position_title' in existing.keys() else '') or '').strip()
        if target_type == 'business' and not position_title:
            position_title = '호점대표'
        conn.execute("UPDATE users SET role = ?, position_title = ? WHERE id = ?", (role_value, position_title, int(payload.user_id)))
        row = conn.execute("SELECT * FROM users WHERE id = ?", (int(payload.user_id),)).fetchone()
    return {'ok': True, 'user': _serialize_admin_user_row(row)}

@app.post("/api/admin/users/details-bulk")
def update_admin_user_details_bulk(payload: AdminUserDetailsBulkIn, admin=Depends(require_admin_or_subadmin)):
    editable_fields = [
        'group_number', 'group_number_text', 'name', 'nickname', 'account_unique_id', 'position_title', 'gender', 'birth_year', 'region', 'phone', 'recovery_email',
        'vehicle_number', 'branch_no', 'marital_status', 'resident_address',
        'business_name', 'business_number', 'business_type', 'business_item', 'business_address',
        'bank_account', 'bank_name', 'mbti', 'login_id', 'email', 'google_email', 'account_status', 'permission_codes_json', 'account_type', 'branch_code', 'resident_id', 'vehicle_available', 'show_in_branch_status', 'show_in_employee_status', 'show_in_field_employee_status', 'show_in_hq_status', 'archived_in_branch_status',
    ]
    with get_conn() as conn:
        for item in payload.users:
            existing = conn.execute("SELECT * FROM users WHERE id = ?", (item.id,)).fetchone()
            if not existing:
                continue
            data = item.model_dump()
            branch_value = data.get('branch_no')
            if branch_value in ('', None):
                data['branch_no'] = None
            else:
                try:
                    data['branch_no'] = int(branch_value)
                except Exception:
                    data['branch_no'] = None
            try:
                data['birth_year'] = int(data.get('birth_year') or 1995)
            except Exception:
                data['birth_year'] = 1995
            raw_group_number = str(data.get('group_number') or '0')
            cleaned_group_number = ''.join(ch for ch in raw_group_number if ch.isdigit()) or '0'
            data['group_number_text'] = cleaned_group_number
            try:
                data['group_number'] = int(cleaned_group_number)
            except Exception:
                data['group_number'] = 0
            data['position_title'] = str(data.get('position_title') or '')
            current_or_next_grade = int(data.get('grade') or existing['grade'] or 6)
            data['archived_in_branch_status'] = 1 if bool(data.get('archived_in_branch_status', False)) else 0
            data['name'] = str(data.get('name') or '').strip()
            data['nickname'] = str(data.get('nickname') or '').strip()
            data['account_unique_id'] = str(data.get('account_unique_id') or '').strip()
            data['login_id'] = _validate_login_id_value(data.get('login_id') or data.get('email') or existing['login_id'] or existing['email'])
            data['email'] = _normalize_email_value(data.get('email') or data['login_id'] or existing['email'] or existing['login_id'])
            data['recovery_email'] = _normalize_email_value(data.get('recovery_email') or '')
            data['gender'] = _validate_gender_value(data.get('gender') or '', allow_empty=True)
            data = _normalize_account_admin_flags({**row_to_dict(existing), **data, 'grade': current_or_next_grade})
            if not data['name']:
                data['name'] = data['nickname']
            if not data['account_unique_id']:
                data['account_unique_id'] = generate_account_unique_id(conn, data['email'] or existing['email'], item.id)
            dup_login_id = _find_user_by_login_id_ci(conn, data['login_id'], item.id)
            if dup_login_id:
                raise HTTPException(status_code=400, detail=f"{data['login_id']} 아이디는 이미 사용 중입니다.")
            dup_email = _find_user_by_email_ci(conn, data['email'], item.id)
            if dup_email:
                raise HTTPException(status_code=400, detail=f"{data['email']} 이메일은 이미 사용 중입니다.")
            if data['account_unique_id']:
                dup_uid = conn.execute("SELECT id FROM users WHERE account_unique_id = ? AND id != ?", (data['account_unique_id'], item.id)).fetchone()
                if dup_uid:
                    raise HTTPException(status_code=400, detail=f"{data['account_unique_id']} 고유ID값은 이미 사용 중입니다.")
            if not data['position_title'] and data.get('branch_no') not in (None, '') and int(data.get('branch_no') or 0) > 0:
                data['position_title'] = '호점대표'
            update_fields = list(editable_fields)
            if str(data.get('new_password') or '').strip():
                data['password_hash'] = hash_password(str(data.get('new_password') or '').strip())
                update_fields.append('password_hash')
            assignments = ', '.join(f"{field} = ?" for field in update_fields)
            values = [data.get(field) for field in update_fields] + [item.id]
            conn.execute(f"UPDATE users SET {assignments} WHERE id = ?", values)
    return {'ok': True}
@app.post('/api/admin/accounts/create')
def create_admin_account(payload: AdminCreateAccountIn, admin=Depends(require_admin_or_subadmin)):
    if payload.grade not in {1,2,3,4,5,6,7}:
        raise HTTPException(status_code=400, detail='허용되지 않는 권한입니다.')
    if int(admin.get('grade') or 6) == 2 and int(payload.grade or 6) <= 2:
        raise HTTPException(status_code=403, detail='부관리자는 관리자/부관리자 계정을 생성할 수 없습니다.')
    if not str(payload.name or '').strip():
        raise HTTPException(status_code=400, detail='이름을 입력해주세요.')
    if not str(payload.login_id or '').strip():
        raise HTTPException(status_code=400, detail='아이디를 입력해주세요.')
    if not str(payload.password or '').strip():
        raise HTTPException(status_code=400, detail='비밀번호를 입력해주세요.')
    if not str(payload.nickname or '').strip():
        raise HTTPException(status_code=400, detail='닉네임을 입력해주세요.')
    payload_gender = _validate_gender_value(payload.gender, allow_empty=False)
    try:
        with get_conn() as conn:
            login_id = _validate_login_id_value(payload.login_id or payload.email)
            normalized_email = _normalize_email_value(payload.email or login_id)
            exists = _find_user_by_login_id_ci(conn, login_id)
            if exists:
                raise HTTPException(status_code=400, detail='이미 존재하는 아이디입니다.')
            exists_email = _find_user_by_email_ci(conn, normalized_email)
            if exists_email:
                raise HTTPException(status_code=400, detail='이미 존재하는 아이디입니다.')
            generated_unique_id = generate_account_unique_id(conn, normalized_email)
            normalized_new_user = _normalize_account_admin_flags({
                'grade': int(payload.grade),
                'branch_no': payload.branch_no,
                'position_title': str(payload.position_title or '').strip(),
                'vehicle_available': payload.vehicle_available,
                'show_in_hq_status': False,
                'account_status': payload.account_status,
            })
            conn.execute(
                """
                INSERT INTO users(login_id, email, google_email, password_hash, name, nickname, role, grade, approved, account_status, permission_codes_json, account_type, branch_code, gender, birth_year, region, phone, recovery_email, vehicle_number, branch_no, position_title, vehicle_available, account_unique_id, group_number, group_number_text, show_in_branch_status, show_in_employee_status, show_in_field_employee_status, show_in_hq_status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (login_id, normalized_email, _normalize_email_value(payload.google_email), hash_password(payload.password), str(payload.name or '').strip(), str(payload.nickname or '').strip(), normalized_new_user['role'], int(payload.grade), int(bool(payload.approved)), _normalize_account_status_value(payload.account_status, payload.approved, payload.grade), normalized_new_user['permission_codes_json'], normalized_new_user['account_type'], normalized_new_user['branch_code'], payload_gender, payload.birth_year, payload.region, payload.phone, _normalize_email_value(payload.recovery_email), payload.vehicle_number, payload.branch_no if payload.branch_no is not None else (-1 if int(payload.grade or 6) == 4 else None), normalized_new_user['position_title'], normalized_new_user['vehicle_available'], generated_unique_id, int(''.join(ch for ch in str(payload.group_number or '0') if ch.isdigit()) or 0), ''.join(ch for ch in str(payload.group_number or '0') if ch.isdigit()) or '0', normalized_new_user['show_in_branch_status'], normalized_new_user['show_in_employee_status'], normalized_new_user['show_in_field_employee_status'], normalized_new_user['show_in_hq_status'], utcnow()),
            )
            user_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            conn.execute('INSERT INTO preferences(user_id, data) VALUES (?, ?)', (user_id, json.dumps({"groupChatNotifications": True, "directChatNotifications": True, "likeNotifications": True, "theme": "dark"}, ensure_ascii=False)))
            row = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        return {'ok': True, 'user': user_public_dict(row)}
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception('create_admin_account failed: %s', exc)
        raise HTTPException(status_code=500, detail='계정 생성 중 서버 오류가 발생했습니다.')
@app.post('/api/admin/accounts/delete')
def delete_admin_accounts(payload: AdminDeleteAccountsIn, admin=Depends(require_admin_or_subadmin)):
    target_ids = [int(item) for item in (payload.ids or []) if int(item or 0) > 0]
    if not target_ids:
        raise HTTPException(status_code=400, detail='삭제할 계정을 선택해주세요.')
    deleted_ids = []
    with get_conn() as conn:
        for target_id in target_ids:
            if int(admin['id']) == target_id:
                continue
            row = conn.execute("SELECT * FROM users WHERE id = ?", (target_id,)).fetchone()
            if not row:
                continue
            if int(admin.get('grade') or 6) == 2 and int(row['grade'] or 6) <= 2:
                continue
            mark_deleted_imported_account(conn, row['email'])
            conn.execute("DELETE FROM users WHERE id = ?", (target_id,))
            deleted_ids.append(target_id)
    return {'ok': True, 'deleted_ids': deleted_ids}

@app.get("/api/admin/reports")
def admin_reports(admin=Depends(require_admin)):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM reports ORDER BY id DESC").fetchall()
        enriched = []
        for r in rows:
            item = row_to_dict(r)
            item["reporter"] = user_basic(conn, r["reporter_id"])
            item["target"] = user_basic(conn, r["target_user_id"])
            enriched.append(item)
        return enriched
@app.post("/api/admin/reports/{report_id}/close")
def close_report(report_id: int, admin=Depends(require_admin)):
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="신고를 찾을 수 없습니다.")
        conn.execute("UPDATE reports SET status = 'closed', closed_at = ?, closed_by = ? WHERE id = ?", (utcnow(), admin["id"], report_id))
        return {"ok": True}
@app.get("/api/blocked-users")
def blocked_users(user=Depends(require_user)):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT b.*, u.* FROM blocks b
            JOIN users u ON u.id = b.blocked_user_id
            WHERE b.blocker_id = ?
            ORDER BY b.id DESC
            """,
            (user["id"],),
        ).fetchall()
        return [
            {
                "id": r["id"],
                "reason": r["reason"],
                "blocked_user": {
                    "id": r["blocked_user_id"],
                    "nickname": r["nickname"],
                    "email": r["email"],
                    "region": r["region"],
                }
            }
            for r in rows
        ]
@app.post("/api/block/{target_user_id}")
def block_user(target_user_id: int, payload: BlockIn, user=Depends(require_user)):
    if target_user_id == user["id"]:
        raise HTTPException(status_code=400, detail="본인을 차단할 수 없습니다.")
    with get_conn() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO blocks(blocker_id, blocked_user_id, reason, created_at) VALUES (?, ?, ?, ?)",
            (user["id"], target_user_id, payload.reason, utcnow()),
        )
        return {"ok": True}
@app.put("/api/blocked-users/{block_id}/reason")
def update_block_reason(block_id: int, payload: BlockIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute("UPDATE blocks SET reason = ? WHERE id = ? AND blocker_id = ?", (payload.reason, block_id, user["id"]))
        return {"ok": True}
@app.post("/api/unblock/{target_user_id}")
def unblock_user(target_user_id: int, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute("DELETE FROM blocks WHERE blocker_id = ? AND blocked_user_id = ?", (user["id"], target_user_id))
        return {"ok": True}
@app.get("/api/preferences")
def get_preferences(user=Depends(require_user)):
    with get_conn() as conn:
        row = conn.execute("SELECT data FROM preferences WHERE user_id = ?", (user["id"],)).fetchone()
        return json.loads(row["data"]) if row else {}

@app.post("/api/preferences")
def save_preferences(payload: PreferenceIn, user=Depends(require_user)):
    with get_conn() as conn:
        conn.execute("INSERT OR REPLACE INTO preferences(user_id, data) VALUES (?, ?)", (user["id"], json.dumps(payload.data, ensure_ascii=False)))
        return {"ok": True}


def _disposal_compact_text(value: str) -> str:
    raw = str(value or '').strip()
    if not raw:
        return ''
    raw = raw.replace(',', ' ')
    raw = re.sub(r'[^가-힣0-9\s]', ' ', raw)
    raw = re.sub(r'\s+', ' ', raw).strip()
    return raw


_DISPOSAL_SHORT_REGION_NAMES = (
    '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종', '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주'
)


_DISPOSAL_LONG_REGION_NAMES = (
    '서울특별시', '부산광역시', '대구광역시', '인천광역시', '광주광역시', '대전광역시', '울산광역시', '세종특별자치시',
    '경기도', '강원특별자치도', '강원도', '충청북도', '충청남도', '전북특별자치도', '전라북도', '전라남도', '경상북도', '경상남도', '제주특별자치도', '제주도'
)


def _normalize_disposal_place_prefix(value: str) -> str:
    raw = _disposal_compact_text(value)
    if not raw:
        return ''
    raw = re.sub(r'([가-힣])(특별시|광역시|특별자치시|특별자치도|도|시|구|군)', r'\1\2 ', raw)
    raw = re.sub(r'\s+', ' ', raw).strip()

    # Prefer the opening address phrase because disposal jurisdiction matching should rely mostly on the first 2~3 words.
    tokens = [token for token in raw.split(' ') if token]
    if tokens:
        region_token = ''
        if tokens[0] in _DISPOSAL_REGION_ALIAS_MAP:
            region_token = _DISPOSAL_REGION_ALIAS_MAP.get(tokens[0], tokens[0])
        else:
            compact_first = re.sub(r'\s+', '', tokens[0])
            for region_alias in sorted(_DISPOSAL_REGION_ALIAS_MAP.keys(), key=len, reverse=True):
                if compact_first.startswith(region_alias):
                    region_token = _DISPOSAL_REGION_ALIAS_MAP.get(region_alias, region_alias)
                    tail = compact_first[len(region_alias):].strip()
                    if tail:
                        if len(tokens) == 1:
                            tokens = [region_token, tail]
                        else:
                            tokens = [region_token, tail] + tokens[1:]
                    else:
                        tokens[0] = region_token
                    break
        if region_token and len(tokens) >= 2:
            second = tokens[1]
            if re.search(r'(시|구|군)$', second):
                if len(tokens) >= 3 and second.endswith('시') and re.search(r'(구|군)$', tokens[2]):
                    return f"{region_token} {second} {tokens[2]}".strip()
                return f"{region_token} {second}".strip()

    match = re.search(r'((?:[가-힣]+(?:특별시|광역시|특별자치시|특별자치도|도|시))\s+[가-힣0-9]+(?:구|군|시))', raw)
    if match:
        return match.group(1).strip()
    short_match = re.search(r'((?:서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주))\s*([가-힣0-9]+(?:구|군|시)?)', raw)
    if short_match:
        return f"{short_match.group(1)} {short_match.group(2)}".strip()

    compact = re.sub(r'\s+', '', raw)
    for region in sorted(_DISPOSAL_LONG_REGION_NAMES, key=len, reverse=True):
        if compact.startswith(region):
            tail = compact[len(region):]
            district_match = re.match(r'([가-힣0-9]+(?:구|군|시)?)', tail)
            if district_match:
                return f"{region} {district_match.group(1)}".strip()
    for region in sorted(_DISPOSAL_SHORT_REGION_NAMES, key=len, reverse=True):
        if compact.startswith(region):
            tail = compact[len(region):]
            district_match = re.match(r'([가-힣0-9]+(?:구|군|시)?)', tail)
            if district_match:
                return f"{region} {district_match.group(1)}".strip()

    region_aliases = sorted(set(_DISPOSAL_REGION_ALIAS_MAP.keys()) | set(_DISPOSAL_REGION_ALIAS_MAP.values()), key=len, reverse=True)
    found_region = ''
    for region_alias in region_aliases:
        if region_alias and region_alias in compact:
            found_region = _DISPOSAL_REGION_ALIAS_MAP.get(region_alias, region_alias)
            break
    district_candidates: list[str] = []
    for token in re.findall(r'[가-힣0-9]{1,12}(?:구|군|시)?', compact):
        stripped = _strip_disposal_region_suffix(token)
        if not stripped:
            continue
        if token in _DISPOSAL_REGION_ALIAS_MAP or stripped in _DISPOSAL_REGION_ALIAS_MAP:
            continue
        district_candidates.append(token)
        district_candidates.append(stripped)
    if found_region and district_candidates:
        for candidate in district_candidates:
            if candidate and candidate != found_region:
                return f"{found_region} {candidate}".strip()

    tokens = raw.split(' ')
    if len(tokens) >= 2:
        return ' '.join(tokens[:2]).strip()
    return raw


_DISPOSAL_REGION_ALIAS_MAP: dict[str, str] = {
    '서울특별시': '서울',
    '서울시': '서울',
    '서울': '서울',
    '부산광역시': '부산',
    '부산시': '부산',
    '부산': '부산',
    '대구광역시': '대구',
    '대구시': '대구',
    '대구': '대구',
    '인천광역시': '인천',
    '인천시': '인천',
    '인천': '인천',
    '광주광역시': '광주',
    '광주시': '광주',
    '광주': '광주',
    '대전광역시': '대전',
    '대전시': '대전',
    '대전': '대전',
    '울산광역시': '울산',
    '울산시': '울산',
    '울산': '울산',
    '세종특별자치시': '세종',
    '세종시': '세종',
    '세종': '세종',
    '경기도': '경기',
    '경기': '경기',
    '강원특별자치도': '강원',
    '강원도': '강원',
    '강원': '강원',
    '충청북도': '충북',
    '충북': '충북',
    '충청남도': '충남',
    '충남': '충남',
    '전북특별자치도': '전북',
    '전라북도': '전북',
    '전북': '전북',
    '전라남도': '전남',
    '전남': '전남',
    '경상북도': '경북',
    '경북': '경북',
    '경상남도': '경남',
    '경남': '경남',
    '제주특별자치도': '제주',
    '제주도': '제주',
    '제주': '제주',
}


def _strip_disposal_region_suffix(value: str) -> str:
    token = str(value or '').strip()
    if not token:
        return ''
    token = re.sub(r'(특별자치시|특별자치도|특별시|광역시|자치시|자치도)$', '', token)
    token = re.sub(r'(시|구|군)$', '', token)
    return token.strip()


def _disposal_place_search_key(value: str) -> str:
    normalized = _normalize_disposal_place_prefix(value)
    if not normalized:
        return ''
    compact = re.sub(r'\s+', ' ', normalized).strip()
    tokens = [token for token in re.split(r'\s+', compact) if token]
    if not tokens:
        return ''
    region = _DISPOSAL_REGION_ALIAS_MAP.get(tokens[0], _strip_disposal_region_suffix(tokens[0]))
    district = ''
    for token in tokens[1:]:
        stripped = _strip_disposal_region_suffix(token)
        if stripped:
            district = stripped
            break
    if not district and len(tokens) == 1:
        short_match = re.match(r'^(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)([가-힣0-9]+)$', tokens[0])
        if short_match:
            region = short_match.group(1)
            district = _strip_disposal_region_suffix(short_match.group(2))
    return ' '.join([part for part in [region, district] if part]).strip()


def _disposal_place_search_parts(value: str) -> tuple[str, str]:
    search_key = _disposal_place_search_key(value)
    parts = [part for part in search_key.split(' ') if part]
    if not parts:
        return ('', '')
    if len(parts) == 1:
        return (parts[0], '')
    return (parts[0], parts[1])


def _disposal_place_keys_match(left: str, right: str) -> bool:
    left_region, left_district = _disposal_place_search_parts(left)
    right_region, right_district = _disposal_place_search_parts(right)
    if not left_region or not right_region or left_region != right_region:
        return False
    if not left_district or not right_district:
        return left_region == right_region
    return left_district == right_district


def _disposal_location_compact_tokens(value: str) -> set[str]:
    compact = re.sub(r'\s+', '', _disposal_compact_text(value))
    tokens: set[str] = set()
    if not compact:
        return tokens

    for region_alias in set(_DISPOSAL_REGION_ALIAS_MAP.values()):
        if region_alias and region_alias in compact:
            tokens.add(region_alias)

    for match in re.finditer(r'([가-힣0-9]{1,12})(구|군|시)', compact):
        whole = ''.join(match.groups())
        stripped = _strip_disposal_region_suffix(whole)
        if whole:
            tokens.add(whole)
        if stripped:
            tokens.add(stripped)

    for chunk in re.findall(r'[가-힣0-9]{2,12}', compact):
        stripped = _strip_disposal_region_suffix(chunk)
        if stripped:
            tokens.add(stripped)
        tokens.add(chunk)
    return {token for token in tokens if token}


def _disposal_front_weighted_tokens(value: str, limit: int = 3) -> tuple[list[str], str, str]:
    raw = _disposal_compact_text(value)
    if not raw:
        return ([], '', '')
    raw = re.sub(r'[,()/\-]+', ' ', raw)
    raw = re.sub(r'\s+', ' ', raw).strip()
    tokens = [token for token in raw.split(' ') if token]
    front_tokens = tokens[:max(1, limit)]
    front_phrase = ' '.join(front_tokens).strip()
    compact_front = re.sub(r'\s+', '', front_phrase)
    return (front_tokens, front_phrase, compact_front)


def _disposal_row_region_district_tokens(place_prefix: str, district_name: str = '') -> tuple[str, set[str]]:
    row_region, row_district = _disposal_place_search_parts(place_prefix)
    aliases: set[str] = set()
    if row_district:
        aliases.add(row_district)
    district_raw = str(district_name or '').strip()
    if district_raw:
        aliases.add(district_raw)
        stripped = _strip_disposal_region_suffix(district_raw)
        if stripped:
            aliases.add(stripped)
    compact_place = re.sub(r'\s+', '', _disposal_compact_text(place_prefix))
    for alias in list(aliases):
        compact_alias = re.sub(r'\s+', '', _disposal_compact_text(alias))
        if compact_alias:
            aliases.add(compact_alias)
    if compact_place:
        aliases.add(compact_place)
    return row_region, {alias for alias in aliases if alias}


def _disposal_similarity_score(location: str, place_prefix: str, district_name: str = '') -> int:
    input_region, input_district = _disposal_place_search_parts(location)
    row_region, row_aliases = _disposal_row_region_district_tokens(place_prefix, district_name)
    district_alias = _strip_disposal_region_suffix(district_name)
    row_district = ''
    for candidate in row_aliases:
        stripped = _strip_disposal_region_suffix(candidate)
        if stripped and (not row_district or len(stripped) < len(row_district)):
            row_district = stripped

    compact_location = re.sub(r'\s+', '', _disposal_compact_text(location))
    location_tokens = _disposal_location_compact_tokens(location)
    front_tokens, front_phrase, compact_front = _disposal_front_weighted_tokens(location, limit=3)
    front_location_tokens = _disposal_location_compact_tokens(front_phrase)

    if not row_region:
        return -1
    if input_region:
        if input_region != row_region:
            return -1
    elif row_region not in location_tokens and row_region not in compact_location:
        return -1

    score = 100
    if input_district and row_district:
        if input_district == row_district:
            score += 100
        elif row_district.startswith(input_district) or input_district.startswith(row_district):
            score += 85
        elif input_district in row_district or row_district in input_district:
            score += 60
        else:
            district_matched = False
            for alias in row_aliases:
                stripped_alias = _strip_disposal_region_suffix(alias)
                compact_alias = re.sub(r'\s+', '', _disposal_compact_text(alias))
                if stripped_alias and stripped_alias in location_tokens:
                    district_matched = True
                    score += 55
                    break
                if compact_alias and compact_alias in compact_location:
                    district_matched = True
                    score += 50
                    break
            if not district_matched:
                return -1
    elif input_district or row_district:
        district_matched = False
        search_targets = {input_district, row_district, district_alias}
        for alias in row_aliases.union(search_targets):
            stripped_alias = _strip_disposal_region_suffix(alias)
            compact_alias = re.sub(r'\s+', '', _disposal_compact_text(alias))
            if stripped_alias and stripped_alias in location_tokens:
                district_matched = True
                score += 35
                break
            if compact_alias and compact_alias in compact_location:
                district_matched = True
                score += 30
                break
        if not district_matched and input_district:
            return -1

    normalized_input = _normalize_disposal_place_prefix(location)
    normalized_row = _normalize_disposal_place_prefix(place_prefix)
    if normalized_input and normalized_row:
        if normalized_input == normalized_row:
            score += 50
        elif normalized_input in normalized_row or normalized_row in normalized_input:
            score += 20

    compact_row = re.sub(r'\s+', '', _disposal_compact_text(place_prefix))
    compact_district_name = re.sub(r'\s+', '', _disposal_compact_text(district_name))
    if compact_input := compact_location:
        if compact_row:
            if compact_input == compact_row:
                score += 30
            elif compact_input in compact_row or compact_row in compact_input:
                score += 15
        if compact_district_name:
            if compact_district_name in compact_input:
                score += 35
            elif any(token and token in compact_input for token in {district_alias, row_district}):
                score += 25

    # Give heavy weight to the opening address phrase so the first 3 words drive ~80% of the match quality.
    # Example: "경기 고양시 일산동구 ..." should strongly favor rows like "경기 고양시" over tokens found later in the sentence.
    if compact_front and compact_row:
        front_score = 0
        if compact_front == compact_row:
            front_score = 220
        elif compact_row.startswith(compact_front) or compact_front.startswith(compact_row):
            front_score = 185
        elif compact_row in compact_front or compact_front in compact_row:
            front_score = 160
        else:
            front_region, front_district = _disposal_place_search_parts(front_phrase)
            if front_region and front_region == row_region:
                front_score += 95
            if front_district and row_district:
                if front_district == row_district:
                    front_score += 95
                elif row_district.startswith(front_district) or front_district.startswith(row_district):
                    front_score += 80
                elif front_district in row_district or row_district in front_district:
                    front_score += 60
            for alias in row_aliases.union({row_region, row_district, district_alias}):
                stripped_alias = _strip_disposal_region_suffix(alias)
                compact_alias = re.sub(r'\s+', '', _disposal_compact_text(alias))
                if stripped_alias and stripped_alias in front_location_tokens:
                    front_score = max(front_score, 140)
                if compact_alias and compact_alias in compact_front:
                    front_score = max(front_score, 130)
        score += front_score

    # Small penalty if the row only matches the full sentence weakly but does not appear near the front.
    if compact_front and compact_row and compact_row not in compact_front and compact_front not in compact_row:
        if row_region not in front_location_tokens and not any(
            (_strip_disposal_region_suffix(alias) and _strip_disposal_region_suffix(alias) in front_location_tokens)
            or (re.sub(r'\s+', '', _disposal_compact_text(alias)) and re.sub(r'\s+', '', _disposal_compact_text(alias)) in compact_front)
            for alias in row_aliases
        ):
            score -= 35

    return score


def _disposal_find_best_jurisdiction_match(rows, location: str):
    best_row = None
    best_score = -1
    for row in rows:
        score = _disposal_similarity_score(location, str(row['place_prefix'] or ''), str(row['district_name'] or ''))
        if score > best_score:
            best_row = row
            best_score = score
    return best_row if best_score >= 0 else None


def _disposal_jurisdiction_row_to_dict(row) -> dict[str, Any]:
    return {
        'id': int(row['id']),
        'category': str(row['category'] or '기본'),
        'place_prefix': str(row['place_prefix'] or ''),
        'district_name': str(row['district_name'] or ''),
        'report_link': str(row['report_link'] or ''),
        'created_at': str(row['created_at'] or ''),
        'updated_at': str(row['updated_at'] or ''),
    }


def _disposal_jurisdiction_order_by_sql() -> str:
    if DB_ENGINE == 'postgresql':
        return 'LOWER(place_prefix), id DESC'
    return 'place_prefix COLLATE NOCASE, id DESC'


@app.get('/api/disposal/jurisdictions')
def list_disposal_jurisdictions(q: str = Query(default=''), user=Depends(require_admin_or_subadmin)):
    keyword = str(q or '').strip()
    order_by_sql = _disposal_jurisdiction_order_by_sql()
    with get_conn() as conn:
        if keyword:
            like = f'%{keyword}%'
            rows = conn.execute(
                f"""
                SELECT id, category, place_prefix, district_name, report_link, created_at, updated_at
                FROM disposal_jurisdiction_mappings
                WHERE category LIKE ? OR place_prefix LIKE ? OR district_name LIKE ? OR report_link LIKE ?
                ORDER BY {order_by_sql}
                """,
                (like, like, like, like),
            ).fetchall()
        else:
            rows = conn.execute(
                f"""
                SELECT id, category, place_prefix, district_name, report_link, created_at, updated_at
                FROM disposal_jurisdiction_mappings
                ORDER BY {order_by_sql}
                """
            ).fetchall()
    return {'rows': [_disposal_jurisdiction_row_to_dict(row) for row in rows]}


@app.post('/api/disposal/jurisdictions/bulk-save')
def bulk_save_disposal_jurisdictions(payload: DisposalJurisdictionBulkSaveIn, user=Depends(require_admin_or_subadmin)):
    rows = payload.rows or []
    saved_rows: list[dict[str, Any]] = []
    normalized_rows: list[tuple[DisposalJurisdictionRowIn, str, str, str, str]] = []
    seen_place_prefixes: set[str] = set()
    for item in rows:
        place_prefix = _normalize_disposal_place_prefix(item.place_prefix)
        district_name = str(item.district_name or '').strip()
        if not place_prefix or not district_name:
            continue
        if place_prefix in seen_place_prefixes:
            raise HTTPException(status_code=400, detail=f'중복된 폐기장소가 있습니다: {place_prefix}')
        seen_place_prefixes.add(place_prefix)
        category = str(item.category or '기본').strip() or '기본'
        report_link = str(item.report_link or '').strip()
        normalized_rows.append((item, place_prefix, district_name, category, report_link))

    with get_conn() as conn:
        for item, place_prefix, district_name, category, report_link in normalized_rows:
            now = utcnow()
            duplicate_row = conn.execute(
                'SELECT id FROM disposal_jurisdiction_mappings WHERE place_prefix = ? AND id <> ?' if item.id else 'SELECT id FROM disposal_jurisdiction_mappings WHERE place_prefix = ?',
                (place_prefix, int(item.id)) if item.id else (place_prefix,),
            ).fetchone()
            if duplicate_row:
                raise HTTPException(status_code=400, detail=f'이미 등록된 폐기장소입니다: {place_prefix}')

            existing = conn.execute(
                'SELECT id FROM disposal_jurisdiction_mappings WHERE id = ?' if item.id else 'SELECT id FROM disposal_jurisdiction_mappings WHERE place_prefix = ?',
                (int(item.id),) if item.id else (place_prefix,),
            ).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE disposal_jurisdiction_mappings
                    SET category = ?, place_prefix = ?, district_name = ?, report_link = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (category, place_prefix, district_name, report_link, now, int(existing['id'])),
                )
                saved_id = int(existing['id'])
            else:
                if DB_ENGINE == 'postgresql':
                    row = conn.execute(
                        """
                        INSERT INTO disposal_jurisdiction_mappings(category, place_prefix, district_name, report_link, created_by, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        RETURNING id
                        """,
                        (category, place_prefix, district_name, report_link, int(user['id']), now, now),
                    ).fetchone()
                    if not row:
                        raise HTTPException(status_code=500, detail='관할구역 데이터를 저장하지 못했습니다.')
                    saved_id = int(row['id'])
                else:
                    cursor = conn.execute(
                        """
                        INSERT INTO disposal_jurisdiction_mappings(category, place_prefix, district_name, report_link, created_by, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (category, place_prefix, district_name, report_link, int(user['id']), now, now),
                    )
                    saved_id = int(cursor.lastrowid)
            row = conn.execute(
                """
                SELECT id, category, place_prefix, district_name, report_link, created_at, updated_at
                FROM disposal_jurisdiction_mappings
                WHERE id = ?
                """,
                (saved_id,),
            ).fetchone()
            if row:
                saved_rows.append(_disposal_jurisdiction_row_to_dict(row))
    return {'ok': True, 'rows': saved_rows}


@app.post('/api/disposal/jurisdictions/delete')
def delete_disposal_jurisdictions(payload: DisposalJurisdictionBulkSaveIn, user=Depends(require_admin_or_subadmin)):
    ids = [int(item.id) for item in (payload.rows or []) if item.id]
    if not ids:
        return {'ok': True, 'deleted_ids': []}
    placeholders = ','.join('?' for _ in ids)
    with get_conn() as conn:
        conn.execute(f'DELETE FROM disposal_jurisdiction_mappings WHERE id IN ({placeholders})', ids)
    return {'ok': True, 'deleted_ids': ids}


@app.get('/api/disposal/jurisdictions/resolve', response_model=DisposalJurisdictionResolveOut)
def resolve_disposal_jurisdiction(location: str = Query(default=''), user=Depends(require_admin_or_subadmin)):
    normalized = _normalize_disposal_place_prefix(location)
    search_key = _disposal_place_search_key(location)
    if not normalized:
        return DisposalJurisdictionResolveOut(matched=False)
    with get_conn() as conn:
        exact = conn.execute(
            """
            SELECT place_prefix, district_name, report_link
            FROM disposal_jurisdiction_mappings
            WHERE place_prefix = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (normalized,),
        ).fetchone()
        if exact:
            return DisposalJurisdictionResolveOut(
                matched=True,
                place_prefix=str(exact['place_prefix'] or ''),
                district_name=str(exact['district_name'] or ''),
                report_link=str(exact['report_link'] or ''),
            )
        rows = conn.execute(
            """
            SELECT place_prefix, district_name, report_link, id
            FROM disposal_jurisdiction_mappings
            ORDER BY id DESC
            """
        ).fetchall()
    if search_key:
        for row in rows:
            row_place_prefix = str(row['place_prefix'] or '')
            if _disposal_place_keys_match(row_place_prefix, search_key):
                return DisposalJurisdictionResolveOut(
                    matched=True,
                    place_prefix=row_place_prefix,
                    district_name=str(row['district_name'] or ''),
                    report_link=str(row['report_link'] or ''),
                )
    best_row = _disposal_find_best_jurisdiction_match(rows, location)
    if best_row:
        return DisposalJurisdictionResolveOut(
            matched=True,
            place_prefix=str(best_row['place_prefix'] or ''),
            district_name=str(best_row['district_name'] or ''),
            report_link=str(best_row['report_link'] or ''),
        )
    return DisposalJurisdictionResolveOut(matched=False, place_prefix=normalized)


@app.get('/api/storage-status/state')
def storage_status_state_api(user=Depends(require_user)):
    with get_conn() as conn:
        state = get_storage_status_state(conn)
    return {'state': state}


@app.post('/api/storage-status/state')
def save_storage_status_state_api(payload: StorageStatusStateIn, user=Depends(require_user)):
    rows = [item.model_dump() for item in (payload.rows or [])]
    with get_conn() as conn:
        state = replace_storage_status_rows(conn, rows)
    return {'ok': True, 'state': state}


@app.get('/api/warehouse/state')
def warehouse_state_api(user=Depends(require_user)):
    with get_conn() as conn:
        state = get_warehouse_state(conn)
    return {'state': state}


@app.post('/api/warehouse/cell')
def warehouse_cell_api(payload: WarehouseCellUpdateIn, user=Depends(require_user)):
    with get_conn() as conn:
        try:
            state = update_warehouse_cell(conn, payload.sheet_name, int(payload.row), int(payload.col), payload.value)
        except ValueError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
    return {'ok': True, 'state': state}


@app.post('/api/warehouse/layout')
def warehouse_layout_api(payload: WarehouseLayoutUpdateIn, user=Depends(require_user)):
    with get_conn() as conn:
        try:
            state = update_warehouse_layout(conn, payload.sheet_name, col_widths=payload.col_widths, row_heights=payload.row_heights)
        except ValueError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
    return {'ok': True, 'state': state}


def _get_materials_table_scale_key() -> str:
    return 'materials_table_scale_json'


def _normalize_materials_table_device(device: str | None) -> str:
    normalized = str(device or 'desktop').strip().lower()
    return 'mobile' if normalized == 'mobile' else 'desktop'


def _get_materials_table_layout_key(device: str | None) -> str:
    return f"materials_table_layout_json:{_normalize_materials_table_device(device)}"


@app.get('/api/policies-content')
def get_policies_content(user=Depends(require_user)):
    try:
        with get_conn() as conn:
            contents = _load_policy_contents(conn)
    except Exception as exc:
        logger.exception('failed to load policy contents: %s', exc)
        contents = dict(POLICY_CONTENT_DEFAULTS)
    return {'contents': contents}

@app.post('/api/policies-content')
def save_policies_content(payload: PreferenceIn, user=Depends(require_admin_or_subadmin)):
    data = payload.data or {}
    category = str(data.get('category') or '').strip()
    target = str(data.get('target') or '').strip()
    content = str(data.get('content') or '')
    normalized = f"{category}:{target}".strip(':')
    if normalized not in POLICY_CONTENT_DEFAULTS:
        raise HTTPException(status_code=400, detail='허용되지 않는 규정 항목입니다.')
    try:
        with get_conn() as conn:
            contents = _save_policy_content(conn, normalized, content)
    except Exception as exc:
        logger.exception('failed to save policy content: %s', exc)
        raise HTTPException(status_code=500, detail='규정 저장 중 오류가 발생했습니다.') from exc
    return {'ok': True, 'contents': contents}

@app.get('/api/materials/table-scale')
def get_materials_table_scale(user=Depends(require_user)):
    with get_conn() as conn:
        raw = _get_admin_setting(conn, _get_materials_table_scale_key(), '{}')
    try:
        scales = json.loads(raw or '{}')
        if not isinstance(scales, dict):
            scales = {}
    except Exception:
        scales = {}
    return {'scales': scales}

@app.post('/api/materials/table-scale')
def save_materials_table_scale(payload: PreferenceIn, user=Depends(require_admin_or_subadmin)):
    scales = (payload.data or {}).get('scales', {})
    if not isinstance(scales, dict):
        raise HTTPException(status_code=400, detail='표 배율 데이터 형식이 올바르지 않습니다.')
    normalized = {}
    for key, value in scales.items():
        try:
            parsed = int(round(float(value)))
        except Exception:
            continue
        normalized[str(key)] = max(80, min(140, parsed))
    with get_conn() as conn:
        now = utcnow()
        conn.execute(
            "INSERT INTO admin_settings(key, value, updated_at) VALUES (?, ?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at",
            (_get_materials_table_scale_key(), json.dumps(normalized, ensure_ascii=False), now),
        )
    return {'ok': True, 'scales': normalized}

@app.get('/api/materials/table-layout')
def get_materials_table_layout(device: str = 'desktop', user=Depends(require_user)):
    normalized = _normalize_materials_table_device(device)
    with get_conn() as conn:
        raw = _get_admin_setting(conn, _get_materials_table_layout_key(normalized), '{}')
    try:
        layouts = json.loads(raw or '{}')
        if not isinstance(layouts, dict):
            layouts = {}
    except Exception:
        layouts = {}
    return {'device': normalized, 'layouts': layouts}

@app.post('/api/materials/table-layout')
def save_materials_table_layout(payload: PreferenceIn, user=Depends(require_admin_or_subadmin)):
    device = _normalize_materials_table_device((payload.data or {}).get('device', 'desktop'))
    layouts = (payload.data or {}).get('layouts', {})
    if not isinstance(layouts, dict):
        raise HTTPException(status_code=400, detail='표 레이아웃 데이터 형식이 올바르지 않습니다.')
    with get_conn() as conn:
        now = utcnow()
        conn.execute(
            "INSERT INTO admin_settings(key, value, updated_at) VALUES (?, ?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at",
            (_get_materials_table_layout_key(device), json.dumps(layouts, ensure_ascii=False), now),
        )
    return {'ok': True, 'device': device, 'layouts': layouts}
@app.get('/api/materials/overview')
def get_materials_overview(user=Depends(require_user)):
    with get_conn() as conn:
        return _material_overview_payload(conn, user)

@app.options('/api/materials/purchase-requests')
def options_material_purchase_requests():
    return {'ok': True}

@app.post('/api/materials/purchase-requests')
def create_material_purchase_request(payload: MaterialPurchaseCreateIn, user=Depends(require_user)):
    _require_materials_scope(user, 'sales')
    if _is_employee_restricted_user(user):
        raise HTTPException(status_code=403, detail='직원 계정은 자재를 구매할 수 없습니다.')
    valid_items = [item for item in payload.items if int(item.quantity or 0) > 0]
    if not valid_items:
        raise HTTPException(status_code=400, detail='구매 개수를 1개 이상 입력해 주세요.')
    now = utcnow()
    with get_conn() as conn:
        products = {int(row['id']): row_to_dict(row) for row in conn.execute("SELECT * FROM material_products WHERE COALESCE(is_active, 1) = 1").fetchall()}
        total_amount = 0
        request_items = []
        for item in valid_items:
            product = products.get(int(item.product_id))
            if not product:
                continue
            qty = max(0, int(item.quantity or 0))
            unit_price = int(product.get('unit_price') or 0)
            line_total = qty * unit_price
            total_amount += line_total
            request_items.append((int(product['id']), qty, unit_price, line_total, str(item.memo or '').strip()))
        if not request_items:
            raise HTTPException(status_code=400, detail='유효한 자재 항목이 없습니다.')
        branch_label = _material_branch_label_from_user(user)
        requester_display_name = _material_requester_display_name_from_user(user)
        requester_name = ' '.join(part for part in [branch_label, requester_display_name] if part).strip()
        if not requester_name:
            requester_name = requester_display_name or '구매신청자'
        request_note = str(payload.request_note or '').strip()
        requester_unique_id = str(user.get('account_unique_id') or user.get('login_id') or user.get('email') or '').strip().lower()
        if DB_ENGINE == 'postgresql':
            inserted_row = conn.execute(
                '''
                INSERT INTO material_purchase_requests(user_id, requester_name, requester_unique_id, request_note, total_amount, status, payment_confirmed, created_at, settled_at, share_snapshot_json)
                VALUES (?, ?, ?, ?, ?, 'pending', 0, ?, '', '')
                RETURNING id
                ''',
                (user['id'], requester_name, requester_unique_id, request_note, total_amount, now),
            ).fetchone()
            request_id = int((inserted_row or {}).get('id') or 0)
        else:
            conn.execute(
                '''
                INSERT INTO material_purchase_requests(user_id, requester_name, requester_unique_id, request_note, total_amount, status, payment_confirmed, created_at, settled_at, share_snapshot_json)
                VALUES (?, ?, ?, ?, ?, 'pending', 0, ?, '', '')
                ''',
                (user['id'], requester_name, requester_unique_id, request_note, total_amount, now),
            )
            request_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0] or 0)
        if request_id <= 0:
            raise HTTPException(status_code=500, detail='구매신청 번호 생성에 실패했습니다.')
        for product_id, qty, unit_price, line_total, memo in request_items:
            conn.execute(
                "INSERT INTO material_purchase_request_items(request_id, product_id, quantity, unit_price, line_total, memo) VALUES (?, ?, ?, ?, ?, ?)",
                (request_id, product_id, qty, unit_price, line_total, memo),
            )
        row = conn.execute("SELECT * FROM material_purchase_requests WHERE id = ?", (request_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=500, detail='구매신청 저장 후 조회에 실패했습니다.')
        detail = _material_request_detail(conn, row_to_dict(row))
        detail['requester_display_name'] = requester_display_name
        detail['requester_user_name'] = str(user.get('name') or '').strip()
        detail['requester_nickname'] = str(user.get('nickname') or '').strip()
        detail['requester_login_id'] = str(user.get('login_id') or '').strip()
        detail['requester_email'] = str(user.get('email') or '').strip()
        detail['requester_google_email'] = str(user.get('google_email') or '').strip()
        detail['requester_branch_no'] = user.get('branch_no')
        detail['requester_branch_code'] = str(user.get('branch_code') or '').strip()
        detail['requester_branch_label'] = branch_label
        detail['requester_account_unique_id'] = requester_unique_id
        detail['requester_unique_id'] = requester_unique_id
        logger.info(
            'Material purchase request saved request_id=%s user_id=%s requester_unique_id=%s status=%s total_amount=%s',
            request_id,
            user.get('id'),
            requester_unique_id,
            detail.get('status'),
            detail.get('total_amount'),
        )
        try:
            _notify_material_purchase_request(conn, user, detail)
        except Exception:
            logger.exception('Failed to notify material purchase request admins request_id=%s user_id=%s', request_id, user.get('id'))
        return {'ok': True, 'request': detail}

@app.post('/api/materials/purchase-requests/settle')
def settle_material_purchase_requests(payload: MaterialSettlementProcessIn, user=Depends(require_user)):
    _require_materials_scope(user, 'requesters')
    request_ids = sorted({int(item) for item in payload.request_ids if int(item or 0) > 0})
    if not request_ids:
        raise HTTPException(status_code=400, detail='결산등록할 구매신청자를 선택해 주세요.')
    now = utcnow()
    placeholders = ','.join('?' for _ in request_ids)
    with get_conn() as conn:
        rows = [
            row_to_dict(row)
            for row in conn.execute(
                f"SELECT * FROM material_purchase_requests WHERE id IN ({placeholders}) ORDER BY created_at, id",
                tuple(request_ids),
            ).fetchall()
        ]
        if not rows:
            raise HTTPException(status_code=404, detail='결산 대상 신청서를 찾을 수 없습니다.')
        settled_rows = []
        for row in rows:
            if row.get('status') == 'settled':
                settled_rows.append(_material_request_detail(conn, row))
                continue
            detail = _material_request_detail(conn, row)
            share_text = _material_share_text([detail])
            _adjust_material_stock_by_request_ids(conn, [int(row['id'])], -1, now)
            conn.execute(
                "UPDATE material_purchase_requests SET status = 'settled', payment_confirmed = 1, settled_at = ?, settled_by_user_id = ?, share_snapshot_json = ? WHERE id = ?",
                (now, user['id'], share_text, row['id']),
            )
            updated = conn.execute("SELECT * FROM material_purchase_requests WHERE id = ?", (row['id'],)).fetchone()
            settled_rows.append(_material_request_detail(conn, row_to_dict(updated)))
        return {'ok': True, 'settled_requests': settled_rows, 'share_text': _material_share_text(settled_rows)}


@app.post('/api/materials/purchase-requests/unsettle')
def unsettle_material_purchase_requests(payload: MaterialSettlementProcessIn, user=Depends(require_user)):
    _require_materials_scope(user, 'requesters')
    request_ids = sorted({int(item) for item in payload.request_ids if int(item or 0) > 0})
    if not request_ids:
        raise HTTPException(status_code=400, detail='결산취소할 신청건을 선택해 주세요.')
    placeholders = ','.join('?' for _ in request_ids)
    with get_conn() as conn:
        rows = [
            row_to_dict(row)
            for row in conn.execute(
                f"SELECT * FROM material_purchase_requests WHERE id IN ({placeholders}) ORDER BY created_at DESC, id DESC",
                tuple(request_ids),
            ).fetchall()
        ]
        if not rows:
            raise HTTPException(status_code=404, detail='결산취소 대상 신청건을 찾을 수 없습니다.')
        updated_rows = []
        for row in rows:
            if str(row.get('status') or '') != 'settled':
                updated_rows.append(_material_request_detail(conn, row))
                continue
            _adjust_material_stock_by_request_ids(conn, [int(row['id'])], 1, now)
            conn.execute(
                "UPDATE material_purchase_requests SET status = 'pending', payment_confirmed = 0, settled_at = '', settled_by_user_id = NULL, share_snapshot_json = '' WHERE id = ?",
                (row['id'],),
            )
            updated = conn.execute("SELECT * FROM material_purchase_requests WHERE id = ?", (row['id'],)).fetchone()
            updated_rows.append(_material_request_detail(conn, row_to_dict(updated)))
        return {'ok': True, 'requests': updated_rows}


@app.post('/api/materials/purchase-requests/reject')
def reject_material_purchase_requests(payload: MaterialSettlementProcessIn, user=Depends(require_user)):
    _require_materials_scope(user, 'requesters')
    request_ids = sorted({int(item) for item in payload.request_ids if int(item or 0) > 0})
    if not request_ids:
        raise HTTPException(status_code=400, detail='결산반려할 구매신청자를 선택해 주세요.')
    placeholders = ','.join('?' for _ in request_ids)
    with get_conn() as conn:
        rows = [
            row_to_dict(row)
            for row in conn.execute(
                f"SELECT * FROM material_purchase_requests WHERE id IN ({placeholders}) ORDER BY created_at, id",
                tuple(request_ids),
            ).fetchall()
        ]
        if not rows:
            raise HTTPException(status_code=404, detail='반려 대상 신청서를 찾을 수 없습니다.')
        rejected_rows = []
        for row in rows:
            if str(row.get('status') or '') == 'settled':
                raise HTTPException(status_code=400, detail='이미 결산완료된 신청건은 반려할 수 없습니다.')
            if str(row.get('status') or '') == 'rejected':
                rejected_rows.append(_material_request_detail(conn, row))
                continue
            conn.execute(
                "UPDATE material_purchase_requests SET status = 'rejected', payment_confirmed = 0, settled_at = '', settled_by_user_id = NULL, share_snapshot_json = '' WHERE id = ?",
                (row['id'],),
            )
            updated = conn.execute("SELECT * FROM material_purchase_requests WHERE id = ?", (row['id'],)).fetchone()
            rejected_rows.append(_material_request_detail(conn, row_to_dict(updated)))
        return {'ok': True, 'rejected_requests': rejected_rows}


@app.put('/api/materials/purchase-requests')
def update_material_purchase_requests(payload: MaterialRequestUpdateIn, user=Depends(require_user)):
    _require_materials_scope(user, 'sales')
    request_ids = sorted({int(item) for item in payload.request_ids if int(item or 0) > 0})
    if not request_ids:
        raise HTTPException(status_code=400, detail='수정/취소할 신청건을 선택해 주세요.')
    quantity_map = {int(row.product_id): max(0, int(row.quantity or 0)) for row in payload.rows}
    if not quantity_map:
        raise HTTPException(status_code=400, detail='수정할 구매수량 정보가 없습니다.')
    now = utcnow()
    placeholders = ','.join('?' for _ in request_ids)
    with get_conn() as conn:
        rows = [row_to_dict(row) for row in conn.execute(
            f"SELECT * FROM material_purchase_requests WHERE id IN ({placeholders}) AND user_id = ? ORDER BY created_at DESC, id DESC",
            tuple(request_ids) + (user['id'],),
        ).fetchall()]
        if not rows:
            raise HTTPException(status_code=404, detail='수정 가능한 신청건을 찾을 수 없습니다.')
        blocked = [row for row in rows if str(row.get('status') or '') == 'settled']
        if blocked:
            raise HTTPException(status_code=400, detail='결산완료된 신청건은 수정 또는 취소할 수 없습니다.')
        valid_product_ids = {int(row['id']) for row in conn.execute("SELECT id FROM material_products WHERE COALESCE(is_active, 1) = 1").fetchall()}
        updated_requests = []
        for request_row in rows:
            items = [row_to_dict(row) for row in conn.execute(
                "SELECT * FROM material_purchase_request_items WHERE request_id = ? ORDER BY id",
                (request_row['id'],),
            ).fetchall()]
            total_amount = 0
            for item in items:
                product_id = int(item.get('product_id') or 0)
                if product_id not in valid_product_ids:
                    continue
                if product_id in quantity_map:
                    qty = max(0, int(quantity_map[product_id]))
                    line_total = qty * int(item.get('unit_price') or 0)
                    conn.execute(
                        "UPDATE material_purchase_request_items SET quantity = ?, line_total = ?, memo = ? WHERE id = ?",
                        (qty, line_total, '취소접수' if qty == 0 else '', int(item['id'])),
                    )
                    total_amount += line_total
                else:
                    total_amount += int(item.get('line_total') or 0)
            conn.execute(
                "UPDATE material_purchase_requests SET total_amount = ?, request_note = ?, created_at = ? WHERE id = ?",
                (total_amount, str(request_row.get('request_note') or ''), now, int(request_row['id'])),
            )
            updated = conn.execute("SELECT * FROM material_purchase_requests WHERE id = ?", (int(request_row['id']),)).fetchone()
            updated_requests.append(_material_request_detail(conn, row_to_dict(updated)))
        return {'ok': True, 'requests': updated_requests}

@app.get('/api/admin/materials/purchase-requests')
def admin_list_material_purchase_requests(user_id: int = 0, status: str = '', start_date: str = '', end_date: str = '', admin=Depends(require_admin_mode_user)):
    allowed_statuses = {'pending', 'settled', 'rejected'}
    normalized_status = str(status or '').strip().lower()
    with get_conn() as conn:
        query = "SELECT * FROM material_purchase_requests WHERE 1=1"
        params = []
        if int(user_id or 0) > 0:
            query += " AND user_id = ?"
            params.append(int(user_id))
        if normalized_status and normalized_status in allowed_statuses:
            query += " AND status = ?"
            params.append(normalized_status)
        if str(start_date or '').strip():
            query += " AND substr(created_at, 1, 10) >= ?"
            params.append(str(start_date).strip())
        if str(end_date or '').strip():
            query += " AND substr(created_at, 1, 10) <= ?"
            params.append(str(end_date).strip())
        query += " ORDER BY created_at DESC, id DESC LIMIT 500"
        rows = [
            _material_request_detail(conn, row_to_dict(row))
            for row in conn.execute(query, tuple(params)).fetchall()
        ]
    return {'requests': rows}

@app.post('/api/admin/materials/purchase-requests/delete')
def admin_delete_material_purchase_requests(payload: MaterialRequestDeleteIn, admin=Depends(require_admin_mode_user)):
    request_ids = sorted({int(item) for item in payload.request_ids if int(item or 0) > 0})
    if not request_ids:
        raise HTTPException(status_code=400, detail='삭제할 신청현황을 선택해 주세요.')
    placeholders = ','.join('?' for _ in request_ids)
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT id FROM material_purchase_requests WHERE id IN ({placeholders})",
            tuple(request_ids),
        ).fetchall()
        if not rows:
            raise HTTPException(status_code=404, detail='삭제할 신청현황을 찾을 수 없습니다.')
        valid_ids = [int(row['id']) for row in rows]
        settled_ids = [
            int(row['id'])
            for row in conn.execute(
                f"SELECT id FROM material_purchase_requests WHERE id IN ({placeholders}) AND status = 'settled'",
                tuple(request_ids),
            ).fetchall()
        ]
        if settled_ids:
            _adjust_material_stock_by_request_ids(conn, settled_ids, 1, utcnow())
        valid_placeholders = ','.join('?' for _ in valid_ids)
        conn.execute(
            f"DELETE FROM material_purchase_request_items WHERE request_id IN ({valid_placeholders})",
            tuple(valid_ids),
        )
        conn.execute(
            f"DELETE FROM material_purchase_requests WHERE id IN ({valid_placeholders})",
            tuple(valid_ids),
        )
    return {'ok': True, 'deleted_ids': valid_ids}

@app.post('/api/materials/inventory')
def save_material_inventory(payload: MaterialInventorySaveIn, user=Depends(require_user)):
    _require_materials_scope(user, 'inventory_manage')
    target_date = datetime.now().date().isoformat()
    now = utcnow()
    with get_conn() as conn:
        valid_product_ids = {int(row['id']) for row in conn.execute("SELECT id FROM material_products WHERE COALESCE(is_active, 1) = 1").fetchall()}
        for row in payload.rows:
            product_id = int(row.product_id or 0)
            if product_id not in valid_product_ids:
                continue
            incoming_qty = max(0, int(row.incoming_qty or 0))
            outgoing_qty = max(0, int(getattr(row, 'outgoing_qty', 0) or 0))
            note = str(row.note or '').strip()
            conn.execute(
                '''
                INSERT INTO material_inventory_daily(inventory_date, product_id, incoming_qty, note, outgoing_qty, is_closed, closed_at, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, 0, '', ?, ?)
                ON CONFLICT(inventory_date, product_id) DO UPDATE SET incoming_qty = excluded.incoming_qty, outgoing_qty = excluded.outgoing_qty, note = excluded.note, updated_at = excluded.updated_at
                ''',
                (target_date, product_id, incoming_qty, note, outgoing_qty, now, now),
            )
        return {'ok': True, 'inventory_rows': _material_today_inventory_rows(conn, target_date)}


def _normalize_material_entry_date(value: str) -> str:
    raw = str(value or '').strip()[:10]
    if not raw:
        return datetime.now().date().isoformat()
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date().isoformat()
    except ValueError:
        raise HTTPException(status_code=400, detail='입고입력일 형식이 올바르지 않습니다.')

@app.post('/api/materials/incoming')
def save_material_incoming(payload: MaterialIncomingSaveIn, user=Depends(require_user)):
    _require_materials_scope(user, 'inventory_manage')
    entry_date = _normalize_material_entry_date(payload.entry_date)
    now = utcnow()
    rows = []
    for row in payload.rows:
        incoming_qty = max(0, int(row.incoming_qty or 0))
        outgoing_qty = max(0, int(getattr(row, 'outgoing_qty', 0) or 0))
        product_id = int(row.product_id or 0)
        note = str(getattr(row, 'note', '') or '').strip()
        if product_id > 0 and (incoming_qty > 0 or outgoing_qty > 0 or note):
            rows.append({'product_id': product_id, 'incoming_qty': incoming_qty, 'outgoing_qty': outgoing_qty, 'note': note})
    if not rows:
        raise HTTPException(status_code=400, detail='입고수량 또는 출고수량을 1개 이상 입력해 주세요.')
    try:
        with get_conn() as conn:
            product_map = {int(r['id']): row_to_dict(r) for r in conn.execute("SELECT * FROM material_products WHERE COALESCE(is_active, 1) = 1").fetchall()}
            valid_rows = [row for row in rows if row['product_id'] in product_map]
            if not valid_rows:
                raise HTTPException(status_code=400, detail='유효한 입출고 품목이 없습니다.')
            existing_rows = {int(r['product_id']): row_to_dict(r) for r in conn.execute("SELECT * FROM material_inventory_daily WHERE inventory_date = ?", (entry_date,)).fetchall()}
            force_apply = bool(getattr(payload, 'force_apply', False))
            for row in valid_rows:
                product = product_map[row['product_id']]
                existing = existing_rows.get(row['product_id'], {})
                if force_apply:
                    delta = int(row['incoming_qty']) - int(row['outgoing_qty'])
                else:
                    prev_incoming = int(existing.get('incoming_qty') or 0)
                    prev_outgoing = int(existing.get('outgoing_qty') or 0)
                    delta = int(row['incoming_qty']) - prev_incoming - (int(row['outgoing_qty']) - prev_outgoing)
                next_stock = max(0, int(product.get('current_stock') or 0) + delta)
                conn.execute(
                    "UPDATE material_products SET current_stock = ?, updated_at = ? WHERE id = ?",
                    (next_stock, now, row['product_id']),
                )
                if force_apply:
                    continue
                conn.execute(
                    '''
                    INSERT INTO material_inventory_daily(inventory_date, product_id, incoming_qty, note, outgoing_qty, is_closed, closed_at, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, 0, '', ?, ?)
                    ON CONFLICT(inventory_date, product_id) DO UPDATE SET incoming_qty = excluded.incoming_qty, outgoing_qty = excluded.outgoing_qty, note = excluded.note, updated_at = excluded.updated_at
                    ''',
                    (entry_date, row['product_id'], int(row['incoming_qty']), row['note'], int(row['outgoing_qty']), now, now),
                )
            return {'ok': True, 'inventory_rows': _material_today_inventory_rows(conn, datetime.now().date().isoformat())}
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception('save_material_incoming failed: %s', exc)
        raise HTTPException(status_code=500, detail='자재입고 처리 중 서버 오류가 발생했습니다.')

@app.post('/api/materials/inventory/close')
def close_material_inventory(user=Depends(require_admin_or_subadmin)):
    target_date = datetime.now().date().isoformat()
    now = utcnow()
    with get_conn() as conn:
        rows = _material_today_inventory_rows(conn, target_date)
        if any(row.get('is_closed') for row in rows):
            raise HTTPException(status_code=400, detail='오늘 재고 결산은 이미 처리되었습니다.')
        for row in rows:
            conn.execute(
                "UPDATE material_products SET current_stock = ?, updated_at = ? WHERE id = ?",
                (int(row.get('expected_stock') or 0), now, int(row['product_id'])),
            )
            conn.execute(
                '''
                INSERT INTO material_inventory_daily(inventory_date, product_id, incoming_qty, note, outgoing_qty, is_closed, closed_at, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
                ON CONFLICT(inventory_date, product_id) DO UPDATE SET outgoing_qty = excluded.outgoing_qty, is_closed = 1, closed_at = excluded.closed_at, updated_at = excluded.updated_at
                ''',
                (target_date, int(row['product_id']), int(row.get('incoming_qty') or 0), str(row.get('note') or ''), int(row.get('outgoing_qty') or 0), now, now, now),
            )
        return {'ok': True, 'inventory_rows': _material_today_inventory_rows(conn, target_date)}

FRONTEND_DIST_DIR = (Path(__file__).resolve().parents[1] / "static").resolve()
FALLBACK_FRONTEND_DIST_DIR = (Path(__file__).resolve().parents[2] / "frontend" / "dist").resolve()
if not FRONTEND_DIST_DIR.exists() and FALLBACK_FRONTEND_DIST_DIR.exists():
    FRONTEND_DIST_DIR = FALLBACK_FRONTEND_DIST_DIR
def _frontend_file(path: str) -> Optional[Path]:
    candidate = (FRONTEND_DIST_DIR / path).resolve()
    if candidate.is_file() and str(candidate).startswith(str(FRONTEND_DIST_DIR)):
        return candidate
    if Path(path).suffix:
        return None
    return FRONTEND_DIST_DIR / "index.html"


def _frontend_response(file_path: Path) -> FileResponse:
    response = FileResponse(file_path)
    suffix = file_path.suffix.lower()
    name = file_path.name.lower()
    is_asset = '/assets/' in file_path.as_posix() or file_path.parent.name == 'assets'
    if name == 'index.html':
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    elif is_asset and suffix in {'.js', '.css', '.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg', '.ico', '.woff', '.woff2'}:
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    elif suffix in {'.json', '.webmanifest'}:
        response.headers['Cache-Control'] = 'public, max-age=3600, stale-while-revalidate=86400'
    else:
        response.headers['Cache-Control'] = 'public, max-age=600, stale-while-revalidate=3600'
    return response
@app.get("/.well-known/appspecific/com.chrome.devtools.json", include_in_schema=False)
def serve_chrome_devtools_probe():
    return {}


@app.get("/privacy-policy", include_in_schema=False)
def serve_privacy_policy():
    file_path = (Path(__file__).resolve().parents[1] / "static" / "legal" / "privacy-policy.html").resolve()
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="privacy-policy.html 파일이 없습니다.")
    return FileResponse(file_path)


@app.get("/account-deletion", include_in_schema=False)
def serve_account_deletion():
    file_path = (Path(__file__).resolve().parents[1] / "static" / "legal" / "account-deletion.html").resolve()
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="account-deletion.html 파일이 없습니다.")
    return FileResponse(file_path)


@app.get("/uploads/{file_path:path}", include_in_schema=False)
def serve_local_upload(file_path: str):
    requested = (settings.upload_root / file_path).resolve()
    if not requested.exists() or not str(requested).startswith(str(settings.upload_root.resolve())):
        raise HTTPException(status_code=404, detail="Not Found")
    return _frontend_response(requested)


@app.get("/", include_in_schema=False)
def serve_root():
    index_file = FRONTEND_DIST_DIR / "index.html"
    if index_file.exists():
        return _frontend_response(index_file)
    return {
        "message": "이청잘 API 서버 실행중",
        "docs": "/docs",
        "health": "/api/health",
        "frontend_built": False,
    }


@app.get("/{full_path:path}", include_in_schema=False)
def serve_frontend(full_path: str):
    if full_path.startswith("api/") or full_path in {"docs", "openapi.json", "redoc"}:
        raise HTTPException(status_code=404, detail="Not Found")
    index_file = FRONTEND_DIST_DIR / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=404, detail="프런트엔드 빌드본이 없습니다. frontend에서 npm run build를 먼저 실행하세요.")
    requested = _frontend_file(full_path) if full_path else index_file
    if requested is None or not requested.exists():
        raise HTTPException(status_code=404, detail="Not Found")
    return _frontend_response(requested)
